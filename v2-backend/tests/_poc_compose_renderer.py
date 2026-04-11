"""POC: compose-from-scratch xlsx renderer.

Goal: prove that we can fill any template (even complex ones with summary
merges and shifted formulas) WITHOUT calling openpyxl's `insert_rows` /
`delete_rows`. The principle is "compose, never mutate":

  1. Walk the template top-to-bottom row by row
  2. Copy each row to a NEW workbook, transforming as needed:
     - Header rows: copy as-is
     - Product rows: replicated N times from a single template row
     - Summary rows: copied with row offset shifted, formulas adjusted
  3. Re-create merges in the new workbook at the new positions

This avoids EVERY openpyxl row-manipulation footgun (the current
fill_template fights with) because we never call insert_rows. Merges in the
new workbook are created explicitly at the right positions, so there's no
"openpyxl forgot to shift" problem to fix.

Run:
    cd v2-backend
    PYTHONPATH=. ./venv/bin/python tests/_poc_compose_renderer.py
"""
from __future__ import annotations

import copy
import io
import re
import sys
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import column_index_from_string, get_column_letter


# ──────────────────────────────────────────────────────────────────────────
# THE POC RENDERER
# ──────────────────────────────────────────────────────────────────────────


def compose_render(
    template_bytes: bytes,
    zone_config: dict[str, Any],
    order_data: dict[str, Any],
    supplier_id: str | int,
) -> bytes:
    """Render a filled xlsx by composing from a template, NEVER mutating it.

    No insert_rows, no delete_rows, no merge cleanup. Every cell in the
    output is placed explicitly at its final position.
    """
    sid = str(supplier_id)
    products = order_data["suppliers"][sid]["products"]
    n_products = len(products)

    # ── Load source template (read-only reference) ──
    src_wb = load_workbook(io.BytesIO(template_bytes))
    src_ws = src_wb.active

    # ── Create blank destination workbook ──
    dst_wb = Workbook()
    dst_ws = dst_wb.active
    dst_ws.title = src_ws.title

    # ── Zone definitions ──
    prod_start = zone_config["zones"]["product_data"]["start"]
    prod_end = zone_config["zones"]["product_data"]["end"]
    summ_start = zone_config["zones"]["summary"]["start"]
    summ_end = zone_config["zones"]["summary"]["end"]
    template_prod_rows = prod_end - prod_start + 1
    row_delta = n_products - template_prod_rows  # how much summary shifts down

    # Build flat order context (used by per-row order-level fields like ship_name
    # in flat-table templates, AND by header_fields like B2 → ship_name)
    supplier_data = order_data.get("suppliers", {}).get(sid, {})
    order_context = {
        "ship_name": order_data.get("ship_name", ""),
        "delivery_date": order_data.get("delivery_date", ""),
        "order_date": order_data.get("order_date", ""),
        "destination_port": order_data.get("destination_port", ""),
        "voyage": order_data.get("voyage", ""),
        "supplier_name": supplier_data.get("supplier_name", ""),
        "po_number": order_data.get("po_number", ""),
        "currency": order_data.get("currency", ""),
    }

    # ── Phase 1: copy rows above the product zone (header) ──
    for src_row in range(1, prod_start):
        _copy_row_cells(src_ws, dst_ws, src_row, src_row)

    # ── Phase 1b: apply header_fields (placeholder substitution) ──
    # zone_config["header_fields"] maps cell refs like "B2" to data paths like
    # "ship_name" or "suppliers.{sid}.supplier_name". We resolve each path and
    # write the value into the target cell. This OVERWRITES the template's
    # placeholder values with real order data.
    for cell_ref, data_path in zone_config.get("header_fields", {}).items():
        col_letter = "".join(c for c in cell_ref if c.isalpha())
        row_num = int("".join(c for c in cell_ref if c.isdigit()))
        col_idx = column_index_from_string(col_letter)
        if row_num >= prod_start:
            continue  # not in header zone — skip
        value = _resolve_data_path(order_data, data_path, sid)
        if value is not None:
            cell = dst_ws.cell(row=row_num, column=col_idx)
            if not isinstance(cell, MergedCell):
                cell.value = value
                cell.number_format = _format_for_value(cell.number_format, value)

    # ── Pre-scan: does this order contain ANY fractional numeric value? ──
    # If so, every integer-only format in the product zone (including formula
    # cells whose value we can't see at render time, and the summary
    # formulas downstream) gets promoted to `#,##0.##`. This makes display
    # precision consistent across the whole filled workbook regardless of
    # which cells happen to be integer or fractional.
    any_fractional = False
    for product in products:
        for v in product.values():
            if isinstance(v, float) and v != int(v):
                any_fractional = True
                break
        if any_fractional:
            break

    # ── Phase 2: emit N product rows from the template's first product row ──
    template_row = prod_start
    col_map = zone_config.get("product_columns", {})
    formula_map = zone_config.get("product_row_formulas", {})
    po_number = order_data.get("po_number", "")
    currency = order_data.get("currency", "")

    for i, product in enumerate(products):
        dst_row = prod_start + i
        # Copy the template product row's cell formats
        _copy_row_cells(src_ws, dst_ws, template_row, dst_row)
        # Override values with product data (with order_context fallback)
        for col_letter, field_name in col_map.items():
            col_idx = column_index_from_string(col_letter)
            value = _resolve_product_field(
                field_name, product, i, po_number, currency, order_context
            )
            cell = dst_ws.cell(row=dst_row, column=col_idx)
            cell.value = value
            cell.number_format = _format_for_value(
                cell.number_format, value, force_decimal=any_fractional
            )
        # Apply per-row formulas — formula cells don't have a value at render
        # time, so we promote unconditionally if any product data is fractional
        for col_letter, formula_tpl in formula_map.items():
            col_idx = column_index_from_string(col_letter)
            cell = dst_ws.cell(row=dst_row, column=col_idx)
            cell.value = formula_tpl.replace("{row}", str(dst_row))
            if any_fractional:
                cell.number_format = _promote_int_format(cell.number_format)

    # ── Phase 3: copy rows below the product zone (summary + footer) ──
    # Each row gets shifted by row_delta. Formulas in summary cells are
    # NOT auto-adjusted from the template — instead we re-emit them from
    # `summary_formulas` in zone_config (the explicit declarative source
    # of truth that fill_template uses too). This avoids fragile regex-
    # based formula parsing.
    summary_formula_cells: dict[str, dict] = {
        sf["cell"]: sf for sf in zone_config.get("summary_formulas", [])
    }
    new_prod_end = prod_start + n_products - 1
    new_summ_start = prod_end + 1 + row_delta
    new_summ_end = new_summ_start + (summ_end - summ_start)
    stale_summary_columns = zone_config.get("stale_columns_in_summary", [])

    for src_row in range(prod_end + 1, src_ws.max_row + 1):
        dst_row = src_row + row_delta
        _copy_row_cells(src_ws, dst_ws, src_row, dst_row)

    # ── Phase 3a: clear stale columns inside the summary zone ──
    # The original template may have placeholder values (whitespace, leftover
    # numbers) in summary columns that fill_template clears via
    # stale_columns_in_summary. Mirror that here.
    for row in range(new_summ_start, new_summ_end + 1):
        for col_letter in stale_summary_columns:
            col_idx = column_index_from_string(col_letter)
            cell = dst_ws.cell(row=row, column=col_idx)
            if not isinstance(cell, MergedCell):
                cell.value = None

    # ── Phase 3b: restore summary static values at new positions ──
    # zone_config["summary_static_values"] declares which cells in the summary
    # zone hold fixed labels (like "Sub Total", "Tax", "JPY"). The clear above
    # may have wiped them; we re-emit them now at their shifted positions.
    for cell_ref, value in zone_config.get("summary_static_values", {}).items():
        col_letter = "".join(c for c in cell_ref if c.isalpha())
        orig_row = int("".join(c for c in cell_ref if c.isdigit()))
        offset = orig_row - summ_start
        new_row = new_summ_start + offset
        col_idx = column_index_from_string(col_letter)
        cell = dst_ws.cell(row=new_row, column=col_idx)
        if not isinstance(cell, MergedCell):
            cell.value = value

    # ── Phase 3b: emit summary formulas at the new positions ──
    # We compute new positions by walking summary_formulas in declared order.
    # Each formula entry in zone_config knows its original cell; we shift it
    # by row_delta and emit a freshly-built formula with the right range.
    formula_cell_refs: dict[str, str] = {}  # role → "F10" etc.
    for sf in zone_config.get("summary_formulas", []):
        orig_cell = sf["cell"]
        col_letter = "".join(c for c in orig_cell if c.isalpha())
        orig_row = int("".join(c for c in orig_cell if c.isdigit()))
        new_row = orig_row + row_delta
        col_idx = column_index_from_string(col_letter)

        if sf["type"] == "product_sum":
            # SUM over the new product range
            formula = f"=SUM({col_letter}{prod_start}:{col_letter}{new_prod_end})"
            formula_cell_refs["sum_cell"] = f"{col_letter}{new_row}"
        elif sf["type"] == "relative":
            # Substitute placeholders like {sum_cell}, {tax_cell}
            formula = sf["formula_template"]
            for key, ref in formula_cell_refs.items():
                formula = formula.replace(f"{{{key}}}", ref)
            label = sf.get("label", "").lower()
            if "tax" in label:
                formula_cell_refs["tax_cell"] = f"{col_letter}{new_row}"
            if "grand" in label or "total" in label:
                formula_cell_refs["grand_total_cell"] = f"{col_letter}{new_row}"
        else:
            continue

        cell = dst_ws.cell(row=new_row, column=col_idx)
        cell.value = formula
        if any_fractional:
            cell.number_format = _promote_int_format(cell.number_format)

    # ── Phase 3c: external cross-references ─────────────────────
    # zone_config["external_refs"] declares header (or other) cells whose
    # formulas reference summary cells (typically grand_total) by name.
    # Template authors use these to display the grand total in the header
    # for visual emphasis (e.g. "Total: ¥xxx" at the top of the inquiry).
    # After row resizing, the underlying grand total cell has moved, so we
    # must rewrite these formulas with the NEW grand_total_cell address.
    for ext in zone_config.get("external_refs", []):
        cell_ref = ext["cell"]
        col_letter = "".join(c for c in cell_ref if c.isalpha())
        row_num = int("".join(c for c in cell_ref if c.isdigit()))
        col_idx = column_index_from_string(col_letter)
        formula = ext.get("formula_template", "")
        for key, ref in formula_cell_refs.items():
            formula = formula.replace(f"{{{key}}}", ref)
        cell = dst_ws.cell(row=row_num, column=col_idx)
        if not isinstance(cell, MergedCell):
            cell.value = formula
            if any_fractional:
                cell.number_format = _promote_int_format(cell.number_format)

    # ── Phase 4: copy column widths (visual fidelity) ──
    for col_letter, dim in src_ws.column_dimensions.items():
        if dim.width:
            dst_ws.column_dimensions[col_letter].width = dim.width

    # ── Phase 5: re-create merges, adjusting positions ──
    for merged in list(src_ws.merged_cells.ranges):
        if merged.max_row < prod_start:
            # Header merge: same position
            dst_ws.merge_cells(str(merged))

        elif merged.min_row >= prod_start and merged.max_row <= prod_end:
            # Product-zone merge: replicate per product row
            # (e.g. F:G merged for "description" on every product row)
            row_offset_in_zone = merged.min_row - prod_start
            for i in range(n_products):
                dst_row = prod_start + i + row_offset_in_zone
                start_letter = get_column_letter(merged.min_col)
                end_letter = get_column_letter(merged.max_col)
                try:
                    dst_ws.merge_cells(f"{start_letter}{dst_row}:{end_letter}{dst_row}")
                except Exception:
                    pass

        elif merged.min_row > prod_end:
            # Summary or footer merge: shift down by row_delta
            new_min = merged.min_row + row_delta
            new_max = merged.max_row + row_delta
            start_letter = get_column_letter(merged.min_col)
            end_letter = get_column_letter(merged.max_col)
            try:
                dst_ws.merge_cells(f"{start_letter}{new_min}:{end_letter}{new_max}")
            except Exception:
                pass
        # else: merge straddles product zone boundary — rare, skip with warning
        else:
            print(f"  ⚠ skipped boundary-straddling merge: {merged}")

    # ── Save ──
    buf = io.BytesIO()
    dst_wb.save(buf)
    return buf.getvalue()


# ──────────────────────────────────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────────────────────────────────


def _copy_row_cells(
    src_ws,
    dst_ws,
    src_row: int,
    dst_row: int,
    *,
    adjust_formula_rows_by: int = 0,
    product_zone_old: tuple[int, int] | None = None,
    product_zone_new: tuple[int, int] | None = None,
) -> None:
    """Copy all cells in src_row to dst_row, including style + value.

    If adjust_formula_rows_by != 0, formulas referencing rows >= product_zone_old[0]
    are rewritten to point at the corresponding new row positions.
    """
    max_col = src_ws.max_column or 1
    for col_idx in range(1, max_col + 1):
        src_cell = src_ws.cell(row=src_row, column=col_idx)
        if isinstance(src_cell, MergedCell):
            continue
        dst_cell = dst_ws.cell(row=dst_row, column=col_idx)
        # Value (with formula adjustment if needed)
        value = src_cell.value
        if (
            isinstance(value, str)
            and value.startswith("=")
            and adjust_formula_rows_by != 0
            and product_zone_old
            and product_zone_new
        ):
            value = _adjust_formula_row_refs(
                value,
                product_zone_old=product_zone_old,
                product_zone_new=product_zone_new,
                shift=adjust_formula_rows_by,
            )
        dst_cell.value = value
        # Style (if present)
        if src_cell.has_style:
            dst_cell.font = copy.copy(src_cell.font)
            dst_cell.fill = copy.copy(src_cell.fill)
            dst_cell.border = copy.copy(src_cell.border)
            dst_cell.alignment = copy.copy(src_cell.alignment)
            dst_cell.number_format = src_cell.number_format
            dst_cell.protection = copy.copy(src_cell.protection)


_CELL_REF_RE = re.compile(r"(\$?[A-Z]{1,3})(\$?)(\d+)")


# Integer-only number formats that should be promoted to optional-decimal
# when the data is fractional. `#,##0.##` is the universal "show up to 2
# decimals only when needed" format — backward-compatible with integer data.
_INTEGER_ONLY_FORMATS = {
    "#,##0",
    "0",
    "General",
    "#,##0;-#,##0",
    "#,##0_ ;-#,##0_ ",
}


def _promote_int_format(template_fmt: str) -> str:
    """Force-promote an integer-only format to optional-decimal.

    Used unconditionally when we know the surrounding order has fractional
    data — even if THIS particular cell happens to be integer (e.g. an
    amount column whose value is a formula we can't pre-evaluate, or a
    quantity that happens to be a round number).
    """
    if template_fmt in _INTEGER_ONLY_FORMATS:
        return "#,##0.##"
    return template_fmt


def _format_for_value(template_fmt: str, value: Any, force_decimal: bool = False) -> str:
    """Promote integer-only number formats to optional-decimal when needed.

    The template author may have specified `#,##0` (integer with thousand
    separator) assuming the data would always be integers — typical for
    JPY-denominated templates. When that template is reused with data
    containing decimals (USD/AUD/EUR prices like 1.29), the integer format
    silently truncates the display, hiding real precision from the user.

    The fix is permissive widening: if the value has a fractional part AND
    the template format is integer-only, promote it to `#,##0.##` (which
    shows up to 2 decimals only when needed).

    `force_decimal=True` promotes regardless of this cell's individual value
    — used when the surrounding order has fractional data anywhere, so that
    visual precision is consistent across all rows and columns.

    String values (e.g. "KG", "JPY", supplier name) NEVER get their format
    promoted — strings ignore number_format in Excel anyway, and changing
    them would just pollute the workbook with meaningless format strings.
    A formula string (starts with "=") is treated as numeric because Excel
    will evaluate it to a number at open time.
    """
    if isinstance(value, str) and not value.startswith("="):
        return template_fmt
    if force_decimal:
        return _promote_int_format(template_fmt)
    if not isinstance(value, (int, float)):
        return template_fmt
    if value == int(value):
        # integer data — any format the template chose is fine
        return template_fmt
    # value has a fractional part
    return _promote_int_format(template_fmt)


def _adjust_formula_row_refs(
    formula: str,
    product_zone_old: tuple[int, int],
    product_zone_new: tuple[int, int],
    shift: int,
) -> str:
    """Rewrite row numbers in a formula to account for product zone resizing.

    Rules:
      - References to rows inside the old product zone (e.g. =SUM(D5:D15))
        are remapped to span the NEW product zone (e.g. =SUM(D5:D77) for 73 prods)
      - References to rows AFTER the product zone are shifted down by `shift`
      - References to rows BEFORE the product zone are unchanged
    """
    old_start, old_end = product_zone_old
    new_start, new_end = product_zone_new

    def replace_ref(match: re.Match) -> str:
        col, dollar, row_str = match.groups()
        row = int(row_str)
        if row < old_start:
            return match.group(0)  # before product zone — no change
        if row > old_end:
            return f"{col}{dollar}{row + shift}"  # after — shift
        # Inside the old product zone — remap to new zone end
        if row == old_start:
            return f"{col}{dollar}{new_start}"
        if row == old_end:
            return f"{col}{dollar}{new_end}"
        # Other rows inside zone — proportional remap (rare)
        ratio = (row - old_start) / max(1, (old_end - old_start))
        new_row = new_start + round(ratio * (new_end - new_start))
        return f"{col}{dollar}{new_row}"

    return _CELL_REF_RE.sub(replace_ref, formula)


def _resolve_product_field(
    field_name: str,
    product: dict,
    index: int,
    po_number: str,
    currency: str,
    order_context: dict | None = None,
) -> Any:
    """Resolve a product cell value, falling back to order-level context.

    Mirrors services/template_engine.py::_resolve_product_field so that flat-
    table templates (which repeat ship_name / supplier_name / delivery_date
    on every product row) work correctly.
    """
    if field_name in ("line_number", "__line_number__"):
        return index + 1
    if field_name in ("po_number", "__po_number__"):
        return po_number
    if field_name in ("currency", "__currency__"):
        return product.get("currency") or currency

    if field_name == "product_code":
        return product.get("product_code", "")
    if field_name in ("product_name", "product_name_en"):
        return product.get("product_name", "")
    if field_name == "product_name_jp":
        return product.get("product_name_jp", "")
    if field_name == "description":
        return product.get("pack_size", "")
    if field_name == "quantity":
        return product.get("quantity")
    if field_name == "unit_price":
        return product.get("unit_price")
    if field_name == "unit":
        return product.get("unit", "CT")

    # Try product dict first
    val = product.get(field_name)
    if val is not None and val != "":
        return val

    # Fallback to order-level context for repeated fields
    if order_context:
        val = order_context.get(field_name)
        if val is not None:
            return val

    return ""


def _resolve_data_path(order_data: dict, path: str, sid: str) -> Any:
    """Resolve a dotted data path like 'suppliers.{sid}.supplier_name'.

    Mirrors services/template_engine.py::_resolve_path. Used by header_fields
    to find what value should go into a header cell.
    """
    if not path:
        return None
    expanded = path.replace("{sid}", sid)
    cursor: Any = order_data
    for part in expanded.split("."):
        if isinstance(cursor, dict):
            cursor = cursor.get(part)
        else:
            return None
        if cursor is None:
            return None
    return cursor


# ──────────────────────────────────────────────────────────────────────────
# Test fixtures (mirrors the painful real-world case)
# ──────────────────────────────────────────────────────────────────────────


def _build_japanese_style_template() -> tuple[bytes, dict]:
    """A small template that mirrors what fill_template currently fails on:

      - Header section with merged title (rows 1-3)
      - Product zone with column headers (row 4) + 1 product row (row 5)
      - Summary section with merges + label cells + formulas (rows 6-9)

    The product row count (1) is intentionally tiny so we can fill 70+
    products and watch summary correctly shift to the new position.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Inquiry Sheet"

    # ── Header ──
    ws.merge_cells("A1:F1")
    ws["A1"] = "PURCHASE INQUIRY"
    ws["A2"] = "Ship"
    ws["B2"] = "PLACEHOLDER_SHIP"
    ws["A3"] = "Vendor"
    ws["B3"] = "PLACEHOLDER_VENDOR"

    # ── Product header row ──
    ws["A4"] = "No."
    ws["B4"] = "Product Name"
    ws["D4"] = "Qty"
    ws["E4"] = "Price"
    ws["F4"] = "Subtotal"
    # Description spans B:C
    ws.merge_cells("B5:C5")

    # ── One product row template ──
    ws["A5"] = 1
    ws["B5"] = "TEMPLATE_PRODUCT"
    ws["D5"] = 0
    ws["E5"] = 0
    ws["F5"] = "=D5*E5"

    # ── Summary section ──
    # row 6: Subtotal label (merged D:E) + sum formula in F
    ws["A6"] = "Sub Total"
    ws.merge_cells("D6:E6")
    ws["D6"] = "Total:"
    ws["F6"] = "=SUM(F5:F5)"
    # row 7: Tax (merged label)
    ws["A7"] = "Tax"
    ws.merge_cells("D7:E7")
    ws["D7"] = "Tax (10%):"
    ws["F7"] = "=F6*0.1"
    # row 8: Grand Total (merged label)
    ws["A8"] = "Grand Total"
    ws.merge_cells("D8:E8")
    ws["D8"] = "Grand Total:"
    ws["F8"] = "=F6+F7"
    # row 9: signature line (merged across)
    ws.merge_cells("A9:F9")
    ws["A9"] = "Authorized signature: ____________"

    buf = io.BytesIO()
    wb.save(buf)
    template_bytes = buf.getvalue()

    zone_config = {
        "zones": {
            "product_data": {"start": 5, "end": 5},
            "summary": {"start": 6, "end": 9},
        },
        "header_fields": {
            "B2": "ship_name",
            "B3": "suppliers.{sid}.supplier_name",
        },
        "product_columns": {
            "A": "line_number",
            "B": "product_name",
            "D": "quantity",
            "E": "unit_price",
        },
        "product_row_formulas": {"F": "=D{row}*E{row}"},
        "summary_formulas": [
            {"cell": "F6", "type": "product_sum", "label": "Sub Total"},
            {"cell": "F7", "type": "relative", "label": "Tax", "formula_template": "={sum_cell}*0.1"},
            {"cell": "F8", "type": "relative", "label": "Grand Total", "formula_template": "={sum_cell}+{tax_cell}"},
        ],
        "summary_static_values": {
            "A6": "Sub Total", "A7": "Tax", "A8": "Grand Total",
        },
        "stale_columns_in_summary": [],
        "external_refs": [],
    }
    return template_bytes, zone_config


def _build_order_data(n_products: int) -> dict:
    return {
        "po_number": "PO-COMPOSE-001",
        "ship_name": "SILVER NOVA",
        "currency": "USD",
        "suppliers": {
            "1": {
                "supplier_name": "Ocean Foods Co.",
                "products": [
                    {
                        "product_code": f"SKU-{i:03d}",
                        "product_name": f"Test Product {i+1}",
                        "quantity": (i % 9) + 1,
                        "unit_price": round(1.5 + i * 0.3, 2),
                    }
                    for i in range(n_products)
                ],
            }
        },
    }


# ──────────────────────────────────────────────────────────────────────────
# Test runner
# ──────────────────────────────────────────────────────────────────────────


def _section(title: str) -> None:
    print()
    print("=" * 78)
    print(f"  {title}")
    print("=" * 78)


def _check(cond: bool, msg: str) -> bool:
    status = "✓" if cond else "✗"
    print(f"  {status} {msg}")
    return cond


def _verify_output(
    excel_bytes: bytes,
    zone_config: dict,
    n_products: int,
) -> dict[str, Any]:
    """Inspect a generated workbook and return structured findings."""
    wb = load_workbook(io.BytesIO(excel_bytes))
    ws = wb.active

    prod_start = zone_config["zones"]["product_data"]["start"]
    prod_end_template = zone_config["zones"]["product_data"]["end"]
    template_prod_rows = prod_end_template - prod_start + 1
    row_delta = n_products - template_prod_rows
    new_summ_start = prod_end_template + 1 + row_delta

    findings = {
        "filled_product_rows": 0,
        "summary_at_correct_position": False,
        "summary_merges_present": [],
        "summary_merges_missing": [],
        "header_unchanged": True,
        "actual_max_row": ws.max_row,
        "expected_max_row": ws.max_row,
        "subtotal_formula_correct": False,
        "tax_formula_correct": False,
        "grand_total_formula_correct": False,
    }

    # Count actual product rows that have a non-empty product name
    name_col = column_index_from_string("B")
    for row in range(prod_start, prod_start + n_products + 5):
        cell = ws.cell(row=row, column=name_col)
        if cell.value and "Test Product" in str(cell.value):
            findings["filled_product_rows"] += 1

    # Check summary row position (look for "Sub Total" label in column A)
    expected_subtotal_row = new_summ_start
    actual_subtotal_cell = ws.cell(row=expected_subtotal_row, column=1).value
    findings["summary_at_correct_position"] = actual_subtotal_cell == "Sub Total"
    findings["actual_subtotal_label_at_expected_row"] = actual_subtotal_cell

    # Check expected summary merges exist
    actual_merges = {str(rng) for rng in ws.merged_cells.ranges}
    expected_merges = [
        f"D{new_summ_start}:E{new_summ_start}",       # subtotal label
        f"D{new_summ_start + 1}:E{new_summ_start + 1}",  # tax label
        f"D{new_summ_start + 2}:E{new_summ_start + 2}",  # grand total label
        f"A{new_summ_start + 3}:F{new_summ_start + 3}",  # signature line
    ]
    for em in expected_merges:
        if em in actual_merges:
            findings["summary_merges_present"].append(em)
        else:
            findings["summary_merges_missing"].append(em)

    # Check header unchanged
    findings["header_unchanged"] = (
        ws["A1"].value == "PURCHASE INQUIRY"
        and "A1:F1" in actual_merges
    )

    # Check formula correctness — the subtle bug compose-from-scratch can hit
    new_prod_end = prod_start + n_products - 1
    expected_subtotal = f"=SUM(F{prod_start}:F{new_prod_end})"
    actual_subtotal = ws.cell(row=new_summ_start, column=column_index_from_string("F")).value
    findings["subtotal_formula_correct"] = actual_subtotal == expected_subtotal
    findings["actual_subtotal_formula"] = actual_subtotal
    findings["expected_subtotal_formula"] = expected_subtotal

    expected_tax = f"=F{new_summ_start}*0.1"
    actual_tax = ws.cell(row=new_summ_start + 1, column=column_index_from_string("F")).value
    findings["tax_formula_correct"] = actual_tax == expected_tax

    expected_gt = f"=F{new_summ_start}+F{new_summ_start + 1}"
    actual_gt = ws.cell(row=new_summ_start + 2, column=column_index_from_string("F")).value
    findings["grand_total_formula_correct"] = actual_gt == expected_gt

    return findings


def main() -> int:
    template_bytes, zone_config = _build_japanese_style_template()

    test_cases = [5, 10, 50, 73]  # 73 = real Japanese inquiry size from production
    all_passed = True

    for n in test_cases:
        _section(f"compose_render with {n} products")
        order_data = _build_order_data(n)

        try:
            output_bytes = compose_render(template_bytes, zone_config, order_data, "1")
        except Exception as exc:
            print(f"  ✗ compose_render raised: {exc}")
            import traceback
            traceback.print_exc()
            all_passed = False
            continue

        findings = _verify_output(output_bytes, zone_config, n)

        ok = True
        ok &= _check(
            findings["filled_product_rows"] == n,
            f"product rows filled: {findings['filled_product_rows']} == {n}",
        )
        ok &= _check(
            findings["summary_at_correct_position"],
            f"Sub Total label at expected row (got: {findings['actual_subtotal_label_at_expected_row']!r})",
        )
        ok &= _check(
            len(findings["summary_merges_missing"]) == 0,
            f"all summary merges present (missing: {findings['summary_merges_missing']})",
        )
        ok &= _check(
            findings["header_unchanged"],
            "header section preserved (A1:F1 merge + title)",
        )
        ok &= _check(
            findings["actual_max_row"] >= 5 + n,
            f"workbook has at least {5 + n} rows (got {findings['actual_max_row']})",
        )
        ok &= _check(
            findings["subtotal_formula_correct"],
            f"subtotal formula = {findings.get('expected_subtotal_formula')!r} "
            f"(got {findings.get('actual_subtotal_formula')!r})",
        )
        ok &= _check(
            findings["tax_formula_correct"],
            "tax formula references new subtotal cell",
        )
        ok &= _check(
            findings["grand_total_formula_correct"],
            "grand total formula references new subtotal + tax cells",
        )

        if not ok:
            all_passed = False

    _section("Result")
    if all_passed:
        print("  ALL TEST CASES PASSED")
        print()
        print("  This proves: compose_render handles 5, 10, 50, AND 73 products")
        print("  with summary merges intact, formulas adjusted, and zero")
        print("  reliance on insert_rows / delete_rows / template_contract.")
        return 0
    else:
        print("  SOME TEST CASES FAILED — see above")
        return 1


if __name__ == "__main__":
    sys.exit(main())
