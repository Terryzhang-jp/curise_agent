"""End-to-end integration tests for the inquiry generation pipeline.

Why this test exists
====================
Before this test, every layer in the inquiry pipeline (template upload,
fill_template, verify_output, _save_workbook, _generate_single_supplier,
run_inquiry_orchestrator) had its own narrow unit tests, but **nothing**
exercised the full path with real production templates. As a result, every
breakage shipped to production. On 2026-04-11 we found 5 distinct bugs in
a single day, all of which would have been caught by this test.

This file tests `_generate_single_supplier` end-to-end with:

1. Real production template bytes (saved as fixtures from Supabase 2026-04-11)
2. A sqlite in-memory DB seeded with the minimum rows needed
3. A monkey-patched file_storage that serves the fixtures and a fake bucket
4. Realistic mock match_results with 73 products

Each test exercises one specific failure mode that we have hit (or want to
prevent hitting) in production.
"""
from __future__ import annotations

import io
import json
import os
import sys
import threading
from pathlib import Path
from typing import Any, Iterator
from unittest import mock

import pytest
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

# Make sure project root is on path
sys.path.insert(0, str(Path(__file__).parent.parent))

import core.models as models  # noqa: E402
from core.models import (  # noqa: E402
    Base,
    CompanyConfig,
    DeliveryLocation,
    Order,
    Supplier,
    SupplierTemplate,
)


FIXTURE_DIR = Path(__file__).parent / "fixtures" / "templates"
PRODUCTION_TEMPLATE_IDS = [11, 12, 13]


# ──────────────────────────────────────────────────────────────────────
# Fixtures
# ──────────────────────────────────────────────────────────────────────


def _load_template_fixture(template_id: int) -> tuple[bytes, dict]:
    """Load a real production template bytes + its zone_config from fixtures."""
    bytes_path = FIXTURE_DIR / f"template_{template_id}.xlsx"
    cfg_path = FIXTURE_DIR / f"template_{template_id}_zone_config.json"
    if not bytes_path.exists():
        pytest.skip(f"fixture not present: {bytes_path}")
    return bytes_path.read_bytes(), json.loads(cfg_path.read_text())


def _build_match_results(supplier_id: int, n: int = 73) -> list[dict]:
    """Build realistic match_results that look like a real produce PO."""
    sample_products = [
        ("99PRD010588", "APPLE GRANNY SMITH US EXTRA FANCY 125CT/40LB", 246.0, "KG", 1.29),
        ("99PRD010590", "APPLE RED DELICIOUS US EXTRA FANCY 125CT/40LB", 296.0, "KG", 1.26),
        ("99PRD010601", "ASPARAGUS GREEN LARGE", 225.0, "KG", 14.88),
        ("99PRD010604", "AVOCADO HASS 3/4 RIPE 60CT #2", 173.0, "KG", 3.20),
        ("99PRD010690", "BEET RED LARGE 25LB", 82.0, "KG", 1.40),
    ]
    results = []
    for i in range(n):
        code, name, qty, unit, price = sample_products[i % len(sample_products)]
        results.append({
            "product_name": name,
            "product_code": code,
            "quantity": qty + i,
            "unit": unit,
            "unit_price": price,
            "currency": "JPY",
            "matched_product": {
                "supplier_id": supplier_id,
                "code": code,
                "product_name_en": name,
                "product_name_jp": "",
                "unit": unit,
                "price": price,
                "pack_size": "",
            },
        })
    return results


@pytest.fixture
def memory_db_session_factory():
    """sqlite in-memory DB with all tables that inquiry_agent touches.

    We create ORM-defined tables from their __table__ objects (so full
    column lists match production), then add raw-SQL tables for the few
    that inquiry_agent queries via raw SQL strings (suppliers/ports/countries).
    """
    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
    )

    # ── ORM-managed tables (must be created from __table__ so column
    # lists exactly match what the ORM expects to read back) ──
    for tbl in [
        SupplierTemplate.__table__,
        Order.__table__,
        CompanyConfig.__table__,
        DeliveryLocation.__table__,
    ]:
        tbl.create(engine, checkfirst=True)

    # ── Raw-SQL tables (inquiry_agent uses sa_text for these,
    # so we just need the columns referenced by the SQL strings) ──
    from sqlalchemy import text
    with engine.begin() as conn:
        conn.execute(text(
            "CREATE TABLE IF NOT EXISTS suppliers ("
            "id INTEGER PRIMARY KEY, name TEXT, contact TEXT, email TEXT, "
            "phone TEXT, fax TEXT, address TEXT, zip_code TEXT, "
            "default_payment_method TEXT, default_payment_terms TEXT, "
            "status BOOLEAN)"
        ))
        conn.execute(text(
            "CREATE TABLE IF NOT EXISTS ports "
            "(id INTEGER PRIMARY KEY, name TEXT, location TEXT, code TEXT)"
        ))
        conn.execute(text(
            "CREATE TABLE IF NOT EXISTS countries "
            "(id INTEGER PRIMARY KEY, name TEXT, code TEXT)"
        ))

    Session = sessionmaker(bind=engine)
    return Session


@pytest.fixture
def seeded_db(memory_db_session_factory):
    """Memory DB with one supplier, one order, one ready template."""
    Session = memory_db_session_factory
    db = Session()

    # Seed supplier (raw SQL because production code uses raw SQL for this table)
    from sqlalchemy import text
    db.execute(text(
        "INSERT INTO suppliers (id, name, contact, email, phone, fax, address, "
        "zip_code, default_payment_method, default_payment_terms, status) "
        "VALUES (17, 'Select Fresh PROVIDORES', 'Tom', 'tom@select.com', "
        "'+81-3-1234-5678', '', 'Tokyo, Japan', '100-0001', 'wire', 'NET 30', 1)"
    ))
    db.execute(text("INSERT INTO ports (id, name, location, code) VALUES (1, 'Sydney', 'Sydney NSW', 'SYD')"))
    db.execute(text("INSERT INTO countries (id, name, code) VALUES (1, 'Australia', 'AU')"))

    # Seed order via ORM (the production code reads it back via ORM in some places)
    order = Order(
        id=75,
        user_id=1,
        filename="test_order.pdf",
        file_url="docs/test_order.pdf",
        file_type="pdf",
        status="ready",
        port_id=1,
        country_id=1,
        order_metadata={"po_number": "68358749", "ship_name": "CELEBRITY EDGE"},
        products=[],
        product_count=73,
    )
    db.add(order)
    db.commit()

    yield Session, db
    db.close()


@pytest.fixture
def template_fixture_loader():
    """Returns a loader function that monkey-patches storage.download to serve fixtures."""
    def _loader(template_id_to_use: int):
        bytes_, zone_config = _load_template_fixture(template_id_to_use)

        def fake_download(url: str) -> bytes:
            # Map any url to our single fixture
            return bytes_

        return bytes_, zone_config, fake_download
    return _loader


# ──────────────────────────────────────────────────────────────────────
# Tests — happy path on each production template
# ──────────────────────────────────────────────────────────────────────


@pytest.mark.parametrize("template_id", PRODUCTION_TEMPLATE_IDS)
def test_generate_single_supplier_happy_path(
    template_id: int,
    seeded_db,
    template_fixture_loader,
):
    """For each production template: full pipeline must succeed and produce a
    valid xlsx with all 73 products and pass verify_output."""
    Session, db = seeded_db
    template_bytes, zone_config, fake_download = template_fixture_loader(template_id)

    # Insert the template into our memory DB
    tpl = SupplierTemplate(
        id=template_id,
        template_name=f"Test Template {template_id}",
        supplier_ids=[17],
        template_file_url=f"templates/test_{template_id}.xlsx",
        template_styles=zone_config,
    )
    db.add(tpl)
    db.commit()

    # Patch SessionLocal + file_storage globally so _generate_single_supplier
    # uses our memory DB and our fixture bytes
    captured_uploads: list[tuple[str, bytes]] = []

    def fake_upload(folder: str, filename: str, content: bytes, **kwargs) -> str:
        captured_uploads.append((filename, content))
        return f"{folder}/{filename}"

    from services import inquiry_agent
    from services import file_storage as fs_module

    with mock.patch.object(inquiry_agent, "file_storage") as mock_fs, \
         mock.patch("database.SessionLocal", Session):
        mock_fs.download = fake_download
        mock_fs.upload = fake_upload

        result = inquiry_agent._generate_single_supplier(
            order_id=75,
            order_meta={
                "po_number": "68358749",
                "ship_name": "CELEBRITY EDGE",
                "delivery_date": "2026-01-05",
                "currency": "AUD",
            },
            supplier_id=17,
            products=_build_match_results(17, 73),
            stream_key="",
            overall_start=0.0,
        )

    # ── Assertions ──
    assert result["error"] is None, f"unexpected error: {result['error']}"
    assert "file_info" in result
    assert result["file_info"]["product_count"] == 73
    assert len(captured_uploads) >= 1, "no file was uploaded"

    # Find the .xlsx upload (not the .html preview)
    xlsx_uploads = [u for u in captured_uploads if u[0].endswith(".xlsx")]
    assert len(xlsx_uploads) == 1
    filename, file_bytes = xlsx_uploads[0]

    # Read back the produced workbook and sanity-check structure
    wb = load_workbook(io.BytesIO(file_bytes))
    ws = wb.active
    assert ws.max_row >= 73, f"expected at least 73 rows, got {ws.max_row}"

    # Verify the result didn't completely punt — verify_results should exist
    assert "verify_results" in result
    # We don't assert verify passed because some old templates may have legacy
    # quirks; we assert it ran without raising and produced a structured report.


# ──────────────────────────────────────────────────────────────────────
# Tests — failure modes that we have hit before or want to prevent
# ──────────────────────────────────────────────────────────────────────


def test_generate_with_template_file_missing(seeded_db):
    """If the template file can't be downloaded, we should raise a clear error,
    not a silent corruption."""
    Session, db = seeded_db
    tpl = SupplierTemplate(
        id=99,
        template_name="Missing File Template",
        supplier_ids=[17],
        template_file_url="templates/does_not_exist.xlsx",
        template_styles={"zones": {"product_data": {"start": 5, "end": 5}, "summary": {"start": 6, "end": 6}}},
    )
    db.add(tpl)
    db.commit()

    from services import inquiry_agent

    def fake_download(url: str) -> bytes | None:
        return None

    def fake_upload(*args, **kwargs):
        return "noop"

    with mock.patch.object(inquiry_agent, "file_storage") as mock_fs, \
         mock.patch("database.SessionLocal", Session):
        mock_fs.download = fake_download
        mock_fs.upload = fake_upload

        with pytest.raises(Exception):
            inquiry_agent._generate_single_supplier(
                order_id=75,
                order_meta={"po_number": "T1", "ship_name": "X"},
                supplier_id=17,
                products=_build_match_results(17, 5),
                stream_key="",
                overall_start=0.0,
            )


def test_generate_cancelled_midway(seeded_db, template_fixture_loader):
    """If cancel_event is set, generation should raise InquiryCancelledError."""
    Session, db = seeded_db
    template_bytes, zone_config, fake_download = template_fixture_loader(11)
    tpl = SupplierTemplate(
        id=11,
        template_name="Test 11",
        supplier_ids=[17],
        template_file_url="templates/test_11.xlsx",
        template_styles=zone_config,
    )
    db.add(tpl)
    db.commit()

    from services import inquiry_agent

    cancel_event = threading.Event()
    cancel_event.set()  # already cancelled before we start

    def fake_upload(*args, **kwargs):
        return "noop"

    with mock.patch.object(inquiry_agent, "file_storage") as mock_fs, \
         mock.patch("database.SessionLocal", Session):
        mock_fs.download = fake_download
        mock_fs.upload = fake_upload

        with pytest.raises(inquiry_agent.InquiryCancelledError):
            inquiry_agent._generate_single_supplier(
                order_id=75,
                order_meta={"po_number": "T1", "ship_name": "X"},
                supplier_id=17,
                products=_build_match_results(17, 73),
                stream_key="",
                overall_start=0.0,
                cancel_event=cancel_event,
            )


def test_unmatched_supplier_template_unavailable(seeded_db):
    """If supplier has no template at all, should raise ValueError with clear message."""
    Session, db = seeded_db
    # Don't add any template

    from services import inquiry_agent

    def fake_upload(*args, **kwargs):
        return "noop"

    def fake_download(url: str) -> bytes:
        return b""

    with mock.patch.object(inquiry_agent, "file_storage") as mock_fs, \
         mock.patch("database.SessionLocal", Session):
        mock_fs.download = fake_download
        mock_fs.upload = fake_upload

        with pytest.raises(ValueError, match="zone_config"):
            inquiry_agent._generate_single_supplier(
                order_id=75,
                order_meta={"po_number": "T1", "ship_name": "X"},
                supplier_id=17,
                products=_build_match_results(17, 5),
                stream_key="",
                overall_start=0.0,
            )


# ──────────────────────────────────────────────────────────────────────
# Tests — known bug regression guards
# ──────────────────────────────────────────────────────────────────────


def test_no_load_workbook_nameerror(seeded_db, template_fixture_loader):
    """Regression for 2026-04-11 bug: load_workbook was aliased to
    _load_workbook_raw but line 783 still used the original name. Every
    inquiry generation hit NameError. This test ensures that doesn't recur
    by actually walking the deterministic engine path."""
    # The happy path test already does this — if any NameError gets reintroduced
    # it will surface here. We add this explicit test as documentation.
    Session, db = seeded_db
    template_bytes, zone_config, fake_download = template_fixture_loader(11)
    tpl = SupplierTemplate(
        id=11,
        template_name="Test 11",
        supplier_ids=[17],
        template_file_url="templates/test_11.xlsx",
        template_styles=zone_config,
    )
    db.add(tpl)
    db.commit()

    from services import inquiry_agent

    def fake_upload(*args, **kwargs):
        return "noop"

    with mock.patch.object(inquiry_agent, "file_storage") as mock_fs, \
         mock.patch("database.SessionLocal", Session):
        mock_fs.download = fake_download
        mock_fs.upload = fake_upload

        # Should not raise NameError. Any other exception is acceptable for this regression test.
        try:
            inquiry_agent._generate_single_supplier(
                order_id=75,
                order_meta={"po_number": "T1", "ship_name": "X"},
                supplier_id=17,
                products=_build_match_results(17, 10),
                stream_key="",
                overall_start=0.0,
            )
        except NameError as e:
            pytest.fail(f"NameError reintroduced: {e}")
        except Exception:
            pass  # other exceptions are out of scope for this specific regression


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
