"""
Excel writer for generating inquiry/PO Excel files.

Uses openpyxl to create Excel files either from a SupplierTemplate or with a generic layout.
Supports loading an original template file to preserve formatting, formulas, and styles.
"""

from __future__ import annotations

import io
import logging
import threading
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

logger = logging.getLogger(__name__)


def generate_inquiry_excel(
    template: Any | None,  # SupplierTemplate or None
    order_metadata: dict[str, Any],
    products: list[dict[str, Any]],
    supplier_id: int,
    template_file_path: str | None = None,
    field_mapping: dict[str, str] | None = None,
) -> bytes:
    """Generate an inquiry Excel file.

    Three paths:
    1. template_file_path exists → load_workbook from file (preserves format/formulas/styles)
    2. template with field_positions → new Workbook with template-driven fill
    3. No template → new Workbook with generic layout

    Returns bytes of the .xlsx file.
    """
    if template_file_path and template and template.field_positions:
        # Path 1: Load original template file (preserves all formatting)
        wb = load_workbook(template_file_path)
        ws = wb.active
        _fill_with_template(ws, template, order_metadata, products, field_mapping)
    elif template and template.field_positions:
        # Path 2: New workbook with template config
        wb = Workbook()
        ws = wb.active
        ws.title = "Inquiry"
        _fill_with_template(ws, template, order_metadata, products, field_mapping)
    else:
        # Path 3: Generic layout
        wb = Workbook()
        ws = wb.active
        ws.title = "Inquiry"
        _fill_generic(ws, order_metadata, products, supplier_id)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _fill_with_template(ws, template, metadata: dict, products: list[dict],
                        field_mapping: dict[str, str] | None = None) -> None:
    """Fill worksheet using SupplierTemplate's field_positions and product_table_config.

    field_mapping: maps template field_key → metadata key (e.g. {"delivery_date": "deliver_on_date"}).
    When None, falls back to using field_key directly as metadata key (backward compatible).
    """
    field_positions = template.field_positions or {}
    table_config = template.product_table_config or {}
    formula_columns = set(table_config.get("formula_columns", []))

    # 1. Fill header fields
    for field_key, pos_info in field_positions.items():
        position = pos_info if isinstance(pos_info, str) else pos_info.get("position", "")
        if not position:
            continue
        mapped_key = field_mapping.get(field_key, field_key) if field_mapping else field_key
        value = metadata.get(mapped_key, "")
        if value:
            cell = ws[position]
            if isinstance(cell, MergedCell):
                logger.debug("Skipping merged cell %s for field %s", position, field_key)
                continue
            cell.value = value

    # 2. Fill product table
    start_row = table_config.get("start_row", 12)
    columns = table_config.get("columns", {})

    for i, product in enumerate(products):
        row = start_row + i
        matched = product.get("matched_product") or {}

        for col_letter, field_key in columns.items():
            if col_letter.upper() in formula_columns or col_letter in formula_columns:
                continue  # Skip formula columns

            # Resolve value: product dict first, then matched (normalized) dict.
            # Special cases only for fields that need cross-dict or metadata fallback.
            if field_key == "line_number":
                value = i + 1
            elif field_key == "po_number":
                value = product.get("po_number") or metadata.get("po_number", "")
            elif field_key == "currency":
                value = product.get("currency") or matched.get("currency") or metadata.get("currency", "")
            elif field_key == "description":
                value = matched.get("pack_size") or product.get("description", "")
            elif field_key == "product_name_en":
                value = product.get("product_name") or matched.get("product_name_en", "")
            else:
                val = product.get(field_key)
                value = val if val is not None else matched.get(field_key, "")

            if value != "" and value is not None:
                cell = ws[f"{col_letter}{row}"]
                if isinstance(cell, MergedCell):
                    continue
                cell.value = value


def _fill_generic(ws, metadata: dict, products: list[dict], supplier_id: int) -> None:
    """Generate a generic inquiry layout."""
    header_font = Font(bold=True, size=14)
    label_font = Font(bold=True, size=10)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_fill = PatternFill(start_color="D4A853", end_color="D4A853", fill_type="solid")
    header_text_font = Font(bold=True, size=10, color="FFFFFF")

    # Title
    ws.merge_cells("A1:G1")
    ws["A1"] = "Purchase Order / 注文書"
    ws["A1"].font = header_font
    ws["A1"].alignment = Alignment(horizontal="center")

    # Metadata
    meta_rows = [
        ("A3", "PO Number:", "B3", metadata.get("po_number", "")),
        ("A4", "Order Date:", "B4", metadata.get("order_date", "")),
        ("A5", "Delivery Date:", "B5", metadata.get("delivery_date", "")),
        ("A6", "Ship Name:", "B6", metadata.get("ship_name", "")),
        ("A7", "Currency:", "B7", metadata.get("currency", "")),
        ("D3", "Supplier ID:", "E3", str(supplier_id)),
        ("D4", "Vendor:", "E4", metadata.get("vendor_name", "")),
        ("D5", "Port:", "E5", metadata.get("destination_port", "")),
    ]

    for label_cell, label, value_cell, value in meta_rows:
        ws[label_cell] = label
        ws[label_cell].font = label_font
        ws[value_cell] = value

    # Product table header
    table_start = 9
    headers = ["No.", "Product Code", "Product Name", "Qty", "Unit", "Unit Price", "Total"]
    col_widths = [6, 15, 35, 10, 8, 12, 12]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=table_start, column=col_idx, value=header)
        cell.font = header_text_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = width

    # Product rows
    for i, product in enumerate(products, 1):
        row = table_start + i
        matched = product.get("matched_product", {})

        values = [
            i,
            product.get("product_code") or matched.get("code", ""),
            product.get("product_name") or matched.get("product_name_en", ""),
            product.get("quantity", ""),
            product.get("unit") or matched.get("unit", ""),
            product.get("unit_price", ""),
            product.get("total_price", ""),
        ]

        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.border = border

    # Total row
    total_row = table_start + len(products) + 1
    ws.cell(row=total_row, column=5, value="Total:").font = label_font
    total_amount = 0
    for p in products:
        try:
            total_amount += float(p.get("total_price", 0) or 0)
        except (ValueError, TypeError):
            pass
    ws.cell(row=total_row, column=7, value=total_amount).font = label_font


# ─── InquiryWorkbook: Mutable wrapper for agentic cell writes ───

class InquiryWorkbook:
    """Mutable workbook wrapper for incremental cell writes + live HTML preview."""

    def __init__(self):
        self._wb: Workbook | None = None
        self._ws = None
        self._source: str = "empty"  # "template_file" | "template_new" | "generic" | "empty"
        self._lock = threading.Lock()

    def load_template(self, file_path: str) -> None:
        """Load .xlsx template file (preserves formatting/formulas)."""
        self._wb = load_workbook(file_path)
        self._ws = self._wb.active
        self._source = "template_file"

    def create_from_config(self) -> None:
        """Create new workbook for template-driven fill without a file."""
        self._wb = Workbook()
        self._ws = self._wb.active
        self._ws.title = "Inquiry"
        self._source = "template_new"

    def create_generic(self, metadata: dict, products: list, supplier_id: int) -> None:
        """Create generic layout workbook (fully filled, no agent needed)."""
        self._wb = Workbook()
        self._ws = self._wb.active
        self._ws.title = "Inquiry"
        _fill_generic(self._ws, metadata, products, supplier_id)
        self._source = "generic"

    def write_cells(self, cells: list[dict]) -> list[dict]:
        """Write values to cells. Returns status per cell.
        cells: [{cell: "B3", value: "xxx"}, ...]
        """
        with self._lock:
            results = []
            for item in cells:
                ref = item.get("cell", "")
                val = item.get("value", "")
                try:
                    cell = self._ws[ref]
                    if isinstance(cell, MergedCell):
                        results.append({"cell": ref, "status": "skipped_merged"})
                        continue
                    cell.value = val
                    results.append({"cell": ref, "status": "ok", "value": str(val)[:50]})
                except Exception as e:
                    results.append({"cell": ref, "status": "error", "error": str(e)})
            return results

    def write_product_rows(self, start_row: int, columns: dict[str, str],
                           products: list[dict], formula_columns: list[str] | None = None,
                           metadata: dict | None = None) -> int:
        """Write product data rows. Returns count of rows written.
        columns: {"A": "line_number", "C": "product_code", ...}
        metadata: optional order metadata for fields like po_number, currency.
        """
        skip = set(c.upper() for c in (formula_columns or []))
        meta = metadata or {}
        with self._lock:
            # ── Preemptive merge cleanup ──────────────────────────
            # Templates may have merged cells (e.g. summary zone I:J merges)
            # in the rows we're about to write to. Remove merge ranges and
            # purge MergedCell objects so every cell is writable.
            write_end = start_row + len(products) - 1
            merges_to_remove = [
                mr for mr in list(self._ws.merged_cells.ranges)
                if mr.min_row <= write_end and mr.max_row >= start_row
            ]
            for mr in merges_to_remove:
                self._ws.merged_cells.remove(mr)
            if merges_to_remove:
                max_col = self._ws.max_column or 12
                for row in range(start_row, write_end + 1):
                    for col in range(1, max_col + 1):
                        if (row, col) in self._ws._cells and isinstance(
                            self._ws._cells[(row, col)], MergedCell
                        ):
                            del self._ws._cells[(row, col)]
                logger.info(
                    "write_product_rows: removed %d merge ranges in rows %d-%d",
                    len(merges_to_remove), start_row, write_end,
                )

            for i, product in enumerate(products):
                row = start_row + i
                matched = product.get("matched_product") or {}
                for col_letter, field_key in columns.items():
                    if col_letter.upper() in skip:
                        continue
                    # Resolve value: matched (authoritative) → product (extracted) → metadata fallback
                    if field_key == "line_number":
                        value = i + 1
                    elif field_key == "po_number":
                        value = product.get("po_number") or meta.get("po_number", "")
                    elif field_key == "currency":
                        value = product.get("currency") or matched.get("currency") or meta.get("currency", "")
                    elif field_key == "description":
                        value = matched.get("pack_size") or product.get("description", "")
                    elif field_key == "product_name_en":
                        value = product.get("product_name") or matched.get("product_name_en", "")
                    elif field_key == "unit":
                        # matched_product.unit is authoritative (clean DB value);
                        # product.unit from extraction can be corrupted (e.g. "KG2.2")
                        value = matched.get("unit") or product.get("unit", "")
                    elif field_key == "unit_price":
                        # matched_product uses "price" not "unit_price"
                        # Use `is not None` — price=0 is valid, `or` treats it as falsy
                        val = product.get("unit_price")
                        value = val if val is not None else matched.get("price", "")
                    else:
                        val = product.get(field_key)
                        value = val if val is not None else matched.get(field_key, "")
                    if value != "" and value is not None:
                        cell = self._ws[f"{col_letter}{row}"]
                        if not isinstance(cell, MergedCell):
                            cell.value = value
                            # Apply 2-decimal format to price/amount columns
                            if field_key in ("unit_price", "total_price", "amount"):
                                cell.number_format = '0.00'

                # Also format formula columns (e.g. L = H*J total) with 2 decimals
                for fc in skip:
                    cell = self._ws[f"{fc}{row}"]
                    if not isinstance(cell, MergedCell) and cell.value is not None:
                        cell.number_format = '0.00'

            return len(products)

    def safe_set_cell(self, ref: str, value, number_format: str | None = None) -> bool:
        """Write a single cell safely, skipping MergedCells.

        Returns True if the write succeeded, False if skipped (merged/error).
        All external code should use this instead of accessing _ws directly.
        """
        try:
            cell = self._ws[ref]
            if isinstance(cell, MergedCell):
                logger.debug("Skipping merged cell %s", ref)
                return False
            cell.value = value
            if number_format:
                cell.number_format = number_format
            return True
        except Exception as e:
            logger.warning("Failed to write cell %s: %s", ref, e)
            return False

    def render_html(self) -> str:
        """Convert current workbook state to HTML via xlsx2html."""
        if self._wb is None:
            return "<p>No workbook loaded</p>"
        from services.template_analyzer import generate_template_html
        buf = io.BytesIO()
        self._wb.save(buf)
        return generate_template_html(buf.getvalue())

    def save_bytes(self) -> bytes:
        """Save workbook to bytes."""
        if self._wb is None:
            raise RuntimeError("No workbook initialized")
        buf = io.BytesIO()
        self._wb.save(buf)
        return buf.getvalue()

    @property
    def is_generic(self) -> bool:
        return self._source == "generic"
