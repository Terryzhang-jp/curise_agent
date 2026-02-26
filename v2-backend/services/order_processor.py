"""
Order Processor — Vision one-step extraction + Agent-based smart matching.

Flow: vision_extract → run_agent_matching
- PDF: Gemini Vision directly extracts structured JSON (metadata + products)
- Excel: openpyxl reads content, then Gemini structures it
- Agent uses get_db_schema + query_db tools to match products intelligently
"""

from __future__ import annotations

import json
import logging
import time
from datetime import datetime

from database import SessionLocal
from models import Order

logger = logging.getLogger(__name__)

# ─── Standard Metadata Schema ────────────────────────────────────
# Single source of truth for the 8 standard order_metadata keys.
# All extraction prompts reference this; normalize_metadata() enforces it.

ORDER_METADATA_SCHEMA = {
    "po_number": "采购订单号 (PO Number / Order No. / 発注番号)",
    "ship_name": "船名 (Ship Name / Vessel Name) — 字符串",
    "vendor_name": "供应商名称 (Vendor / Supplier Name) — 字符串，不是对象",
    "delivery_date": "交货日期 (Delivery Date / Deliver On) — 格式 YYYY-MM-DD",
    "order_date": "订单日期 (Order Date / PO Date) — 格式 YYYY-MM-DD",
    "currency": "币种 (Currency) — 如 USD, AUD, JPY, THB",
    "destination_port": "目的港 (Destination Port / Final Destination)",
    "total_amount": "总金额 (Total Amount) — 数字类型",
}

# Aliases: AI sometimes uses these instead of the standard keys above.
_FIELD_ALIASES = {
    "po_number": ["order_number", "po_no", "purchase_order_number", "po_num"],
    "vendor_name": ["vendor", "supplier_name", "supplier"],
    "delivery_date": ["expected_delivery_date", "deliver_on_date", "deliver_date"],
    "order_date": ["purchase_date", "purchase_order_date"],
    "destination_port": ["destination", "port_name", "final_destination"],
    "total_amount": ["grand_total", "order_total"],
    "ship_name": ["vessel_name", "vessel"],
    "currency": ["ccy"],
}

_STANDARD_KEYS = set(ORDER_METADATA_SCHEMA.keys())


def _metadata_schema_prompt() -> str:
    """Return a reusable prompt fragment listing the 8 required metadata keys."""
    lines = [
        "## order_metadata 字段要求（必须严格使用以下键名）",
        "以下 8 个键名是固定的，不允许使用任何替代名称：",
    ]
    for k, v in ORDER_METADATA_SCHEMA.items():
        lines.append(f"  - {k}: {v}")
    lines.append("  - extra_fields: {} — 收纳上述 8 个键之外的其他可见元数据（如 port_code, voyage_number, loading_date 等）")
    lines.append("")
    lines.append("禁止使用以下替代键名：order_number, vendor, supplier_name, expected_delivery_date, deliver_on_date, destination, final_destination, grand_total, vessel_name 等。")
    lines.append("vendor_name 必须是纯字符串，不是对象。日期格式必须为 YYYY-MM-DD。total_amount 必须是数字。看不到的字段用 null。")
    return "\n".join(lines)


def normalize_metadata(raw: dict) -> dict:
    """Normalize extracted metadata to standard keys. Lightweight safety net.

    1. Standard key already has a value → keep it.
    2. Otherwise check aliases → promote to standard key.
    3. vendor_name is an object → extract .name or .company.
    4. Non-standard keys → move into extra_fields.
    5. Missing standard keys → null.
    """
    if not raw:
        return {k: None for k in _STANDARD_KEYS} | {"extra_fields": {}}

    result = {}
    consumed = set()

    # Pass 1: collect standard keys already present
    for key in _STANDARD_KEYS:
        if key in raw and raw[key] is not None:
            result[key] = raw[key]
            consumed.add(key)

    # Pass 2: fill missing standard keys from aliases
    for std_key, aliases in _FIELD_ALIASES.items():
        if std_key in result and result[std_key] is not None:
            continue
        for alias in aliases:
            if alias in raw and raw[alias] is not None:
                result[std_key] = raw[alias]
                consumed.add(alias)
                break

    # Pass 3: fix vendor_name if it's an object
    vn = result.get("vendor_name")
    if isinstance(vn, dict):
        result["vendor_name"] = vn.get("name") or vn.get("company") or vn.get("company_name") or str(vn)

    # Pass 4: collect extra_fields (non-standard, non-alias keys that have values)
    alias_keys = set()
    for aliases in _FIELD_ALIASES.values():
        alias_keys.update(aliases)

    extra = raw.get("extra_fields", {}) if isinstance(raw.get("extra_fields"), dict) else {}
    for k, v in raw.items():
        if k in _STANDARD_KEYS or k in consumed or k == "extra_fields" or v is None:
            continue
        extra[k] = v
    result["extra_fields"] = extra

    # Pass 5: fill missing standard keys with None
    for key in _STANDARD_KEYS:
        if key not in result:
            result[key] = None

    return result


def _parse_delivery_date(date_str: str | None) -> datetime | None:
    """Parse delivery_date string into datetime. Returns None if unparseable."""
    if not date_str:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y", "%m/%d/%Y",
                "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(date_str.strip(), fmt)
        except ValueError:
            continue
    return None


# ─── Vision Extraction ───────────────────────────────────────────

VISION_EXTRACT_PROMPT = f"""分析这个采购订单文档，提取所有可见信息。

{_metadata_schema_prompt()}

## 返回格式
返回纯 JSON（不要 markdown 代码块）：
{{
  "order_metadata": {{
    "po_number": "...",
    "ship_name": "...",
    "vendor_name": "...",
    "delivery_date": "YYYY-MM-DD",
    "order_date": "YYYY-MM-DD",
    "currency": "...",
    "destination_port": "...",
    "total_amount": 数字,
    "extra_fields": {{}}
  }},
  "products": [
    {{
      "line_number": 行号,
      "product_code": "产品代码/编号",
      "product_name": "产品名称/描述",
      "quantity": 数量,
      "unit": "单位",
      "unit_price": 单价,
      "total_price": 总价
    }}
  ]
}}
只提取文档中可见的信息，不要编造。数字使用数值类型。"""


def vision_extract(file_bytes: bytes, file_type: str) -> dict:
    """One-step extraction: PDF/Excel → structured order JSON via Gemini Vision."""

    if file_type != "pdf":
        return _extract_and_structure_excel(file_bytes)

    from services.pdf_analyzer import _get_model, _pdf_bytes_to_images, _parse_json_response

    start_time = time.time()
    images = _pdf_bytes_to_images(file_bytes)
    model = _get_model()
    content = [VISION_EXTRACT_PROMPT] + images

    last_error = None
    for attempt in range(2):
        try:
            response = model.generate_content(content)
            result = _parse_json_response(response.text.strip())
            processing_time = time.time() - start_time

            return {
                "order_metadata": normalize_metadata(result.get("order_metadata", {})),
                "products": result.get("products", []),
                "extraction_method": "gemini_vision",
                "page_count": len(images),
                "processing_time": round(processing_time, 2),
            }
        except (ValueError, Exception) as e:
            last_error = e
            logger.warning("Vision extraction attempt %d failed: %s", attempt + 1, str(e))
            if attempt == 0:
                time.sleep(1)
    raise last_error


def _extract_and_structure_excel(file_bytes: bytes) -> dict:
    """Extract from Excel: read with openpyxl, then structure with Gemini."""
    import io
    from openpyxl import load_workbook
    from services.pdf_analyzer import _get_model, _parse_json_response

    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    text_parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(cells):
                text_parts.append(" | ".join(cells))

    full_text = "\n".join(text_parts)
    if not full_text.strip():
        return {
            "order_metadata": {},
            "products": [],
            "extraction_method": "openpyxl_empty",
            "page_count": len(wb.sheetnames),
        }

    model = _get_model()
    schema_prompt = _metadata_schema_prompt()
    prompt = f"""你是订单数据提取专家。请从以下 Excel 内容中提取结构化的采购订单数据。

## Excel 内容
{full_text[:12000]}

{schema_prompt}

## 返回格式
返回纯 JSON（不要 markdown 代码块）：
{{
  "order_metadata": {{
    "po_number": "...",
    "ship_name": "...",
    "vendor_name": "...",
    "delivery_date": "YYYY-MM-DD",
    "order_date": "YYYY-MM-DD",
    "currency": "...",
    "destination_port": "...",
    "total_amount": 数字,
    "extra_fields": {{}}
  }},
  "products": [
    {{
      "line_number": 行号,
      "product_code": "产品代码/编号",
      "product_name": "产品名称/描述",
      "quantity": 数量,
      "unit": "单位",
      "unit_price": 单价,
      "total_price": 总价
    }}
  ]
}}
提取所有产品行，保留原始值。数字使用数值类型。看不到的字段用 null。"""

    last_error = None
    for attempt in range(2):
        try:
            response = model.generate_content([prompt])
            result = _parse_json_response(response.text.strip())
            return {
                "order_metadata": normalize_metadata(result.get("order_metadata", {})),
                "products": result.get("products", []),
                "extraction_method": "openpyxl_gemini",
                "page_count": len(wb.sheetnames),
            }
        except (ValueError, Exception) as e:
            last_error = e
            logger.warning("Excel structuring attempt %d failed: %s", attempt + 1, str(e))
            if attempt == 0:
                time.sleep(1)
    raise last_error


# ─── Matching (code-first, LLM-minimal) ─────────────────────────
#
# Step 1: Deterministic geo matching (code) + single LLM fallback
# Step 2: Batch code matching via _match_products_against_db (0 LLM)
# Step 3: Single LLM call for ambiguous items (0-1 LLM call)

# Currency string → country code mapping
_CURRENCY_TO_COUNTRY = {
    "JPY": "JPN", "YEN": "JPN", "Japanese Yen": "JPN", "円": "JPN",
    "AUD": "AUS", "Australian Dollar": "AUS", "A$": "AUS",
    "USD": "USA", "US Dollar": "USA", "US$": "USA",
    "THB": "THA", "Thai Baht": "THA", "Baht": "THA",
    "CNY": "CHN", "RMB": "CHN", "Chinese Yuan": "CHN", "人民币": "CHN",
    "VND": "VNM", "Vietnamese Dong": "VNM",
    "EUR": None,  # Ambiguous — multiple European countries
    "GBP": "GBR", "British Pound": "GBR",
    "CAD": "CAN", "Canadian Dollar": "CAN",
    "NZD": "NZL", "New Zealand Dollar": "NZL",
    "KRW": "KOR", "Korean Won": "KOR",
    "TWD": "TWN", "Taiwan Dollar": "TWN",
    "INR": "IND", "Indian Rupee": "IND",
    "IDR": "IDN", "Indonesian Rupiah": "IDN",
    "MYR": "MYS", "Malaysian Ringgit": "MYS",
    "PHP": "PHL", "Philippine Peso": "PHL",
    "MXN": "MEX", "Mexican Peso": "MEX",
    "BRL": "BRA", "Brazilian Real": "BRA",
    "SEK": "SWE", "Swedish Krona": "SWE",
    "NOK": "NOR", "Norwegian Krone": "NOR",
    "DKK": "DNK", "Danish Krone": "DNK",
    "CHF": "CHE", "Swiss Franc": "CHE",
    "HKD": "HKG", "Hong Kong Dollar": "HKG",
    "PLN": "POL", "Polish Zloty": "POL",
    "TRY": "TUR", "Turkish Lira": "TUR",
    "CLP": "CHL", "Chilean Peso": "CHL",
}


def _resolve_geo(order_id: int, metadata: dict, db) -> dict:
    """Step 1: Deterministic geo matching with LLM fallback."""
    from sqlalchemy import text

    start = time.time()

    # --- Load reference data ---
    countries_rows = db.execute(text("SELECT id, name, code FROM countries")).fetchall()
    ports_rows = db.execute(text("SELECT id, name, code, country_id FROM ports")).fetchall()

    countries = [{"id": r[0], "name": r[1], "code": r[2]} for r in countries_rows]
    ports = [{"id": r[0], "name": r[1], "code": r[2], "country_id": r[3]} for r in ports_rows]

    # --- Extract delivery_date from metadata (standard key + extra_fields fallback) ---
    delivery_date = metadata.get("delivery_date")
    if delivery_date:
        delivery_date = str(delivery_date).strip() or None
    if not delivery_date:
        extra = metadata.get("extra_fields", {})
        if isinstance(extra, dict):
            for fallback_key in ("loading_date", "deliver_on_date"):
                val = extra.get(fallback_key)
                if val and str(val).strip():
                    delivery_date = str(val).strip()
                    break

    # --- Build searchable text from all metadata values ---
    meta_text = " ".join(str(v) for v in metadata.values() if v).upper()

    extra = metadata.get("extra_fields", {}) if isinstance(metadata.get("extra_fields"), dict) else {}

    # --- Try deterministic port matching ---
    matched_port = None
    # Priority 1: destination_port (standard key) against port names
    dest_port = (metadata.get("destination_port") or "").strip().upper()
    if dest_port:
        for p in ports:
            port_name_upper = (p["name"] or "").upper()
            if dest_port and port_name_upper and (dest_port in port_name_upper or port_name_upper in dest_port):
                matched_port = p
                break

    # Priority 2: extra_fields.port_code against port codes
    if not matched_port:
        port_code = (extra.get("port_code") or "").strip().upper()
        if port_code:
            for p in ports:
                if p["code"] and p["code"].upper() == port_code:
                    matched_port = p
                    break

    # Priority 3: scan ALL metadata values for port name substring
    if not matched_port:
        for p in ports:
            port_name_upper = (p["name"] or "").upper()
            if port_name_upper and len(port_name_upper) >= 4 and port_name_upper in meta_text:
                matched_port = p
                break

    # --- Try deterministic country matching ---
    matched_country = None
    if matched_port:
        # Country is determined by port
        for c in countries:
            if c["id"] == matched_port["country_id"]:
                matched_country = c
                break
    else:
        # Try currency → country mapping (standard key)
        currency_val = (metadata.get("currency") or "").strip()
        if currency_val:
            country_code = _CURRENCY_TO_COUNTRY.get(currency_val) or _CURRENCY_TO_COUNTRY.get(currency_val.upper())
            if country_code:
                for c in countries:
                    if c["code"] == country_code:
                        matched_country = c
                        break

        # Try country name in metadata text
        if not matched_country:
            for c in countries:
                c_name = (c["name"] or "").upper()
                if c_name and len(c_name) >= 4 and c_name in meta_text:
                    matched_country = c
                    break

    country_id = matched_country["id"] if matched_country else None
    port_id = matched_port["id"] if matched_port else None

    elapsed = time.time() - start
    logger.info("Order %d: Step 1 — deterministic geo done in %.3fs (country_id=%s, port_id=%s)",
                order_id, elapsed, country_id, port_id)

    # --- Fallback: single LLM call if we couldn't determine both ---
    if country_id is None or port_id is None:
        logger.info("Order %d: Step 1 — deterministic match incomplete, falling back to LLM", order_id)
        fallback = _geo_llm_fallback(order_id, metadata, countries, ports)
        if country_id is None:
            country_id = fallback.get("country_id")
        if port_id is None:
            port_id = fallback.get("port_id")
        if delivery_date is None:
            delivery_date = fallback.get("delivery_date")

    return {"country_id": country_id, "port_id": port_id, "delivery_date": delivery_date}


def _geo_llm_fallback(order_id: int, metadata: dict, countries: list, ports: list) -> dict:
    """Single Gemini call to identify country/port when code matching fails."""
    from services.pdf_analyzer import _get_model, _parse_json_response

    countries_str = ", ".join(f'{c["id"]}={c["name"]}({c["code"]})' for c in countries)
    ports_str = ", ".join(f'{p["id"]}={p["name"]}(country_id={p["country_id"]})' for p in ports)

    prompt = f"""根据以下订单元数据，从可选列表中确定 country_id 和 port_id。

## 元数据
{json.dumps(metadata, ensure_ascii=False)}

## 可选国家
{countries_str}

## 可选港口
{ports_str}

返回纯 JSON（不要 markdown 代码块）：
{{"country_id": N或null, "port_id": N或null, "delivery_date": "日期或null"}}"""

    start = time.time()
    try:
        model = _get_model()
        response = model.generate_content([prompt])
        result = _parse_json_response(response.text.strip())
        elapsed = time.time() - start
        logger.info("Order %d: Step 1 — LLM fallback done in %.1fs", order_id, elapsed)
        return result
    except Exception as e:
        elapsed = time.time() - start
        logger.warning("Order %d: Step 1 — LLM fallback failed in %.1fs: %s", order_id, elapsed, e)
        return {}


def _batch_match(products: list[dict], db, country_id, port_id, delivery_date=None) -> tuple[list, list]:
    """Step 2: Code-based batch matching (0 LLM calls)."""
    from services.tools.product_matching import _match_products_against_db

    results = _match_products_against_db(products, db, country_id, port_id, delivery_date)
    ambiguous = [r for r in results if r["match_status"] in ("possible_match", "not_matched")]
    return results, ambiguous


def _refine_with_llm(order_id: int, ambiguous: list[dict], db, country_id, port_id, delivery_date=None) -> list:
    """Step 3: Single Gemini call to refine ambiguous matches. No Agent loop."""
    import re
    from models import ProductReadOnly
    from sqlalchemy import or_
    from services.pdf_analyzer import _get_model

    # Pre-fetch candidate products from DB
    query = db.query(ProductReadOnly).filter(ProductReadOnly.status == True)
    if country_id:
        query = query.filter(ProductReadOnly.country_id == country_id)
    if port_id:
        query = query.filter(ProductReadOnly.port_id == port_id)
    if delivery_date:
        query = query.filter(
            or_(ProductReadOnly.effective_from.is_(None),
                ProductReadOnly.effective_from <= delivery_date)
        ).filter(
            or_(ProductReadOnly.effective_to.is_(None),
                ProductReadOnly.effective_to >= delivery_date)
        )
    db_products = query.all()

    # Build candidate list for the prompt (compact format)
    candidates = []
    for p in db_products:
        candidates.append({
            "id": p.id,
            "code": p.code or "",
            "en": p.product_name_en or "",
            "jp": p.product_name_jp or "",
            "price": float(p.price) if p.price else None,
            "currency": p.currency,
            "supplier_id": p.supplier_id,
            "category_id": p.category_id,
            "pack_size": p.pack_size,
            "unit": p.unit,
        })

    # Build compact ambiguous items
    items = []
    for r in ambiguous:
        item = {"code": r.get("product_code", ""), "name": r.get("product_name", "")}
        if r.get("matched_product"):
            item["best_id"] = r["matched_product"].get("id")
        items.append(item)

    prompt = f"""你是产品匹配专家。以下 {len(items)} 个订单产品未能自动匹配，请判断每个是否对应数据库中的某个产品。

## 订单中的待匹配产品
{json.dumps(items, ensure_ascii=False)}

## 数据库候选产品（共 {len(candidates)} 个，同一 country/port）
{json.dumps(candidates, ensure_ascii=False)}

## 规则
- 产品名称语义相似（考虑缩写、多语言、同义词）且确信 = matched
- 有一定相似但不确定 = possible_match
- 完全不同 = not_matched

返回纯 JSON 数组（不要 markdown 代码块），每个元素：
{{"product_code":"订单代码","product_name":"订单名","match_status":"matched|possible_match|not_matched","match_reason":"原因","matched_product":{{"id":N,"code":"...","product_name_en":"...","product_name_jp":"...","price":N,"currency":"...","supplier_id":N,"category_id":N,"pack_size":"...","unit":"..."}}}}
未匹配到的 matched_product 设为 null。"""

    start = time.time()
    try:
        model = _get_model()
        response = model.generate_content([prompt])
        elapsed = time.time() - start
        logger.info("Order %d: Step 3 — single LLM call done in %.1fs", order_id, elapsed)

        # Parse response — handle both raw JSON and markdown-wrapped
        text = response.text.strip()
        m = re.search(r"```(?:json)?\s*([\[\{].*?[\]\}])\s*```", text, re.DOTALL)
        if m:
            text = m.group(1)
        result = json.loads(text)
        if isinstance(result, list):
            return result
        if isinstance(result, dict) and "match_results" in result:
            return result["match_results"]
        return [result] if isinstance(result, dict) else []
    except Exception as e:
        elapsed = time.time() - start
        logger.warning("Order %d: Step 3 — LLM refine failed in %.1fs: %s", order_id, elapsed, e)
        return []


def run_agent_matching(order_id: int, extracted_data: dict, db) -> dict:
    """3-step matching: deterministic geo → batch code match → single LLM refine."""
    metadata = extracted_data.get("order_metadata", {})
    products = extracted_data.get("products", [])

    logger.info("Order %d: starting optimized matching (%d products)", order_id, len(products))
    total_start = time.time()

    # Step 1: Deterministic geo matching + LLM fallback
    geo = _resolve_geo(order_id, metadata, db)
    country_id = geo.get("country_id")
    port_id = geo.get("port_id")
    delivery_date = geo.get("delivery_date")
    logger.info("Order %d: geo result — country_id=%s, port_id=%s, delivery_date=%s",
                order_id, country_id, port_id, delivery_date)

    # Parse delivery_date for effective_date filtering
    delivery_date_dt = _parse_delivery_date(delivery_date)

    # No delivery_date → skip matching
    if not delivery_date_dt:
        logger.warning("Order %d: no delivery_date found, skipping matching", order_id)
        return {
            "country_id": country_id,
            "port_id": port_id,
            "delivery_date": delivery_date,
            "match_results": None,
            "statistics": None,
            "skipped_reason": "missing_delivery_date",
        }

    # Step 2: Code-based batch matching (~0.4s, 0 LLM calls)
    step2_start = time.time()
    all_results, ambiguous = _batch_match(products, db, country_id, port_id, delivery_date_dt)
    step2_elapsed = time.time() - step2_start
    logger.info("Order %d: batch match done in %.1fs — %d matched, %d ambiguous",
                order_id, step2_elapsed,
                len(all_results) - len(ambiguous), len(ambiguous))

    # Step 3: Single LLM call for ambiguous items (optional)
    if 0 < len(ambiguous) <= 20:
        refined = _refine_with_llm(order_id, ambiguous, db, country_id, port_id, delivery_date_dt)
        if refined:
            refined_lookup = {}
            for r in refined:
                key = (r.get("product_code", ""), r.get("product_name", ""))
                refined_lookup[key] = r

            for i, result in enumerate(all_results):
                key = (result.get("product_code", ""), result.get("product_name", ""))
                if key in refined_lookup:
                    ref = refined_lookup[key]
                    # Only upgrade match status
                    if ref.get("match_status") == "matched" and result["match_status"] != "matched":
                        all_results[i]["match_status"] = ref["match_status"]
                        all_results[i]["match_reason"] = ref.get("match_reason", "")
                        if ref.get("matched_product"):
                            all_results[i]["matched_product"] = ref["matched_product"]
                    elif ref.get("match_status") == "possible_match" and result["match_status"] == "not_matched":
                        all_results[i]["match_status"] = ref["match_status"]
                        all_results[i]["match_reason"] = ref.get("match_reason", "")
                        if ref.get("matched_product"):
                            all_results[i]["matched_product"] = ref["matched_product"]
    elif len(ambiguous) > 20:
        logger.info("Order %d: skipping LLM refine — %d ambiguous items (too many)", order_id, len(ambiguous))

    # Compute statistics
    matched = sum(1 for r in all_results if r.get("match_status") == "matched")
    possible = sum(1 for r in all_results if r.get("match_status") == "possible_match")
    not_matched = sum(1 for r in all_results if r.get("match_status") == "not_matched")
    total = len(all_results)

    total_elapsed = time.time() - total_start
    logger.info("Order %d: matching complete in %.1fs — %d/%d matched (%.1f%%)",
                order_id, total_elapsed, matched, total,
                round(matched / total * 100, 1) if total else 0)

    return {
        "country_id": country_id,
        "port_id": port_id,
        "delivery_date": delivery_date,
        "match_results": all_results,
        "statistics": {
            "total": total,
            "matched": matched,
            "possible_match": possible,
            "not_matched": not_matched,
            "match_rate": round(matched / total * 100, 1) if total else 0,
        },
    }


# ─── Template-Guided Extraction ──────────────────────────────────

def _template_guided_extract(file_bytes: bytes, file_type: str, template, db) -> dict:
    """Template-guided extraction. Excel with full column_mapping skips LLM entirely."""
    from services.template_matcher import build_guided_prompt, extract_excel_deterministic

    try:
        # Excel + complete column_mapping → 0 LLM deterministic extraction
        if file_type != "pdf" and template.column_mapping:
            logger.info("Template %d: using deterministic Excel extraction", template.id)
            return extract_excel_deterministic(file_bytes, template)

        # PDF (or Excel without column_mapping) → LLM with enhanced prompt
        field_defs = None
        if template.field_schema_id:
            from models import FieldDefinition
            field_defs = db.query(FieldDefinition).filter(
                FieldDefinition.schema_id == template.field_schema_id
            ).order_by(FieldDefinition.sort_order).all()

        prompt = build_guided_prompt(template, field_defs)

        if file_type == "pdf":
            from services.pdf_analyzer import _get_model, _pdf_bytes_to_images, _parse_json_response

            images = _pdf_bytes_to_images(file_bytes)
            model = _get_model()
            response = model.generate_content([prompt] + images)
            result = _parse_json_response(response.text.strip())
        else:
            # Excel without column_mapping but with layout info
            from services.pdf_analyzer import _get_model, _parse_json_response

            text = _excel_to_text(file_bytes)
            model = _get_model()
            response = model.generate_content([prompt + "\n\n## Excel 内容\n" + text[:12000]])
            result = _parse_json_response(response.text.strip())

        return {
            "order_metadata": normalize_metadata(result.get("order_metadata", {})),
            "products": result.get("products", []),
            "extraction_method": "template_guided",
            "template_id": template.id,
        }
    except Exception as e:
        logger.warning("Template-guided extraction failed, fallback to generic: %s", e)
        return vision_extract(file_bytes, file_type)


def _excel_to_text(file_bytes: bytes) -> str:
    """Convert Excel to text for LLM prompt (reuses openpyxl logic)."""
    import io
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(cells):
                parts.append(" | ".join(cells))
    return "\n".join(parts)


# ─── Main Process ────────────────────────────────────────────────

def process_order(order_id: int, file_bytes: bytes) -> None:
    """Background task: template_match → extract → agent_matching."""
    db = SessionLocal()
    try:
        order = db.query(Order).get(order_id)
        if not order:
            logger.error("Order %d not found", order_id)
            return

        # Step 0: Template matching (0 LLM — keyword-based)
        template = None
        try:
            from services.template_matcher import find_matching_template, get_scannable_text

            scannable = get_scannable_text(file_bytes, order.file_type)
            if scannable:
                template, match_method = find_matching_template(scannable, db)
                if template:
                    order.template_id = template.id
                    order.template_match_method = match_method
                    db.commit()
                    logger.info("Order %d: matched template '%s' (id=%d) via %s",
                                order_id, template.name, template.id, match_method)
        except Exception as e:
            logger.warning("Order %d: template matching failed (non-fatal): %s", order_id, e)

        # Step 1: Extraction (template-guided or generic)
        order.status = "extracting"
        db.commit()
        logger.info("Order %d: starting extraction (file_type=%s, template=%s)",
                     order_id, order.file_type, template.name if template else "none")

        if template:
            extracted = _template_guided_extract(file_bytes, order.file_type, template, db)
        else:
            extracted = vision_extract(file_bytes, order.file_type)

        order.extraction_data = extracted
        order.order_metadata = extracted.get("order_metadata")
        order.products = extracted.get("products")
        order.product_count = len(extracted.get("products", []))
        total_amount = extracted.get("order_metadata", {}).get("total_amount")
        if total_amount is not None:
            try:
                order.total_amount = float(total_amount)
            except (ValueError, TypeError):
                pass
        db.commit()
        logger.info("Order %d: extraction done — %d products found", order_id, order.product_count)

        # Step 2: Agent-based smart matching
        order.status = "matching"
        db.commit()
        logger.info("Order %d: starting agent matching", order_id)

        match_result = run_agent_matching(order_id, extracted, db)

        order.match_results = match_result.get("match_results")
        order.match_statistics = match_result.get("statistics")
        order.country_id = match_result.get("country_id")
        order.port_id = match_result.get("port_id")
        order.delivery_date = match_result.get("delivery_date")

        if match_result.get("skipped_reason") == "missing_delivery_date":
            order.status = "ready"
            order.processing_error = "缺少交货日期(delivery_date)，请编辑订单元数据补充后重新匹配"
        else:
            order.status = "ready"
            order.processing_error = None

            # Auto-run financial analysis
            if order.match_results:
                try:
                    order.financial_data = run_financial_analysis(order)
                except Exception as e:
                    logger.warning("Order %d: financial analysis failed: %s", order_id, str(e))

            # Auto-run delivery environment (if port + delivery_date available)
            if order.port_id and order.delivery_date:
                try:
                    from services.weather_service import fetch_delivery_environment
                    from models import Port, Country
                    port = db.query(Port).get(order.port_id)
                    country = db.query(Country).get(port.country_id) if port and port.country_id else None
                    if port and country:
                        order.delivery_environment = fetch_delivery_environment(
                            port.name, country.name, order.delivery_date, db
                        )
                except Exception as e:
                    logger.warning("Order %d: delivery environment fetch failed: %s", order_id, str(e))

        order.processed_at = datetime.utcnow()
        db.commit()
        logger.info(
            "Order %d: processing complete — %s",
            order_id,
            json.dumps(match_result.get("statistics", {}), ensure_ascii=False),
        )

    except Exception as e:
        logger.error("Order %d processing failed: %s", order_id, str(e), exc_info=True)
        db.rollback()
        try:
            order = db.query(Order).get(order_id)
            if order:
                order.status = "error"
                order.processing_error = str(e)
                db.commit()
        except Exception:
            db.rollback()
    finally:
        db.close()


# ─── Preserved Functions (unchanged) ────────────────────────────

def run_anomaly_check(order: Order) -> dict:
    """Run anomaly detection on an order's data. Returns anomaly_data dict."""
    match_results = order.match_results or []
    products = order.products or []
    order_meta = order.order_metadata or {}
    match_stats = order.match_statistics or {}

    price_threshold = 0.20
    price_anomalies = []
    quantity_anomalies = []
    completeness_issues = []

    # Price anomalies
    for item in match_results:
        matched = item.get("matched_product")
        if not matched:
            continue
        order_price = item.get("unit_price")
        db_price = matched.get("price")
        if order_price is not None and db_price is not None and db_price > 0:
            try:
                order_price = float(order_price)
                db_price = float(db_price)
                deviation = abs(order_price - db_price) / db_price
                if deviation > price_threshold:
                    direction = "高于" if order_price > db_price else "低于"
                    price_anomalies.append({
                        "type": "price",
                        "product_name": item.get("product_name", ""),
                        "product_code": item.get("product_code", ""),
                        "order_value": order_price,
                        "db_value": db_price,
                        "deviation": round(deviation * 100, 1),
                        "description": f"{item.get('product_name', '')}: 订单价格 {order_price} {direction}数据库价格 {db_price} ({round(deviation * 100, 1)}%)",
                    })
            except (ValueError, TypeError):
                pass

    # Quantity anomalies
    for p in products:
        qty = p.get("quantity")
        if qty is None:
            quantity_anomalies.append({
                "type": "quantity",
                "product_name": p.get("product_name", ""),
                "issue": "missing",
                "description": f"{p.get('product_name', '')}: 数量缺失",
            })
        else:
            try:
                qty = float(qty)
                if qty <= 0:
                    quantity_anomalies.append({
                        "type": "quantity",
                        "product_name": p.get("product_name", ""),
                        "value": qty,
                        "issue": "non_positive",
                        "description": f"{p.get('product_name', '')}: 数量为 {qty} (非正数)",
                    })
                elif qty > 10000:
                    quantity_anomalies.append({
                        "type": "quantity",
                        "product_name": p.get("product_name", ""),
                        "value": qty,
                        "issue": "very_large",
                        "description": f"{p.get('product_name', '')}: 数量 {qty} 异常大",
                    })
            except (ValueError, TypeError):
                quantity_anomalies.append({
                    "type": "quantity",
                    "product_name": p.get("product_name", ""),
                    "issue": "invalid",
                    "description": f"{p.get('product_name', '')}: 数量格式无效 ({qty})",
                })

    # Completeness checks
    if not order_meta.get("po_number"):
        completeness_issues.append("缺少 PO 号码")
    if not order_meta.get("delivery_date"):
        completeness_issues.append("缺少交货日期")
    if not order_meta.get("ship_name"):
        completeness_issues.append("缺少船名")
    if not products:
        completeness_issues.append("没有产品数据")
    else:
        missing_names = sum(1 for p in products if not p.get("product_name"))
        if missing_names:
            completeness_issues.append(f"{missing_names} 个产品缺少名称")
    unmatched = match_stats.get("not_matched", 0)
    if unmatched > 0:
        completeness_issues.append(f"{unmatched} 个产品未匹配")

    total = len(price_anomalies) + len(quantity_anomalies) + len(completeness_issues)

    return {
        "total_anomalies": total,
        "price_anomalies": price_anomalies,
        "quantity_anomalies": quantity_anomalies,
        "completeness_issues": completeness_issues,
    }


def run_financial_analysis(order: Order) -> dict:
    """Run financial analysis on an order's matched products. Returns financial_data dict.

    Calculates per-product revenue/cost/profit/margin, aggregated by supplier and category.
    Only processes products with match_status == "matched" and valid prices.
    """
    match_results = order.match_results or []
    order_meta = order.order_metadata or {}
    order_currency = (order_meta.get("currency") or "").strip().upper()

    product_analyses = []
    warnings = []
    skipped_unmatched = 0
    skipped_currency_mismatch = 0
    skipped_missing_price = 0

    # Aggregation accumulators
    supplier_agg: dict[int, dict] = {}  # supplier_id -> {revenue, cost, profit, count}
    category_agg: dict[int, dict] = {}  # category_id -> {revenue, cost, profit, count}

    for item in match_results:
        if item.get("match_status") != "matched":
            skipped_unmatched += 1
            continue

        matched = item.get("matched_product")
        if not matched:
            skipped_unmatched += 1
            continue

        product_name = item.get("product_name", "")
        product_code = item.get("product_code", "")

        # Parse prices
        order_price_raw = item.get("unit_price")
        supplier_price_raw = matched.get("price")

        try:
            order_price = float(order_price_raw) if order_price_raw is not None else None
        except (ValueError, TypeError):
            order_price = None

        try:
            supplier_price = float(supplier_price_raw) if supplier_price_raw is not None else None
        except (ValueError, TypeError):
            supplier_price = None

        if order_price is None or supplier_price is None:
            skipped_missing_price += 1
            warnings.append({
                "type": "missing_price",
                "product_name": product_name,
                "product_code": product_code,
                "description": f"{product_name}: 缺少{'订单价格' if order_price is None else '供应商价格'}",
            })
            continue

        # Currency check
        product_currency = (matched.get("currency") or "").strip().upper()
        if order_currency and product_currency and order_currency != product_currency:
            skipped_currency_mismatch += 1
            warnings.append({
                "type": "currency_mismatch",
                "product_name": product_name,
                "product_code": product_code,
                "order_currency": order_currency,
                "product_currency": product_currency,
                "description": f"{product_name}: 订单币种({order_currency})与供应商币种({product_currency})不一致",
            })
            continue

        # Parse quantity (default 1.0)
        try:
            quantity = float(item.get("quantity") or 1.0)
            if quantity <= 0:
                quantity = 1.0
        except (ValueError, TypeError):
            quantity = 1.0

        revenue = round(order_price * quantity, 2)
        cost = round(supplier_price * quantity, 2)
        profit = round(revenue - cost, 2)
        margin = round((profit / revenue) * 100, 1) if revenue != 0 else 0.0

        currency = order_currency or product_currency or ""

        supplier_id = matched.get("supplier_id")
        category_id = matched.get("category_id")

        product_analyses.append({
            "product_name": product_name,
            "product_code": product_code,
            "order_price": order_price,
            "supplier_price": supplier_price,
            "quantity": quantity,
            "revenue": revenue,
            "cost": cost,
            "profit": profit,
            "margin": margin,
            "currency": currency,
            "supplier_id": supplier_id,
            "category_id": category_id,
        })

        # Negative margin warning
        if margin < 0:
            warnings.append({
                "type": "negative_margin",
                "product_name": product_name,
                "product_code": product_code,
                "margin": margin,
                "order_price": order_price,
                "supplier_price": supplier_price,
                "description": f"{product_name}: 利润率为 {margin}%（卖价 {order_price} < 成本 {supplier_price}）",
            })

        # Aggregate by supplier
        if supplier_id is not None:
            if supplier_id not in supplier_agg:
                supplier_agg[supplier_id] = {"revenue": 0, "cost": 0, "profit": 0, "count": 0}
            agg = supplier_agg[supplier_id]
            agg["revenue"] += revenue
            agg["cost"] += cost
            agg["profit"] += profit
            agg["count"] += 1

        # Aggregate by category
        if category_id is not None:
            if category_id not in category_agg:
                category_agg[category_id] = {"revenue": 0, "cost": 0, "profit": 0, "count": 0}
            agg = category_agg[category_id]
            agg["revenue"] += revenue
            agg["cost"] += cost
            agg["profit"] += profit
            agg["count"] += 1

    # Build breakdowns
    total_revenue = round(sum(p["revenue"] for p in product_analyses), 2)
    total_cost = round(sum(p["cost"] for p in product_analyses), 2)
    total_profit = round(total_revenue - total_cost, 2)
    overall_margin = round((total_profit / total_revenue) * 100, 1) if total_revenue != 0 else 0.0

    # Resolve supplier/category names from DB
    supplier_names: dict[int, str] = {}
    category_names: dict[int, str] = {}
    if supplier_agg or category_agg:
        from sqlalchemy import text as sql_text
        name_db = SessionLocal()
        try:
            if supplier_agg:
                ids = list(supplier_agg.keys())
                rows = name_db.execute(
                    sql_text("SELECT id, name FROM suppliers WHERE id = ANY(:ids)"),
                    {"ids": ids},
                ).fetchall()
                supplier_names = {r[0]: r[1] for r in rows}
            if category_agg:
                ids = list(category_agg.keys())
                rows = name_db.execute(
                    sql_text("SELECT id, name FROM categories WHERE id = ANY(:ids)"),
                    {"ids": ids},
                ).fetchall()
                category_names = {r[0]: r[1] for r in rows}
        except Exception as e:
            logger.warning("Failed to resolve supplier/category names: %s", e)
        finally:
            name_db.close()

    supplier_breakdown = []
    for sid, agg in supplier_agg.items():
        rev = round(agg["revenue"], 2)
        cst = round(agg["cost"], 2)
        prf = round(agg["profit"], 2)
        supplier_breakdown.append({
            "supplier_id": sid,
            "supplier_name": supplier_names.get(sid, f"供应商 #{sid}"),
            "revenue": rev,
            "cost": cst,
            "profit": prf,
            "margin": round((prf / rev) * 100, 1) if rev != 0 else 0.0,
            "product_count": agg["count"],
        })

    category_breakdown = []
    for cid, agg in category_agg.items():
        rev = round(agg["revenue"], 2)
        cst = round(agg["cost"], 2)
        prf = round(agg["profit"], 2)
        category_breakdown.append({
            "category_id": cid,
            "category_name": category_names.get(cid, f"品类 #{cid}"),
            "revenue": rev,
            "cost": cst,
            "profit": prf,
            "margin": round((prf / rev) * 100, 1) if rev != 0 else 0.0,
            "product_count": agg["count"],
        })

    return {
        "summary": {
            "total_revenue": total_revenue,
            "total_cost": total_cost,
            "total_profit": total_profit,
            "overall_margin": overall_margin,
            "currency": order_currency or "",
            "analyzed_count": len(product_analyses),
            "skipped_unmatched": skipped_unmatched,
            "skipped_currency_mismatch": skipped_currency_mismatch,
            "skipped_missing_price": skipped_missing_price,
            "total_products": len(match_results),
        },
        "product_analyses": product_analyses,
        "supplier_breakdown": supplier_breakdown,
        "category_breakdown": category_breakdown,
        "warnings": warnings,
    }


