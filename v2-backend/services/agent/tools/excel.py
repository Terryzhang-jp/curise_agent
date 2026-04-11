"""
Excel modification tool — read, write, and format Excel files in workspace.

Eliminates the need for bash + openpyxl scripts for simple modifications
like changing tax rates, number formats, or cell values.

Design: Single tool with action parameter (Anthropic recommendation:
"fewer tools with clear actions" > "many single-purpose tools")
"""

from __future__ import annotations

import json
import logging
import os
import re

from services.tools.registry_loader import ToolMetaInfo

logger = logging.getLogger(__name__)

TOOL_META = {
    "modify_excel": ToolMetaInfo(
        display_name="修改 Excel",
        group="utility",
        description="直接读取或修改工作目录中的 Excel 文件（改值、改公式、改格式）",
        prompt_description="读取/修改 Excel 文件（改税率、改格式等，无需 bash）",
        summary="修改 Excel",
    ),
}


def register(registry, ctx=None):
    """Register the modify_excel tool."""

    @registry.tool(
        description=(
            "读取或修改工作目录中的 Excel 文件。支持 4 种操作：\n"
            "- read: 读取单元格值（如 read J22, L34）\n"
            "- write: 写入值或公式（如 write L34 =L33*0.10）\n"
            "- format: 设置数字格式（如 format J22:J71 0.0）\n"
            "- list: 列出工作目录中的所有 Excel 文件\n\n"
            "修改后自动保存为新文件（不覆盖原文件）。"
        ),
        parameters={
            "filename": {
                "type": "STRING",
                "description": "Excel 文件名（如 inquiry_xxx.xlsx）。list 操作可留空。",
            },
            "action": {
                "type": "STRING",
                "description": "操作类型：read / write / format / list",
            },
            "cells": {
                "type": "STRING",
                "description": (
                    "JSON 格式的操作内容。\n"
                    'read: ["L34", "J22", "A4"]\n'
                    'write: {"L34": "=L33*0.10", "A4": "新名称"}\n'
                    'format: {"J22:J71": "0.0", "L22:L71": "#,##0.00"}'
                ),
                "required": False,
            },
        },
        group="utility",
    )
    def modify_excel(filename: str = "", action: str = "list", cells: str = "") -> str:
        workspace = getattr(ctx, 'workspace_dir', None) if ctx else None

        # === LIST: show Excel files in workspace ===
        if action == "list":
            if not workspace or not os.path.isdir(workspace):
                return "工作目录为空或不存在。"
            xlsx_files = []
            for f in os.listdir(workspace):
                if f.endswith(('.xlsx', '.xls')):
                    size = os.path.getsize(os.path.join(workspace, f))
                    xlsx_files.append(f"{f} ({size // 1024}KB)")
            if not xlsx_files:
                return "工作目录中没有 Excel 文件。"
            return "Excel 文件:\n" + "\n".join(f"  - {f}" for f in xlsx_files)

        # Validate filename
        if not filename:
            return "Error: 请指定文件名。用 action=list 查看可用文件。"
        if ".." in filename or "/" in filename:
            return "Error: 文件名不允许包含路径。"
        if not workspace:
            return "Error: 工作目录未配置。"

        filepath = os.path.join(workspace, filename)
        if not os.path.isfile(filepath):
            # Cloud Run fallback: download from Supabase Storage to workspace
            downloaded = _download_from_storage(filename, filepath)
            if not downloaded:
                candidates = [f for f in os.listdir(workspace) if f.endswith('.xlsx')]
                if candidates:
                    return f"Error: 文件 '{filename}' 不存在。可用文件: {', '.join(candidates)}"
                return f"Error: 文件 '{filename}' 不存在，工作目录中也没有其他 Excel 文件。"

        try:
            import openpyxl
            from openpyxl.cell.cell import MergedCell
        except ImportError:
            return "Error: openpyxl 未安装。"

        # === READ: read cell values ===
        if action == "read":
            try:
                cell_refs = json.loads(cells) if cells else []
            except json.JSONDecodeError:
                # Support comma-separated format: "L34, J22, A4"
                cell_refs = [c.strip() for c in cells.split(",") if c.strip()]

            if not cell_refs:
                return "Error: 请指定要读取的单元格，如 [\"L34\", \"J22\"]"

            wb = openpyxl.load_workbook(filepath, data_only=False)
            ws = wb.active
            results = []
            for ref in cell_refs:
                try:
                    cell = ws[ref]
                    if isinstance(cell, MergedCell):
                        results.append(f"{ref}: (merged cell)")
                    else:
                        val = cell.value
                        fmt = cell.number_format
                        results.append(f"{ref}: value={val}, format={fmt}")
                except Exception as e:
                    results.append(f"{ref}: Error - {e}")
            wb.close()
            return "\n".join(results)

        # === WRITE: set cell values or formulas ===
        if action == "write":
            try:
                changes = json.loads(cells) if cells else {}
            except json.JSONDecodeError:
                return 'Error: cells 格式错误。期望 JSON: {"L34": "=L33*0.10", "A4": "新值"}'

            if not changes:
                return "Error: 没有指定要写入的内容。"

            wb = openpyxl.load_workbook(filepath)
            ws = wb.active
            written = []
            for ref, value in changes.items():
                try:
                    cell = ws[ref]
                    if isinstance(cell, MergedCell):
                        written.append(f"{ref}: 跳过（合并单元格）")
                        continue
                    # Auto-detect numeric values
                    if isinstance(value, str) and not value.startswith("="):
                        try:
                            value = float(value)
                        except (ValueError, TypeError):
                            pass
                    cell.value = value
                    written.append(f"{ref}: ← {value}")
                except Exception as e:
                    written.append(f"{ref}: Error - {e}")

            # Save as new file (don't overwrite original)
            base, ext = os.path.splitext(filename)
            new_name = f"{base}_modified{ext}"
            new_path = os.path.join(workspace, new_name)
            wb.save(new_path)
            wb.close()

            # Cloud Run: also upload to Supabase Storage for persistence
            _upload_to_storage(new_name, new_path)

            summary = (
                "写入完成:\n" + "\n".join(f"  {w}" for w in written)
                + f"\n\n已保存为: {new_name}"
            )

            # Emit structured card so frontend shows download button
            session_id = getattr(ctx, 'pipeline_session_id', '') or getattr(ctx, 'session_id', '') if ctx else ''
            if session_id:
                card = json.dumps({
                    "card_type": "generated_file",
                    "filename": new_name,
                    "session_id": session_id,
                })
                summary += f"\n__STRUCTURED__\n{card}"

            return summary

        # === FORMAT: set number format ===
        if action == "format":
            try:
                formats = json.loads(cells) if cells else {}
            except json.JSONDecodeError:
                return 'Error: cells 格式错误。期望 JSON: {"J22:J71": "0.0", "L33": "#,##0"}'

            if not formats:
                return "Error: 没有指定要设置的格式。"

            wb = openpyxl.load_workbook(filepath)
            ws = wb.active
            formatted = []
            for ref, fmt in formats.items():
                try:
                    # Support range (e.g., "J22:J71") or single cell
                    if ":" in ref:
                        for row_cells in ws[ref]:
                            for cell in (row_cells if isinstance(row_cells, tuple) else [row_cells]):
                                if not isinstance(cell, MergedCell):
                                    cell.number_format = fmt
                        formatted.append(f"{ref}: format='{fmt}'")
                    else:
                        cell = ws[ref]
                        if not isinstance(cell, MergedCell):
                            cell.number_format = fmt
                            formatted.append(f"{ref}: format='{fmt}'")
                except Exception as e:
                    formatted.append(f"{ref}: Error - {e}")

            base, ext = os.path.splitext(filename)
            new_name = f"{base}_formatted{ext}"
            new_path = os.path.join(workspace, new_name)
            wb.save(new_path)
            wb.close()

            _upload_to_storage(new_name, new_path)

            summary_fmt = (
                "格式设置完成:\n" + "\n".join(f"  {f}" for f in formatted)
                + f"\n\n已保存为: {new_name}"
            )

            session_id = getattr(ctx, 'pipeline_session_id', '') or getattr(ctx, 'session_id', '') if ctx else ''
            if session_id:
                card = json.dumps({
                    "card_type": "generated_file",
                    "filename": new_name,
                    "session_id": session_id,
                })
                summary_fmt += f"\n__STRUCTURED__\n{card}"

            return summary_fmt

        return f"Error: 未知操作 '{action}'。支持: read / write / format / list"


def _upload_to_storage(filename: str, local_path: str):
    """Cloud Run: upload modified file to Supabase Storage for persistence + download."""
    try:
        from services.common.file_storage import storage
        with open(local_path, "rb") as f:
            content = f.read()
        storage.upload(
            "inquiries", filename, content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        logger.info("Uploaded modified file %s to storage", filename)
    except Exception as e:
        logger.debug("Storage upload failed for %s: %s", filename, e)


def _download_from_storage(filename: str, local_path: str) -> bool:
    """Cloud Run fallback: download file from Supabase Storage to local workspace."""
    try:
        from services.common.file_storage import storage
        for prefix in ("inquiries", "chat"):
            try:
                content = storage.download(f"{prefix}/{filename}")
                os.makedirs(os.path.dirname(local_path), exist_ok=True)
                with open(local_path, "wb") as f:
                    f.write(content)
                logger.info("Downloaded %s from storage to workspace", filename)
                return True
            except (FileNotFoundError, Exception):
                continue
    except Exception as e:
        logger.debug("Storage download failed for %s: %s", filename, e)
    return False
