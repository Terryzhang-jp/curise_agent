"""
Inquiry Generator v6.2 — Single LLM call + deterministic code enforcement.

Replaces the multi-turn ReActAgent approach (v5) with:
1. Single Gemini JSON-mode call for semantic field mapping (~2s)
2. Deterministic enforce_annotation() for format correctness
3. Code-driven workbook write + formula rebuild + save

Each supplier still gets its own thread via the orchestrator, but the per-supplier
logic is now a straight-line function (~6 stages) instead of an agent loop.
"""

from __future__ import annotations

import io
import json
import logging
import os
import re
import time
import uuid

from services.common.file_storage import storage as file_storage
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from typing import Any

from google import genai
from google.genai import types
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

from services.agent.config import load_api_key
from services.agent.stream_queue import get_or_create_cancel_event, push_event

logger = logging.getLogger(__name__)

# Matches external workbook references like [Book1.xlsx] or [RecoveredExternalLink1]
EXTERNAL_REF_RE = re.compile(r'\[.*?\]')


class InquiryCancelledError(RuntimeError):
    """Raised when an inquiry generation run is cancelled."""


def _ensure_not_cancelled(cancel_event) -> None:
    if cancel_event is not None and cancel_event.is_set():
        raise InquiryCancelledError("询价生成已取消")


def _sanitize_external_refs(ws) -> int:
    """Strip formulas that reference external workbooks. Returns count removed."""
    if ws is None:
        return 0
    count = 0
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                if EXTERNAL_REF_RE.search(cell.value):
                    logger.info("Sanitized external ref at %s: %s", cell.coordinate, cell.value[:80])
                    cell.value = None
                    count += 1
    if count:
        logger.info("Sanitized %d external reference formula(s) from template", count)
    return count


# ─── Pre-Analysis (no LLM, pure code) ────────────────────────

def run_inquiry_pre_analysis(order, db) -> dict:
    """Analyze order for inquiry generation — pure code, no LLM.

    Groups products by supplier, resolves templates, checks data completeness.
    Returns dict to be stored as order.inquiry_data.
    """
    from core.models import SupplierTemplate
    import sqlalchemy

    match_results = order.match_results or []
    order_meta = order.order_metadata or {}

    # Group products by supplier_id
    products_by_supplier: dict[int, list] = {}
    for p in match_results:
        mp = p.get("matched_product") or {}
        sid = mp.get("supplier_id")
        if not sid:
            continue
        products_by_supplier.setdefault(sid, []).append(p)

    # Load all templates
    all_templates = db.query(SupplierTemplate).all()

    # Load supplier info in one query
    supplier_ids = list(products_by_supplier.keys())
    supplier_rows = {}
    if supplier_ids:
        rows = db.execute(
            sqlalchemy.text(
                "SELECT id, name, contact, email, phone FROM suppliers WHERE id = ANY(:ids)"
            ),
            {"ids": supplier_ids},
        ).fetchall()
        for row in rows:
            supplier_rows[row[0]] = {
                "name": row[1],
                "contact": row[2],
                "email": row[3],
                "phone": row[4],
            }

    # Analyze each supplier
    suppliers = {}
    for sid, products in products_by_supplier.items():
        # Subtotal
        subtotal = 0.0
        for p in products:
            mp = p.get("matched_product") or {}
            qty = p.get("quantity") if p.get("quantity") is not None else 0
            val = p.get("unit_price")
            price = val if val is not None else (mp.get("price") if mp.get("price") is not None else 0)
            try:
                subtotal += float(qty) * float(price)
            except (TypeError, ValueError):
                pass

        # Template resolution
        template, method, candidates = select_template(sid, all_templates)
        template_info = None
        if template:
            template_info = {"id": template.id, "name": template.template_name, "method": method}
        elif candidates:
            template_info = {"method": method, "count": len(candidates)}
        else:
            template_info = {"method": "unavailable"}

        # Supplier data completeness
        info = supplier_rows.get(sid, {})
        missing_fields = []
        for field in ["contact", "email", "phone"]:
            if not info.get(field):
                missing_fields.append(field)

        suppliers[str(sid)] = {
            "status": "pending",
            "supplier_name": info.get("name") or f"供应商 #{sid}",
            "product_count": len(products),
            "subtotal": round(subtotal, 2),
            "currency": order_meta.get("currency", ""),
            "template": template_info,
            "missing_fields": missing_fields if missing_fields else None,
        }

    return {
        "status": "pre_analyzed",
        "supplier_count": len(suppliers),
        "total_products": len(match_results),
        "suppliers": suppliers,
    }


# ─── Pure Functions ───────────────────────────────────────────

def resolve_template(supplier_id: int, all_templates: list) -> tuple[Any | None, str, list]:
    """Resolve which template to use for a supplier.

    Returns (template, method, candidates):
    - Exact match found:  (template, "exact", [])
    - No exact match:     (None, "candidates", [{id, name, country_id}, ...])
    """
    all_templates = get_production_templates(all_templates)
    if not all_templates:
        return None, "unavailable", []

    # Step 1: exact binding — supplier_ids array
    for t in all_templates:
        if t.supplier_ids and supplier_id in t.supplier_ids:
            return t, "exact", []

    # Step 2: exact binding — legacy supplier_id field
    for t in all_templates:
        if t.supplier_id == supplier_id:
            return t, "exact", []

    # Step 3: no exact match → return candidate list
    candidates = []
    for t in all_templates:
        candidates.append({
            "id": t.id,
            "name": t.template_name,
            "country_id": t.country_id,
        })
    return None, "candidates", candidates


def template_has_zone_config(template: Any | None) -> bool:
    """Production templates must carry a root-level zone_config."""
    if not template:
        return False
    styles = getattr(template, "template_styles", None)
    return isinstance(styles, dict) and isinstance(styles.get("zones"), dict)


def get_production_templates(all_templates: list) -> list:
    """Only zone_config templates are allowed in production inquiry flow."""
    return [template for template in all_templates if template_has_zone_config(template)]


def select_template(
    supplier_id: int,
    all_templates: list,
    template_id_override: int | None = None,
) -> tuple[Any | None, str, list]:
    """Choose the production template for a supplier.

    Returns (template, selection_method, candidates). When no exact binding exists,
    the first production-ready candidate is chosen to preserve the current fallback
    behavior, but legacy templates are never considered.
    """
    production_templates = get_production_templates(all_templates)

    if template_id_override:
        override = next((t for t in all_templates if t.id == template_id_override), None)
        if not override:
            raise ValueError(f"模板 {template_id_override} 不存在")
        if not template_has_zone_config(override):
            raise ValueError(
                f"模板 {template_id_override} 已下架：当前只允许使用带 zone_config 的模板"
            )
        return override, "user_selected", []

    template, method, candidates = resolve_template(supplier_id, production_templates)
    if template:
        return template, method, candidates
    if candidates:
        first_candidate = next(
            (t for t in production_templates if t.id == candidates[0]["id"]),
            None,
        )
        if first_candidate:
            return first_candidate, "candidate_auto", candidates
    return None, "unavailable", []


def _try_parse_date(val: str) -> datetime | None:
    """Try to parse a date string from various formats.

    Supports: Reiwa (R8.02.23), YYYY/MM/DD, YYYY-MM-DD, Chinese 年月日,
    ISO format, English month names, etc.
    """
    if not val or not val.strip():
        return None

    # Try Reiwa format first: R8.02.23 or R8/02/23
    reiwa_match = re.match(r"R(\d+)[./](\d+)[./](\d+)", val.strip())
    if reiwa_match:
        year = int(reiwa_match.group(1)) + 2018
        month = int(reiwa_match.group(2))
        day = int(reiwa_match.group(3))
        try:
            return datetime(year, month, day)
        except ValueError:
            pass

    for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y.%m.%d", "%m/%d/%Y", "%d/%m/%Y",
                "%Y年%m月%d日", "%B %d, %Y", "%d %B %Y"):
        try:
            return datetime.strptime(val.strip(), fmt)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(val.strip())
    except (ValueError, TypeError):
        return None


def enforce_annotation(value: str, annotation: str) -> str:
    """Deterministic format enforcement based on annotation rules.

    Handles 10 patterns in priority order:
    1. 和暦 (Reiwa era dates)
    2. YYYY/MM/DD
    3. DD/MM/YYYY
    4. Decimal places (小数点N位)
    5. Integer only (整数のみ)
    6. Uppercase (大写/大文字のみ)
    7. Strip prefix (不含XX前缀)
    8. Remove hyphens (ハイフンなし)
    9. ISO 4217 currency code
    10. Max length (N文字以内)
    """
    if not value or not annotation:
        return value
    val = str(value).strip()

    # 1. 和暦 (Japanese era): "和暦表記" / "令和" → convert to R{year}.MM.DD
    if re.search(r"和暦|令和|Reiwa", annotation, re.IGNORECASE):
        parsed = _try_parse_date(val)
        if parsed:
            reiwa_year = parsed.year - 2018
            if reiwa_year > 0:
                return f"R{reiwa_year}.{parsed.month:02d}.{parsed.day:02d}"
        return val

    # 2. YYYY/MM/DD
    if re.search(r"YYYY[/\-.]MM[/\-.]DD", annotation, re.IGNORECASE):
        parsed = _try_parse_date(val)
        if parsed:
            return parsed.strftime("%Y/%m/%d")
        return val

    # 3. DD/MM/YYYY
    if re.search(r"DD[/\-.]MM[/\-.]YYYY", annotation, re.IGNORECASE):
        parsed = _try_parse_date(val)
        if parsed:
            return parsed.strftime("%d/%m/%Y")
        return val

    # 4. Decimal places: "小数点3位", "小数点后面两位", "3 decimal"
    decimal_match = re.search(
        r"小数点\s*(?:后面|まで)?\s*(\d+|[一二两三四五六七八九十]+)\s*位|(\d+)\s*decimal",
        annotation, re.IGNORECASE,
    )
    if decimal_match:
        raw = decimal_match.group(1) or decimal_match.group(2)
        cn_digits = {"一": 1, "二": 2, "两": 2, "三": 3, "四": 4, "五": 5}
        places = cn_digits.get(raw, None) if not raw.isdigit() else int(raw)
        if places is not None:
            try:
                return f"{float(val):.{places}f}"
            except (ValueError, TypeError):
                return val

    # 5. Integer only: "整数のみ" / "integer only" / "小数点不可"
    if re.search(r"整数のみ|integer only|小数点不可", annotation, re.IGNORECASE):
        try:
            return str(int(float(val)))
        except (ValueError, TypeError):
            return val

    # 6. Uppercase: "大写" / "uppercase" / "英語大文字のみ"
    if re.search(r"大写|uppercase|大文字のみ", annotation, re.IGNORECASE):
        return val.upper()

    # 7. Strip prefix: "不含ROL-前缀" / "strip prefix XXX"
    prefix_match = re.search(r"不含(.+?)前缀|strip prefix[: ]*(.+)", annotation, re.IGNORECASE)
    if prefix_match:
        prefix = (prefix_match.group(1) or prefix_match.group(2)).strip()
        if val.upper().startswith(prefix.upper()):
            return val[len(prefix):].strip()
        for sep in ("-", " ", "_"):
            pf = prefix + sep
            if val.upper().startswith(pf.upper()):
                return val[len(pf):].strip()

    # 8. Remove hyphens: "ハイフンなし"
    if re.search(r"ハイフンなし|no hyphen|remove hyphen", annotation, re.IGNORECASE):
        return val.replace("-", "").replace("\u2010", "").replace("\u2212", "")

    # 9. ISO currency code: "ISO 4217" / "3文字コード"
    if re.search(r"ISO\s*4217|3文字コード", annotation, re.IGNORECASE):
        return val.upper().strip()[:3]

    # 10. Max length: "XX文字以内" / "max XX chars"
    len_match = re.search(r"(\d+)\s*文字以内|max\s*(\d+)\s*char", annotation, re.IGNORECASE)
    if len_match:
        max_len = int(len_match.group(1) or len_match.group(2))
        if len(val) > max_len:
            return val[:max_len]

    return val


def _code_check(value, annotation: str) -> dict:
    """Code-level check of a cell value against its annotation.

    Returns {status: "pass"|"fail"|"unchecked", reason, suggestion?}
    """
    val_str = str(value).strip() if value is not None and value != "" else ""

    # 和暦 format check
    if re.search(r"和暦|令和", annotation, re.IGNORECASE):
        if not val_str:
            return {"status": "fail", "reason": "日期为空", "suggestion": "填写和暦格式日期 (例: R8.02.23)"}
        if re.match(r"^R\d+\.\d{2}\.\d{2}$", val_str):
            return {"status": "pass", "reason": "和暦格式正确"}
        return {"status": "fail", "reason": f"格式不符: {val_str}", "suggestion": "应为 R{年}.MM.DD 格式"}

    # YYYY/MM/DD format check
    date_match = re.search(r'YYYY[/\-.]MM[/\-.]DD', annotation, re.IGNORECASE)
    if date_match:
        if not val_str:
            return {"status": "fail", "reason": "日期为空", "suggestion": "填写 YYYY/MM/DD 格式日期"}
        if re.match(r'^\d{4}/\d{2}/\d{2}$', val_str):
            try:
                datetime.strptime(val_str, "%Y/%m/%d")
                return {"status": "pass", "reason": "日期格式正确"}
            except ValueError:
                return {"status": "fail", "reason": f"日期无效: {val_str}", "suggestion": "检查日期是否合法"}
        return {"status": "fail", "reason": f"格式不符: {val_str}", "suggestion": "应为 YYYY/MM/DD 格式"}

    # DD/MM/YYYY format check
    if re.search(r'DD[/\-.]MM[/\-.]YYYY', annotation, re.IGNORECASE):
        if not val_str:
            return {"status": "fail", "reason": "日期为空", "suggestion": "填写 DD/MM/YYYY 格式日期"}
        if re.match(r'^\d{2}/\d{2}/\d{4}$', val_str):
            return {"status": "pass", "reason": "日期格式正确"}
        return {"status": "fail", "reason": f"格式不符: {val_str}", "suggestion": "应为 DD/MM/YYYY 格式"}

    # Decimal places check (e.g. "小数点2位", "2 decimal places")
    decimal_match = re.search(r'小数点\s*(?:后面|まで)?\s*(\d+)\s*位|(\d+)\s*decimal', annotation, re.IGNORECASE)
    if decimal_match:
        required_places = int(decimal_match.group(1) or decimal_match.group(2))
        if not val_str:
            return {"status": "fail", "reason": "值为空", "suggestion": f"填写保留{required_places}位小数的数字"}
        dot_match = re.match(r'^-?\d+\.(\d+)$', val_str)
        if dot_match:
            actual_places = len(dot_match.group(1))
            if actual_places == required_places:
                return {"status": "pass", "reason": f"小数点{required_places}位正确"}
            return {
                "status": "fail",
                "reason": f"小数点{actual_places}位，需要{required_places}位",
                "suggestion": f"改为 {float(val_str):.{required_places}f}",
            }
        try:
            float(val_str)
            return {
                "status": "fail",
                "reason": "缺少小数点",
                "suggestion": f"改为 {float(val_str):.{required_places}f}",
            }
        except (ValueError, TypeError):
            return {"status": "fail", "reason": f"非数字: {val_str}", "suggestion": "应为数字"}

    # Integer only check
    if re.search(r"整数のみ|integer only|小数点不可", annotation, re.IGNORECASE):
        if not val_str:
            return {"status": "fail", "reason": "值为空", "suggestion": "填写整数"}
        try:
            f = float(val_str)
            if f == int(f):
                return {"status": "pass", "reason": "整数正确"}
            return {"status": "fail", "reason": f"非整数: {val_str}", "suggestion": str(int(f))}
        except (ValueError, TypeError):
            return {"status": "fail", "reason": f"非数字: {val_str}", "suggestion": "应为整数"}

    # Uppercase check
    if re.search(r"大写|uppercase|大文字のみ", annotation, re.IGNORECASE):
        if not val_str:
            return {"status": "unchecked"}
        if val_str == val_str.upper():
            return {"status": "pass", "reason": "大文字正確"}
        return {"status": "fail", "reason": f"非大文字: {val_str}", "suggestion": val_str.upper()}

    # Remove hyphens check
    if re.search(r"ハイフンなし|no hyphen", annotation, re.IGNORECASE):
        if not val_str:
            return {"status": "unchecked"}
        if "-" in val_str or "\u2010" in val_str or "\u2212" in val_str:
            return {"status": "fail", "reason": f"含ハイフン: {val_str}", "suggestion": val_str.replace("-", "").replace("\u2010", "").replace("\u2212", "")}
        return {"status": "pass", "reason": "ハイフンなし正確"}

    # Max length check
    len_match = re.search(r"(\d+)\s*文字以内|max\s*(\d+)\s*char", annotation, re.IGNORECASE)
    if len_match:
        max_len = int(len_match.group(1) or len_match.group(2))
        if not val_str:
            return {"status": "unchecked"}
        if len(val_str) <= max_len:
            return {"status": "pass", "reason": f"{len(val_str)}文字 <= {max_len}文字"}
        return {"status": "fail", "reason": f"{len(val_str)}文字 > {max_len}文字", "suggestion": val_str[:max_len]}

    # No matching pattern → unchecked
    return {"status": "unchecked"}


# ─── LLM Mapping Prompt ──────────────────────────────────────

MAPPING_PROMPT = """你是询价单填写专家。根据模板字段定义和订单数据，决定每个字段应该填入什么值。

## 模板字段（需要你填写的）
{fields_json}

## 订单数据
{order_data_json}

## 供应商信息
{supplier_json}

## 规则
1. **order 类字段**（ship_name, delivery_date, po_number, voyage 等）→ 从订单数据中取对应值。**如果数据中没有对应值，填空字符串 ""**（清除模板残留）
2. **supplier 类字段**（supplier_name, supplier_contact, supplier_tel, supplier_email, supplier_address）→ 从供应商信息中取值。**如果供应商信息中该字段为空字符串，必须填空字符串 ""**（清除模板残留）
3. **company 类字段**（company_name, company_address, company_tel, company_fax, company_email, company_zip_code）→ 填 null（保留模板原值，这是买方公司固定信息）
4. **formula 类字段** → 填 null（Excel 自动计算）
5. **delivery 类字段**（delivery_company_name, delivery_contact, delivery_time_notes）→ 从订单数据取值（如有），没有则填 ""（清除模板残留）。delivery_address → 从订单数据取值
6. **payment 类字段**（payment_date, payment_method）→ 从订单数据取值，没有则填 ""
7. 注意 annotation 中的格式要求，但代码会做最终格式强制，你只需填正确的原始值
8. 如果 annotation 要求特殊格式（和暦、去前缀等），你可以尝试转换，但不必完美（代码会修正）

## 特殊字段
- 如果某个字段 annotation 说"前面是XX 后面是YY"，组合两个值（缺失的部分省略，不要写 "null"）
- invoice_number 语义上等同于 po_number

## 语义映射参考
- delivery_date = deliver_on_date = 納期 = Delivery Date
- ship_name = vessel_name = 船名 = Vessel
- po_number = order_number = invoice_number = 注文番号
- destination_port = port_name = 納品先
- supplier_tel = phone, supplier_email = email
- remarks = 備考 = notes

## 输出 JSON
返回 key=单元格位置, value=字符串或 null：
- "" = 清空
- null = 保留模板原值
只返回 JSON。"""


# ─── Row formatting copy helper ────────────────────────────────

def _copy_row_formatting(ws, source_row: int, insert_start: int, num_rows: int) -> None:
    """Copy row height and cell styles from source_row to inserted rows.

    After ws.insert_rows(), newly inserted rows have no formatting.
    This copies the source row's height and each cell's style (border, font,
    alignment, fill, number_format) to make inserted rows visually consistent.
    """
    from copy import copy
    src_height = ws.row_dimensions[source_row].height
    # Get max column from the source row
    max_col = ws.max_column or 14  # fallback to N

    for offset in range(num_rows):
        target_row = insert_start + offset
        if src_height:
            ws.row_dimensions[target_row].height = src_height

        for col_idx in range(1, max_col + 1):
            src_cell = ws.cell(row=source_row, column=col_idx)
            tgt_cell = ws.cell(row=target_row, column=col_idx)
            if src_cell.has_style:
                tgt_cell.font = copy(src_cell.font)
                tgt_cell.border = copy(src_cell.border)
                tgt_cell.fill = copy(src_cell.fill)
                tgt_cell.number_format = src_cell.number_format
                tgt_cell.alignment = copy(src_cell.alignment)
                tgt_cell.protection = copy(src_cell.protection)


def _build_order_data_for_engine(
    order_id: int, order_meta: dict, supplier_id: int,
    products: list[dict], supplier_info: dict,
    company_info: dict | None = None,
    delivery_info: dict | None = None,
    _db=None,
) -> dict:
    """Build order_data dict in the format expected by template_engine.fill_template().

    Mirrors the structure from prepare_inquiry_workspace (services/tools/inquiry.py).
    company_info / delivery_info should be pre-fetched by the caller.
    """
    sid = str(supplier_id)
    currency = order_meta.get("currency") or "JPY"

    # Clean products for engine format
    cleaned = []
    for p in products:
        mp = p.get("matched_product") or {}
        cleaned.append({
            "product_code": mp.get("code") or p.get("product_code", ""),
            "product_name": p.get("product_name", ""),
            "product_name_jp": mp.get("product_name_jp", ""),
            "quantity": p.get("quantity"),
            "unit": mp.get("unit") or p.get("unit", ""),
            "unit_price": mp.get("price") if mp.get("price") is not None else p.get("unit_price"),
            "pack_size": mp.get("pack_size", ""),
            "currency": currency,
        })

    return {
        "order_id": order_id,
        "po_number": order_meta.get("po_number", ""),
        "ship_name": order_meta.get("ship_name", ""),
        "delivery_date": order_meta.get("delivery_date", ""),
        "order_date": order_meta.get("order_date", ""),
        "currency": currency,
        "destination_port": order_meta.get("destination_port") or order_meta.get("port_name", ""),
        "delivery_address": order_meta.get("delivery_address", ""),
        "voyage": order_meta.get("voyage", ""),
        "company": company_info or {},
        "delivery_location": delivery_info or {},
        "suppliers": {
            sid: {
                "supplier_name": supplier_info.get("name", ""),
                "supplier_info": supplier_info,
                "product_count": len(cleaned),
                "products": cleaned,
            }
        },
    }


# ─── Per-Supplier Generation (replaces Agent) ─────────────────

def _generate_single_supplier(
    order_id: int, order_meta: dict, supplier_id: int, products: list[dict],
    stream_key: str, overall_start: float,
    template_id_override: int | None = None,
    cancel_event=None,
) -> dict:
    """Generate inquiry for one supplier via single LLM call + deterministic code.

    Thread-safe — creates its own DB session.

    Returns {"file_info": {...}, "verify_results": [...], "error": str|None}
    """
    from core.database import SessionLocal
    from core.models import SupplierTemplate
    from services.excel.excel_writer import InquiryWorkbook
    from services.agent.stream_queue import push_event
    from sqlalchemy import text as sa_text

    upload_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "uploads")
    os.makedirs(upload_dir, exist_ok=True)
    _ensure_not_cancelled(cancel_event)

    # ── Stage 1: Data loading ──
    _db = SessionLocal()
    try:
        # Supplier info — query ALL fields that templates may reference
        row = _db.execute(
            sa_text(
                "SELECT name, contact, email, phone, fax, address, zip_code,"
                " default_payment_method, default_payment_terms"
                " FROM suppliers WHERE id = :sid"
            ),
            {"sid": supplier_id},
        ).fetchone()
        supplier_info = {
            "name": row[0] or "" if row else "",
            "contact": row[1] or "" if row else "",
            "email": row[2] or "" if row else "",
            "phone": row[3] or "" if row else "",
            "fax": row[4] or "" if row else "",
            "address": row[5] or "" if row else "",
            "zip_code": row[6] or "" if row else "",
            "default_payment_method": row[7] or "" if row else "",
            "default_payment_terms": row[8] or "" if row else "",
        }

        # Port/country enrichment
        enriched_meta = dict(order_meta)
        order_row = _db.execute(
            sa_text("SELECT port_id, country_id FROM v2_orders WHERE id = :oid"),
            {"oid": order_id},
        ).fetchone()
        if order_row:
            if order_row[0]:
                p_row = _db.execute(
                    sa_text("SELECT name, location, code FROM ports WHERE id = :pid"),
                    {"pid": order_row[0]},
                ).fetchone()
                if p_row:
                    enriched_meta.update({
                        "port_name": p_row[0] or "", "delivery_address": p_row[1] or "",
                        "port_code": p_row[2] or "",
                    })
            if order_row[1]:
                c_row = _db.execute(
                    sa_text("SELECT name, code FROM countries WHERE id = :cid"),
                    {"cid": order_row[1]},
                ).fetchone()
                if c_row:
                    enriched_meta.update({"country_name": c_row[0] or "", "country_code": c_row[1] or ""})

        # Pre-fetch company config + delivery location (before closing DB)
        _prefetched_company = {}
        _prefetched_delivery = {}
        try:
            from core.models import CompanyConfig, DeliveryLocation
            for c in _db.query(CompanyConfig).order_by(CompanyConfig.sort_order).all():
                _prefetched_company[c.key] = c.value
            loc = _db.query(DeliveryLocation).filter(DeliveryLocation.is_default == True).first()
            if loc:
                _prefetched_delivery = {
                    "name": loc.name, "address": loc.address,
                    "contact_person": loc.contact_person,
                    "contact_phone": loc.contact_phone,
                    "delivery_notes": loc.delivery_notes,
                    "ship_name_label": loc.ship_name_label,
                }
        except Exception:
            pass

        # ── Stage 2: Template resolution + parse ──
        all_templates = _db.query(SupplierTemplate).all()
    finally:
        _db.close()

    chosen_template, selection_method, _ = select_template(
        supplier_id,
        all_templates,
        template_id_override=template_id_override,
    )

    if not chosen_template:
        raise ValueError(
            f"供应商 {supplier_id} 没有可用的 zone_config 模板，询价单生成已阻断"
        )

    # Build workbook + fields
    wb = InquiryWorkbook()
    wb._generation_path = "generic"  # Default; overridden below

    # ── Fast path: deterministic engine (no LLM) ──
    # If template has zone_config in template_styles, use template_engine
    zone_config = None
    if chosen_template.template_styles and isinstance(chosen_template.template_styles, dict):
        if "zones" in chosen_template.template_styles:
            zone_config = chosen_template.template_styles

    if zone_config and chosen_template.template_file_url:
        try:
            from services.templates.template_engine_legacy import fill_template as engine_fill
            from services.templates.template_engine_legacy import verify_output as engine_verify

            # ── Feature flag: choose renderer ──
            # INQUIRY_RENDERER=compose (default) → new compose-from-scratch renderer
            # INQUIRY_RENDERER=fill              → legacy fill_template (kept as fallback)
            # The legacy path is preserved during the migration window so we can
            # roll back instantly if compose_render exposes any unforeseen issue.
            renderer = os.environ.get("INQUIRY_RENDERER", "compose").lower()

            engine_start = time.time()
            _ensure_not_cancelled(cancel_event)

            # Download template
            suffix = ".xls" if chosen_template.template_file_url.lower().endswith(".xls") else ".xlsx"
            template_bytes = file_storage.download(chosen_template.template_file_url)

            # Build order_data for engine (reuses prepare_inquiry_workspace data structure)
            engine_order_data = _build_order_data_for_engine(
                order_id, enriched_meta, supplier_id, products, supplier_info,
                company_info=_prefetched_company, delivery_info=_prefetched_delivery,
            )

            # Load user field overrides if any
            _fo_db = SessionLocal()
            try:
                from core.models import Order as _OrderModel
                _fo_order = _fo_db.query(_OrderModel).get(order_id)
                _fo_overrides = (
                    (_fo_order.inquiry_data or {})
                    .get("suppliers", {})
                    .get(str(supplier_id), {})
                    .get("field_overrides")
                ) if _fo_order else None
            finally:
                _fo_db.close()

            # ── Render via the selected backend ──
            if renderer == "fill":
                excel_bytes = engine_fill(
                    template_bytes, zone_config, engine_order_data, supplier_id,
                    field_overrides=_fo_overrides,
                )
                generation_path = "template_engine_fill"
            else:  # compose (default)
                from services.templates.template_engine import compose_render
                excel_bytes = compose_render(
                    template_bytes, zone_config, engine_order_data, supplier_id,
                )
                generation_path = "template_engine_compose"

            _ensure_not_cancelled(cancel_event)

            engine_elapsed = time.time() - engine_start
            logger.info(
                "Template engine [%s]: supplier %d, %d products in %.2fs (no LLM)",
                renderer, supplier_id, len(products), engine_elapsed,
            )

            if stream_key:
                push_event(stream_key, {
                    "type": "tool_call", "tool_name": "generate", "tool_label": "生成询价单",
                    "content": f"{renderer} 引擎: {chosen_template.template_name}, {len(products)} 产品, {engine_elapsed:.1f}s",
                    "elapsed_seconds": round(time.time() - overall_start, 1),
                    "supplier_id": supplier_id,
                })

            # Verify (same verify regardless of renderer)
            vr = engine_verify(excel_bytes, zone_config, engine_order_data, supplier_id)
            verify_results = _engine_verify_to_results(vr)
            _ensure_not_cancelled(cancel_event)

            # Save and return — deterministic path (0 LLM), whether it passed or failed.
            # Failed verification enters review/repair instead of silently switching paths.
            wb = InquiryWorkbook()
            wb._wb = load_workbook(io.BytesIO(excel_bytes))
            wb._ws = wb._wb.active
            wb._generation_path = generation_path
            saved = _save_workbook(
                wb, chosen_template, selection_method, order_meta, supplier_id,
                products, upload_dir, {}, stream_key, overall_start,
                verify_results=verify_results,
            )
            if not vr["ok"]:
                saved["status"] = "repair_required"
                saved["error"] = (
                    f"模板校验失败：{len(vr.get('errors', []))} 项问题，"
                    "请在 Review/Repair 中修正后重试"
                )
            return saved
        except Exception as e:
            logger.warning(
                "Template engine failed for supplier %d: %s",
                supplier_id, e, exc_info=True,
            )
            raise

    # Load template file from Supabase Storage
    template_file_path = None
    if chosen_template.template_file_url:
        try:
            suffix = ".xlsx"
            if chosen_template.template_file_url.lower().endswith(".xls"):
                suffix = ".xls"
            template_file_path = file_storage.download_to_temp(
                chosen_template.template_file_url, suffix=suffix
            )
        except FileNotFoundError:
            logger.warning("Template file not found in storage: %s", chosen_template.template_file_url)
            template_file_path = None

    if template_file_path:
        wb.load_template(template_file_path)
        # Layer 2.2: Sanitize — strip external workbook references to prevent #N/A
        _sanitize_external_refs(wb._ws)
        # Clean up temp file
        try:
            os.unlink(template_file_path)
        except OSError:
            pass
    else:
        wb.create_from_config()

    # Parse annotations
    field_positions = chosen_template.field_positions or {}
    fmm = chosen_template.field_mapping_metadata or {}
    annotations: dict[str, str] = {}
    if fmm.get("annotations"):
        annotations.update(fmm["annotations"])
    for item in fmm.get("items", []):
        if item.get("note") and item.get("position"):
            annotations[item["position"]] = item["note"]

    # Detect formula cells
    formula_cells: set[str] = set()
    template_formulas: dict[str, str] = {}
    if wb._ws:
        for row in wb._ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    # Layer 1: Skip external workbook references (already sanitized by Layer 2.2,
                    # but double-check in case load_workbook re-introduces them)
                    if EXTERNAL_REF_RE.search(cell.value):
                        continue
                    formula_cells.add(cell.coordinate)
                    template_formulas[cell.coordinate] = cell.value

    # Build fields list (excluding formula cells)
    fields = []
    for field_key, pos_info in field_positions.items():
        position = pos_info.get("position", "") if isinstance(pos_info, dict) else pos_info
        description = pos_info.get("description", "") if isinstance(pos_info, dict) else ""
        if not position or position in formula_cells:
            continue
        field = {"field_key": field_key, "position": position, "description": description}
        if position in annotations:
            field["annotation"] = annotations[position]
        fields.append(field)

    # SSE: emit "generate" step
    if stream_key:
        push_event(stream_key, {
            "type": "tool_call", "tool_name": "generate", "tool_label": "生成询价单",
            "content": f"模板: {chosen_template.template_name}, {len(fields)} 字段, {len(products)} 产品",
            "elapsed_seconds": round(time.time() - overall_start, 1),
            "supplier_id": supplier_id,
        })

    # ── Stage 3: Single LLM call for header field mapping ──
    wb._generation_path = "llm_mapping"
    cell_mapping = {}
    _ensure_not_cancelled(cancel_event)

    if fields:
        order_data = {k: str(v)[:200] for k, v in enriched_meta.items() if v}
        prompt = MAPPING_PROMPT.format(
            fields_json=json.dumps(fields, ensure_ascii=False, indent=2),
            order_data_json=json.dumps(order_data, ensure_ascii=False, indent=2),
            supplier_json=json.dumps(supplier_info, ensure_ascii=False, indent=2),
        )

        api_key = load_api_key("gemini")
        client = genai.Client(api_key=api_key)

        max_retries = 3
        for attempt in range(max_retries):
            try:
                _ensure_not_cancelled(cancel_event)
                llm_start = time.time()
                response = client.models.generate_content(
                    model="gemini-3-flash-preview",
                    contents=prompt,
                    config=types.GenerateContentConfig(
                        response_mime_type="application/json",
                        temperature=0.1,
                        max_output_tokens=5000,
                        thinking_config=types.ThinkingConfig(thinking_budget=2048),
                    ),
                )
                llm_elapsed = time.time() - llm_start
                logger.info("Inquiry v6.2: supplier %d LLM mapping in %.1fs", supplier_id, llm_elapsed)

                resp_text = getattr(response, "text", None)
                if not resp_text or not resp_text.strip():
                    raise ValueError("LLM returned empty response")

                parsed = json.loads(resp_text.strip())
                if isinstance(parsed, list):
                    if len(parsed) == 1 and isinstance(parsed[0], dict):
                        parsed = parsed[0]
                    else:
                        merged = {}
                        for item in parsed:
                            if isinstance(item, dict):
                                merged.update(item)
                        if merged:
                            parsed = merged
                        else:
                            raise ValueError(f"LLM returned list with no usable dicts")
                if not isinstance(parsed, dict):
                    raise ValueError(f"LLM returned {type(parsed).__name__}, expected dict")

                cell_mapping = parsed
                break
            except (json.JSONDecodeError, ValueError) as e:
                logger.warning(
                    "Inquiry v6.2: supplier %d attempt %d parse failed: %s\nLLM raw output (first 500): %s",
                    supplier_id, attempt + 1, e,
                    (resp_text or "")[:500],
                )
                if attempt == max_retries - 1:
                    raise
            except Exception as e:
                logger.warning("Inquiry v6.2: supplier %d attempt %d API error: %s", supplier_id, attempt + 1, e)
                if attempt == max_retries - 1:
                    raise
                time.sleep(2 ** attempt)

        # ── Stage 4: Deterministic format enforcement ──
        date_field_positions = set()
        for f in fields:
            fk = f.get("field_key", "")
            if "date" in fk.lower() or fk in ("order_date", "delivery_date", "payment_date"):
                date_field_positions.add(f["position"])

        for cell_ref, value in list(cell_mapping.items()):
            _ensure_not_cancelled(cancel_event)
            if value is None or value == "":
                continue
            val_str = str(value)

            if cell_ref in date_field_positions:
                ann = annotations.get(cell_ref, "")
                parsed = _try_parse_date(val_str)
                if parsed:
                    if re.search(r"和暦|令和", ann):
                        reiwa_year = parsed.year - 2018
                        cell_mapping[cell_ref] = f"R{reiwa_year}.{parsed.month:02d}.{parsed.day:02d}"
                    elif re.search(r"DD[/\-.]MM[/\-.]YYYY", ann, re.IGNORECASE):
                        cell_mapping[cell_ref] = parsed.strftime("%d/%m/%Y")
                    else:
                        cell_mapping[cell_ref] = parsed.strftime("%Y/%m/%d")
                    continue

            ann = annotations.get(cell_ref, "")
            if not ann:
                for ann_cell, ann_text in annotations.items():
                    if ann_cell[0] == cell_ref[0] and len(ann_cell) <= 3 and len(cell_ref) <= 3:
                        ann = ann_text
                        break
            if ann:
                enforced = enforce_annotation(val_str, ann)
                if enforced != val_str:
                    cell_mapping[cell_ref] = enforced
    else:
        logger.info("Inquiry v6.2: supplier %d — no header fields, skipping LLM", supplier_id)

    # ── Stage 5: Write workbook ──
    _ensure_not_cancelled(cancel_event)
    # Write header cells (skip null = preserve template)
    cells_to_write = [
        {"cell": c, "value": str(v)}
        for c, v in cell_mapping.items()
        if v is not None
    ]
    wb.write_cells(cells_to_write)

    # Write product rows
    table_config = chosen_template.product_table_config or {}
    columns = table_config.get("columns", {})
    start_row = table_config.get("start_row", 22)
    formula_cols = table_config.get("formula_columns", [])

    # ── Auto-correct formula_columns/columns conflict ──
    # AI template analysis sometimes marks data columns (D, E, K) as formula_columns.
    # If a column is in both `columns` and `formula_columns`, it creates a contradiction:
    # the data won't be written because formula_columns means "skip writing".
    # Fix: only keep columns in formula_columns if they are NOT in the data columns mapping,
    # OR if they have actual formulas in template_formulas.
    if formula_cols and columns:
        actual_formula_cols = set()
        for fc in formula_cols:
            fc_upper = fc.upper()
            # Keep it if there's an actual formula in the template for this column
            has_real_formula = any(
                re.match(r"([A-Z]+)", ref).group(1).upper() == fc_upper
                for ref in (template_formulas or {})
            )
            if has_real_formula:
                actual_formula_cols.add(fc)
            elif fc_upper in {c.upper() for c in columns}:
                logger.warning(
                    "Auto-corrected: column %s removed from formula_columns "
                    "(no actual formula found, but it's in data columns)", fc
                )
            else:
                actual_formula_cols.add(fc)
        if len(actual_formula_cols) != len(formula_cols):
            logger.info(
                "formula_columns corrected: %s → %s",
                formula_cols, sorted(actual_formula_cols),
            )
            formula_cols = sorted(actual_formula_cols)

    count = len(products)

    # ── Dynamic row expansion ──
    # If products exceed template capacity, insert rows to push summary/footer down.
    # NOTE: openpyxl.insert_rows does NOT reliably adjust formula references.
    # We must manually fix header formulas and summary formulas after insertion.
    extra_rows = 0
    if template_formulas and wb._ws and count > 0:
        # Find the first summary formula row (= boundary between product area and summary)
        summary_row_nums = []
        for fc_ref in template_formulas:
            row_num = int(re.search(r"(\d+)", fc_ref).group(1))
            if row_num > start_row + 10:  # outside per-row formula range
                summary_row_nums.append(row_num)

        if summary_row_nums:
            first_summary_row = min(summary_row_nums)
            template_capacity = first_summary_row - start_row
            if count > template_capacity:
                extra_rows = count - template_capacity
                wb._ws.insert_rows(first_summary_row, extra_rows)
                logger.info(
                    "Inserted %d rows at row %d (template_capacity=%d, products=%d)",
                    extra_rows, first_summary_row, template_capacity, count,
                )
                # Copy formatting from last template product row to inserted rows
                _copy_row_formatting(wb._ws, first_summary_row - 1, first_summary_row, extra_rows)

    wb.write_product_rows(start_row, columns, products, formula_cols, metadata=enriched_meta)

    # Post-fill: pack_size → description column
    if wb._ws:
        desc_col = None
        for col_letter, field_key in columns.items():
            if field_key == "description":
                desc_col = col_letter
                break
        if desc_col:
            for i, product in enumerate(products):
                _ensure_not_cancelled(cancel_event)
                matched = product.get("matched_product") or {}
                pack_size = matched.get("pack_size", "")
                if pack_size:
                    wb.safe_set_cell(f"{desc_col}{start_row + i}", pack_size)

    # Enforce annotations on product rows + set number_format
    if wb._ws and annotations:
        for ann_cell, ann_text in annotations.items():
            m = re.match(r"([A-Z]+)(\d+)", ann_cell)
            if not m:
                continue
            ann_col = m.group(1)
            ann_row_num = int(m.group(2))
            if ann_row_num >= start_row:
                for ri in range(start_row, start_row + count):
                    _ensure_not_cancelled(cancel_event)
                    cell_ref = f"{ann_col}{ri}"
                    cell = wb._ws[cell_ref]
                    if isinstance(cell, MergedCell):
                        continue
                    cell_val = cell.value
                    if cell_val is not None:
                        original = str(cell_val)
                        enforced = enforce_annotation(original, ann_text)
                        if enforced != original:
                            try:
                                wb.safe_set_cell(cell_ref, float(enforced))
                            except (ValueError, TypeError):
                                wb.safe_set_cell(cell_ref, enforced)
                        # number_format for decimals
                        dec_match = re.search(
                            r"小数点\s*(?:后面|まで)?\s*(\d+|[一二两三四五六七八九十]+)\s*位|(\d+)\s*decimal",
                            ann_text, re.IGNORECASE,
                        )
                        if dec_match:
                            raw_d = dec_match.group(1) or dec_match.group(2)
                            cn_d = {"一": 1, "二": 2, "两": 2, "三": 3, "四": 4, "五": 5}
                            pl = cn_d.get(raw_d, None) if not raw_d.isdigit() else int(raw_d)
                            if pl:
                                try:
                                    val = float(cell.value)
                                    wb.safe_set_cell(cell_ref, val, number_format="0." + "0" * pl)
                                except (ValueError, TypeError):
                                    pass

    # Rebuild formulas
    # openpyxl.insert_rows does NOT reliably adjust formula references.
    # We must manually handle all three categories:
    #   1. Header formulas (row < start_row): update references to shifted summary rows
    #   2. Per-row formulas: replicate for ALL product rows
    #   3. Summary formulas: relocate + adjust SUM ranges
    allowed_formula_cols = set(c.upper() for c in formula_cols) if formula_cols else set()
    if template_formulas and wb._ws:
        last_data_row = start_row + count - 1
        header_formulas: list[tuple[str, str]] = []  # [(ref, formula), ...]
        per_row_formulas: dict[str, tuple[str, int]] = {}
        summary_formulas: dict[str, tuple[str, str]] = {}  # {orig_ref: (orig_formula, new_ref)}

        for fc_ref, fc_val in template_formulas.items():
            col_letter = re.match(r"([A-Z]+)", fc_ref).group(1)
            row_num = int(re.search(r"(\d+)", fc_ref).group(1))
            if row_num < start_row:
                # Header formula (e.g., H16=L35) — must manually adjust references
                header_formulas.append((fc_ref, fc_val))
                continue
            elif start_row <= row_num <= start_row + 10:
                # Per-row formula (within product area)
                if allowed_formula_cols and col_letter.upper() not in allowed_formula_cols:
                    logger.debug("Skipping non-whitelisted per-row formula at %s: %s", fc_ref, fc_val[:60])
                    continue
                if col_letter not in per_row_formulas:
                    per_row_formulas[col_letter] = (fc_val, row_num)
            else:
                # Summary formula — after insert_rows it's at row_num + extra_rows
                new_row = row_num + extra_rows
                new_ref = f"{col_letter}{new_row}"
                summary_formulas[fc_ref] = (fc_val, new_ref)

        # 1. Per-row formulas: write for each product row
        for col, (formula_template, orig_row) in per_row_formulas.items():
            for row_idx in range(start_row, start_row + count):
                _ensure_not_cancelled(cancel_event)
                new_formula = re.sub(str(orig_row), str(row_idx), formula_template)
                wb.safe_set_cell(f"{col}{row_idx}", new_formula)
            # Clear leftover formula rows (between last product and first summary)
            clear_end = start_row + count + 10
            if summary_formulas:
                first_new_summary = min(
                    int(re.search(r"(\d+)", nr).group(1))
                    for _, (_, nr) in summary_formulas.items()
                )
                clear_end = min(clear_end, first_new_summary)
            for row_idx in range(start_row + count, clear_end):
                c = wb._ws[f"{col}{row_idx}"]
                if not isinstance(c, MergedCell) and c.value and isinstance(c.value, str) and c.value.startswith("="):
                    wb.safe_set_cell(f"{col}{row_idx}", None)

        # 2. Summary formulas: adjust SUM ranges to cover all product rows
        # Build row shift map for inter-formula references (e.g., L33*0.08 → L72*0.08)
        row_shift_map: dict[int, int] = {}
        for fc_ref, (_, new_ref) in summary_formulas.items():
            old_row = int(re.search(r"(\d+)", fc_ref).group(1))
            new_row = int(re.search(r"(\d+)", new_ref).group(1))
            if old_row != new_row:
                row_shift_map[old_row] = new_row

        for fc_ref, (orig_formula, new_ref) in summary_formulas.items():
            _ensure_not_cancelled(cancel_event)
            # Expand range references (e.g., SUM(L22:L32) → SUM(L22:L71))
            new_formula = re.sub(
                r"([A-Z]+)(\d+):([A-Z]+)(\d+)",
                lambda m_: f"{m_.group(1)}{start_row}:{m_.group(3)}{last_data_row}",
                orig_formula,
            )
            # Shift single cell references for inter-formula deps (e.g., L33 → L72)
            if row_shift_map:
                def _shift_cell_ref(m_):
                    col_part = m_.group(1)
                    row_part = int(m_.group(2))
                    return f"{col_part}{row_shift_map.get(row_part, row_part)}"
                new_formula = re.sub(r"([A-Z]+)(\d+)", _shift_cell_ref, new_formula)
            wb.safe_set_cell(new_ref, new_formula)

        # 3. Header formulas: update references to shifted summary rows
        if extra_rows > 0 and header_formulas and row_shift_map:
            for hf_ref, hf_formula in header_formulas:
                new_formula = hf_formula
                def _shift_header_ref(m_, _map=row_shift_map):
                    col_part = m_.group(1)
                    row_part = int(m_.group(2))
                    return f"{col_part}{_map.get(row_part, row_part)}"
                new_formula = re.sub(r"([A-Z]+)(\d+)", _shift_header_ref, new_formula)
                if new_formula != hf_formula:
                    wb.safe_set_cell(hf_ref, new_formula)
                    logger.info("Header formula %s updated: %s → %s", hf_ref, hf_formula, new_formula)

    # ── Stage 6: Save ──
    _ensure_not_cancelled(cancel_event)
    return _save_workbook(wb, chosen_template, selection_method, order_meta, supplier_id,
                          products, upload_dir, annotations, stream_key, overall_start)


def _save_workbook(
    wb, template, selection_method: str, order_meta: dict,
    supplier_id: int, products: list[dict], upload_dir: str,
    annotations: dict[str, str], stream_key: str, overall_start: float,
    verify_results: list[dict[str, Any]] | None = None,
) -> dict:
    """Save workbook + preview HTML, run verify, return result dict."""
    from services.agent.stream_queue import get_or_create_cancel_event, push_event

    excel_bytes = wb.save_bytes()

    po_number = str(order_meta.get("po_number") or "unknown").replace("/", "_").replace("\\", "_")
    filename = f"inquiry_{po_number}_supplier{supplier_id}_{uuid.uuid4().hex[:6]}.xlsx"
    file_url = file_storage.upload(
        "inquiries", filename, excel_bytes,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    # Also save to session workspace so agent's bash tool can access/modify the file
    # (DeerFlow pattern: all tools share the same persistent directory per thread)
    if stream_key:
        try:
            import os
            from core.config import settings
            ws_dir = os.path.join(settings.AGENT_WORKSPACE_ROOT, stream_key)
            os.makedirs(ws_dir, exist_ok=True)
            ws_path = os.path.join(ws_dir, filename)
            with open(ws_path, "wb") as f:
                f.write(excel_bytes)
            logger.info("Saved inquiry to workspace: %s", ws_path)
        except Exception as e:
            logger.warning("Failed to save inquiry to workspace: %s", e)

    # Save preview HTML
    preview_filename = filename.replace(".xlsx", ".html")
    preview_url = None
    try:
        html = wb.render_html()
        html_bytes = html.encode("utf-8")
        preview_url = file_storage.upload(
            "inquiries", preview_filename, html_bytes, content_type="text/html; charset=utf-8",
        )
    except Exception as e:
        logger.warning("Preview HTML save failed for supplier %d: %s", supplier_id, e)

    # Run verify
    if verify_results is None:
        verify_results = []
    if not verify_results and annotations and wb._ws:
        for cell_ref, annotation in annotations.items():
            try:
                cell = wb._ws[cell_ref]
                current_value = cell.value if cell.value is not None else ""
                check = _code_check(current_value, annotation)
                verify_results.append({
                    "cell": cell_ref,
                    "annotation": annotation,
                    "value": str(current_value)[:100],
                    **check,
                })
            except Exception:
                verify_results.append({
                    "cell": cell_ref,
                    "annotation": annotation,
                    "value": "(无法读取)",
                    "status": "fail",
                    "reason": "单元格读取失败",
                })

    file_info = {
        "supplier_id": supplier_id,
        "filename": filename,
        "file_url": file_url,
        "preview_url": preview_url,
        "product_count": len(products),
        "has_template": template is not None,
        "template_name": template.template_name if template else None,
        "template_id": template.id if template else None,
        "selection_method": selection_method,
        "generation_path": getattr(wb, '_generation_path', "unknown"),  # template_engine / llm_mapping / generic
    }

    # SSE: emit "save" step
    if stream_key:
        push_event(stream_key, {
            "type": "tool_result", "tool_name": "save", "tool_label": "保存文件",
            "content": f"{filename} ({len(products)} 产品)",
            "elapsed_seconds": round(time.time() - overall_start, 1),
            "supplier_id": supplier_id,
        })

    return {
        "file_info": file_info,
        "verify_results": verify_results,
        "error": None,
    }


def _engine_verify_to_results(verify_report: dict[str, Any]) -> list[dict[str, Any]]:
    """Convert template_engine.verify_output report to frontend-facing verify_results."""
    errors = verify_report.get("errors", []) or []
    checks = int(verify_report.get("checks") or 0)

    if not errors:
        return [{
            "cell": "SYSTEM",
            "annotation": "模板结构与字段校验",
            "value": f"{checks} 项检查",
            "status": "pass",
            "reason": "所有检查通过",
        }]

    results: list[dict[str, Any]] = []
    for error in errors:
        if ":" in error:
            label, reason = error.split(":", 1)
            cell = _extract_cell_from_error(label) or "SYSTEM"
            annotation = label.strip()
            reason_text = reason.strip()
        else:
            cell = _extract_cell_from_error(error) or "SYSTEM"
            annotation = error
            reason_text = error

        results.append({
            "cell": cell,
            "annotation": annotation,
            "value": "",
            "status": "fail",
            "reason": reason_text,
            "suggestion": _suggest_repair_from_error(annotation, reason_text),
        })
    return results


def _extract_cell_from_error(text: str) -> str | None:
    match = re.search(r"\b([A-Z]{1,3}\d+)\b", text)
    return match.group(1) if match else None


def _suggest_repair_from_error(annotation: str, reason: str) -> str:
    lowered = f"{annotation} {reason}".lower()
    if "header anchor" in lowered or "header " in lowered:
        return "检查表头字段映射或在“编辑字段”中补正确值"
    if "product header" in lowered:
        return "检查产品列映射与模板表头是否一致"
    if "formula" in lowered:
        return "检查模板公式列与汇总区公式配置"
    if "summary" in lowered:
        return "检查汇总区标签、税额/总额配置和模板静态文本"
    if "merge" in lowered:
        return "检查模板合并单元格结构是否被破坏"
    if "stale data" in lowered:
        return "检查汇总区是否残留旧产品数据"
    return "打开 Review/Repair 检查字段和模板结构后重试"


# ─── Orchestrator ─────────────────────────────────────────────

_MAX_INQUIRY_WORKERS = int(os.environ.get("INQUIRY_CONCURRENCY", "3"))


def run_inquiry_orchestrator(
    order,
    db,
    stream_key: str = "",
    template_overrides=None,
    supplier_ids: list[int] | None = None,
) -> dict:
    """Orchestrator: run per-supplier generation in parallel via ThreadPoolExecutor.

    Returns inquiry_data dict with per-supplier results.
    """
    from services.agent.stream_queue import push_event

    from services.tools.product_matching import normalize_matched_product

    overall_start = time.time()
    order_id = order.id
    match_results = order.match_results or []
    order_meta = order.order_metadata or {}
    cancel_event = get_or_create_cancel_event(stream_key) if stream_key else None

    # Normalize matched_product dicts (ensures canonical field names like
    # "unit_price" exist even for orders matched before the normalize fix)
    for item in match_results:
        if item.get("matched_product"):
            normalize_matched_product(item["matched_product"])

    # Group products by supplier_id
    supplier_groups: dict[int, list[dict]] = {}
    for item in match_results:
        matched = item.get("matched_product")
        if matched and matched.get("supplier_id"):
            sid = matched["supplier_id"]
            supplier_groups.setdefault(sid, []).append(item)

    if supplier_ids is not None:
        allowed = set(int(sid) for sid in supplier_ids)
        supplier_groups = {
            sid: products
            for sid, products in supplier_groups.items()
            if sid in allowed
        }

    unassigned = sum(
        1 for item in match_results
        if not (item.get("matched_product") or {}).get("supplier_id")
    )

    # Pre-fetch supplier info (names, contact completeness) for all suppliers
    import sqlalchemy
    supplier_info_map: dict[int, dict] = {}
    supplier_ids_list = list(supplier_groups.keys())
    if supplier_ids_list:
        try:
            rows = db.execute(
                sqlalchemy.text(
                    "SELECT id, name, contact, email, phone FROM suppliers WHERE id = ANY(:ids)"
                ),
                {"ids": supplier_ids_list},
            ).fetchall()
            for row in rows:
                sid_val, name, contact, email, phone = row[0], row[1], row[2], row[3], row[4]
                missing = []
                if not contact:
                    missing.append("contact")
                if not email:
                    missing.append("email")
                if not phone:
                    missing.append("phone")
                supplier_info_map[sid_val] = {
                    "name": name or f"供应商 #{sid_val}",
                    "missing_fields": missing if missing else None,
                }
        except Exception as e:
            logger.warning("Failed to fetch supplier info: %s", e)

    suppliers_result: dict[str, dict] = {}
    generated_files: list[dict] = []

    max_workers = min(len(supplier_groups), _MAX_INQUIRY_WORKERS) or 1

    def _worker(sid: int, sid_products: list[dict]) -> tuple[int, dict]:
        """Thread worker: generates one supplier's inquiry and returns (sid, result_dict)."""
        supplier_start = time.time()
        _ensure_not_cancelled(cancel_event)

        info = supplier_info_map.get(sid, {})

        if stream_key:
            push_event(stream_key, {
                "type": "supplier_start",
                "supplier_id": sid,
                "supplier_name": info.get("name") or f"供应商 #{sid}",
                "product_count": len(sid_products),
            })
        subtotal = 0.0
        for p in sid_products:
            mp = p.get("matched_product") or {}
            qty = p.get("quantity") if p.get("quantity") is not None else 0
            val = p.get("unit_price")
            price = val if val is not None else (mp.get("price") if mp.get("price") is not None else 0)
            try:
                subtotal += float(qty) * float(price)
            except (TypeError, ValueError):
                pass

        base_info = {
            "supplier_name": info.get("name") or f"供应商 #{sid}",
            "product_count": len(sid_products),
            "subtotal": round(subtotal, 2),
            "currency": order_meta.get("currency", ""),
            "missing_fields": info.get("missing_fields"),
        }

        override_tid = (template_overrides or {}).get(sid)
        try:
            result = _generate_single_supplier(
                order_id,
                order_meta,
                sid,
                sid_products,
                stream_key,
                overall_start,
                template_id_override=override_tid,
                cancel_event=cancel_event,
            )
            elapsed = round(time.time() - supplier_start, 1)

            fi = result.get("file_info") or {}
            final_status = "error" if result.get("status") == "repair_required" else "completed"
            supplier_data = {
                **base_info,
                "status": final_status,
                "file": result.get("file_info"),
                "template": {
                    "id": fi.get("template_id"),
                    "name": fi.get("template_name"),
                    "selection_method": fi.get("selection_method", "none"),
                },
                "verify_results": result.get("verify_results", []),
                "elapsed_seconds": elapsed,
            }
            if result.get("error"):
                supplier_data["error"] = result["error"]
            logger.info("Inquiry orchestrator: supplier %d completed in %.1fs", sid, elapsed)
        except InquiryCancelledError as e:
            elapsed = round(time.time() - supplier_start, 1)
            supplier_data = {
                **base_info,
                "status": "cancelled",
                "error": str(e),
                "elapsed_seconds": elapsed,
            }
        except Exception as e:
            elapsed = round(time.time() - supplier_start, 1)
            logger.error("Inquiry orchestrator: supplier %d failed: %s", sid, e, exc_info=True)
            supplier_data = {
                **base_info,
                "status": "error",
                "error": str(e),
                "elapsed_seconds": elapsed,
            }

        if stream_key:
            push_event(stream_key, {
                "type": "supplier_done",
                "supplier_id": sid,
                "status": supplier_data["status"],
            })

        return sid, supplier_data

    logger.info("Inquiry orchestrator: %d suppliers, max_workers=%d", len(supplier_groups), max_workers)

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {
            executor.submit(_worker, sid, prods): sid
            for sid, prods in supplier_groups.items()
        }

        for future in as_completed(futures):
            sid = futures[future]
            if cancel_event is not None and cancel_event.is_set():
                for pending in futures:
                    pending.cancel()
            try:
                _, supplier_data = future.result()
            except Exception as e:
                logger.error("Inquiry orchestrator: future for supplier %d raised: %s", sid, e, exc_info=True)
                supplier_data = {"status": "error", "error": str(e), "elapsed_seconds": 0}

            suppliers_result[str(sid)] = supplier_data
            file_info = supplier_data.get("file")
            if file_info:
                generated_files.append(file_info)
            elif supplier_data["status"] == "error":
                generated_files.append({
                    "supplier_id": sid,
                    "filename": None,
                    "error": supplier_data.get("error"),
                    "product_count": len(supplier_groups.get(sid, [])),
                })

    total_elapsed = round(time.time() - overall_start, 1)

    inquiry_data = {
        "suppliers": suppliers_result,
        "generated_files": generated_files,
        "supplier_count": len(supplier_groups),
        "unassigned_count": unassigned,
        "total_elapsed_seconds": total_elapsed,
    }

    return inquiry_data


def run_inquiry_single_supplier(order, db, supplier_id: int, stream_key: str = "", template_id: int | None = None) -> dict:
    """Run inquiry for a single supplier (for re-do). Returns per-supplier result dict."""
    from services.agent.stream_queue import get_or_create_cancel_event, push_event
    from services.tools.product_matching import normalize_matched_product

    match_results = order.match_results or []
    order_meta = order.order_metadata or {}
    cancel_event = get_or_create_cancel_event(stream_key) if stream_key else None
    _ensure_not_cancelled(cancel_event)

    # Normalize matched_product dicts (ensures canonical field names for old orders)
    for item in match_results:
        if item.get("matched_product"):
            normalize_matched_product(item["matched_product"])

    sid_products = [
        item for item in match_results
        if (item.get("matched_product") or {}).get("supplier_id") == supplier_id
    ]

    if not sid_products:
        raise ValueError(f"供应商 {supplier_id} 没有匹配的产品")

    # Fetch supplier info
    import sqlalchemy
    supplier_name = f"供应商 #{supplier_id}"
    missing_fields = None
    try:
        row = db.execute(
            sqlalchemy.text("SELECT name, contact, email, phone FROM suppliers WHERE id = :sid"),
            {"sid": supplier_id},
        ).fetchone()
        if row:
            supplier_name = row[0] or supplier_name
            missing = []
            if not row[1]:
                missing.append("contact")
            if not row[2]:
                missing.append("email")
            if not row[3]:
                missing.append("phone")
            missing_fields = missing if missing else None
    except Exception:
        pass

    # Compute subtotal
    subtotal = 0.0
    for p in sid_products:
        mp = p.get("matched_product") or {}
        qty = p.get("quantity") or 0
        price = p.get("unit_price") or mp.get("price") or 0
        try:
            subtotal += float(qty) * float(price)
        except (TypeError, ValueError):
            pass

    start_time = time.time()

    if stream_key:
        push_event(stream_key, {
            "type": "supplier_start",
            "supplier_id": supplier_id,
            "supplier_name": supplier_name,
            "product_count": len(sid_products),
        })

    result = _generate_single_supplier(
        order.id,
        order_meta,
        supplier_id,
        sid_products,
        stream_key,
        start_time,
        template_id_override=template_id,
        cancel_event=cancel_event,
    )
    elapsed = round(time.time() - start_time, 1)

    fi = result.get("file_info") or {}
    final_status = "error" if result.get("status") == "repair_required" else "completed"
    supplier_result = {
        "supplier_name": supplier_name,
        "product_count": len(sid_products),
        "subtotal": round(subtotal, 2),
        "currency": order_meta.get("currency", ""),
        "missing_fields": missing_fields,
        "status": final_status,
        "file": result.get("file_info"),
        "template": {
            "id": fi.get("template_id"),
            "name": fi.get("template_name"),
            "selection_method": fi.get("selection_method", "none"),
        },
        "verify_results": result.get("verify_results", []),
        "elapsed_seconds": elapsed,
    }
    if result.get("error"):
        supplier_result["error"] = result["error"]

    if stream_key:
        push_event(stream_key, {
            "type": "supplier_done",
            "supplier_id": supplier_id,
            "status": supplier_result["status"],
        })

    return supplier_result
