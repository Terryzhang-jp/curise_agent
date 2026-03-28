"""Generate product upload template Excel with reference data from DB."""

from __future__ import annotations

import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

logger = logging.getLogger(__name__)

# ─── Styles ───
_HEADER_FONT = Font(name="Arial", bold=True, size=11, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
_REQUIRED_FILL = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
_REQUIRED_FONT = Font(name="Arial", bold=True, size=11, color="FFFFFF")
_EXAMPLE_FONT = Font(name="Arial", size=10, color="808080", italic=True)
_THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
_REF_HEADER_FONT = Font(name="Arial", bold=True, size=10, color="2F5496")
_REF_FONT = Font(name="Arial", size=10)
_TITLE_FONT = Font(name="Arial", bold=True, size=14, color="2F5496")
_SECTION_FONT = Font(name="Arial", bold=True, size=11, color="2F5496")
_BODY_FONT = Font(name="Arial", size=10)
_BOLD_BODY = Font(name="Arial", size=10, bold=True)

# ─── Column definitions ───
_COLUMNS = [
    ("product_name",      30, True,  "产品英文名称（必填）"),
    ("product_name_jp",   25, False, "产品日文名称"),
    ("product_code",      15, False, "产品代码 / SKU"),
    ("country_id",        12, True,  "国家ID（必填）见参考表"),
    ("category_id",       12, True,  "类别ID（必填）见参考表"),
    ("port_id",           10, False, "港口ID，见参考表"),
    ("supplier_id",       12, False, "供应商ID，见参考表"),
    ("unit",              10, False, "KG / L / PCS / CT / BOX"),
    ("price",             12, False, "单价（纯数字）"),
    ("currency",          10, False, "默认 JPY"),
    ("pack_size",         18, False, "包装规格"),
    ("unit_size",         12, False, "单位规格"),
    ("brand",             15, False, "品牌"),
    ("country_of_origin", 18, False, "产地 / 原産地"),
    ("effective_from",    15, True,  "生效日期 YYYY-MM-DD（必填）"),
    ("effective_to",      15, False, "失效日期 YYYY-MM-DD"),
]

_EXAMPLES = [
    ["FROZEN SHRIMP 26/30", "冷凍エビ 26/30", "FRZ-SHR-001", 9, 21, 19, 2,
     "KG", 1250, "JPY", "2LB/bag x 10", "20LB", "Oceanfresh", "Thailand", "2025-01-15", "2025-12-31"],
    ["BEET GOLD LARGE", "ビーツ ゴールド", "VEG-BET-002", 11, 14, 29, 15,
     "KG", 25.50, "AUD", "25LB bulk", "25LB", "", "Australia", "2025-02-01", ""],
    ["CANNED MUSHROOM SLICED", "缶詰マッシュルーム", "CAN-MSH-003", 9, 15, 22, 3,
     "CT", 890, "JPY", "24 x 400g", "", "Golden Harvest", "China", "2025-03-01", "2026-02-28"],
]


def generate_product_upload_template(output_path: str) -> None:
    """Generate product upload template Excel with DB reference data."""
    from database import SessionLocal
    from sqlalchemy import text

    db = SessionLocal()
    try:
        countries = db.execute(
            text("SELECT id, name FROM countries WHERE name NOT LIKE '测试%' ORDER BY name")
        ).fetchall()
        categories = db.execute(
            text("SELECT id, name FROM categories WHERE name NOT LIKE '测试%' ORDER BY name")
        ).fetchall()
        ports = db.execute(text(
            "SELECT p.id, p.name, c.name FROM ports p "
            "JOIN countries c ON c.id = p.country_id "
            "WHERE p.name NOT LIKE '测试%' ORDER BY c.name, p.name"
        )).fetchall()
        suppliers = db.execute(
            text("SELECT id, name FROM suppliers ORDER BY name")
        ).fetchall()
    finally:
        db.close()

    wb = Workbook()

    # ─── Sheet 1: Products ───
    ws = wb.active
    ws.title = "Products"

    for col_idx, (header, width, required, comment_text) in enumerate(_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _REQUIRED_FONT if required else _HEADER_FONT
        cell.fill = _REQUIRED_FILL if required else _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
        cell.comment = Comment(comment_text, "System", width=220, height=50)

    for row_offset, data in enumerate(_EXAMPLES):
        row_num = 2 + row_offset
        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.font = _EXAMPLE_FONT
            cell.border = _THIN_BORDER
            cell.alignment = Alignment(vertical="center")

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 28

    # ─── Sheet 2: Reference ───
    ws_ref = wb.create_sheet("参考表 Reference")

    # Countries (A-B)
    ws_ref.column_dimensions["A"].width = 10
    ws_ref.column_dimensions["B"].width = 30
    ws_ref.cell(row=1, column=1, value="country_id").font = _REF_HEADER_FONT
    ws_ref.cell(row=1, column=2, value="country_name").font = _REF_HEADER_FONT
    for i, row in enumerate(countries):
        ws_ref.cell(row=2 + i, column=1, value=row[0]).font = _REF_FONT
        ws_ref.cell(row=2 + i, column=2, value=row[1]).font = _REF_FONT

    # Categories (D-E)
    ws_ref.column_dimensions["D"].width = 10
    ws_ref.column_dimensions["E"].width = 20
    ws_ref.cell(row=1, column=4, value="category_id").font = _REF_HEADER_FONT
    ws_ref.cell(row=1, column=5, value="category_name").font = _REF_HEADER_FONT
    for i, row in enumerate(categories):
        ws_ref.cell(row=2 + i, column=4, value=row[0]).font = _REF_FONT
        ws_ref.cell(row=2 + i, column=5, value=row[1]).font = _REF_FONT

    # Ports (G-I)
    ws_ref.column_dimensions["G"].width = 10
    ws_ref.column_dimensions["H"].width = 28
    ws_ref.column_dimensions["I"].width = 20
    ws_ref.cell(row=1, column=7, value="port_id").font = _REF_HEADER_FONT
    ws_ref.cell(row=1, column=8, value="port_name").font = _REF_HEADER_FONT
    ws_ref.cell(row=1, column=9, value="country").font = _REF_HEADER_FONT
    for i, row in enumerate(ports):
        ws_ref.cell(row=2 + i, column=7, value=row[0]).font = _REF_FONT
        ws_ref.cell(row=2 + i, column=8, value=row[1]).font = _REF_FONT
        ws_ref.cell(row=2 + i, column=9, value=row[2]).font = _REF_FONT

    # Suppliers (K-L)
    ws_ref.column_dimensions["K"].width = 10
    ws_ref.column_dimensions["L"].width = 30
    ws_ref.cell(row=1, column=11, value="supplier_id").font = _REF_HEADER_FONT
    ws_ref.cell(row=1, column=12, value="supplier_name").font = _REF_HEADER_FONT
    for i, row in enumerate(suppliers):
        ws_ref.cell(row=2 + i, column=11, value=row[0]).font = _REF_FONT
        ws_ref.cell(row=2 + i, column=12, value=row[1]).font = _REF_FONT

    # ─── Sheet 3: Instructions ───
    ws2 = wb.create_sheet("使用说明 Instructions")
    ws2.column_dimensions["A"].width = 24
    ws2.column_dimensions["B"].width = 52
    ws2.column_dimensions["C"].width = 30

    lines = [
        ("title", "产品上传模板 Product Upload Template", "", ""),
        ("", "", "", ""),
        ("section", "使用步骤", "", ""),
        ("", "1.", "在 Products 表中从第 2 行开始填写数据（第 1 行为列头）", ""),
        ("", "2.", "红色列头为必填字段，蓝色为选填", ""),
        ("", "3.", "country_id / category_id 等请查看「参考表 Reference」获取 ID", ""),
        ("", "4.", "灰色斜体为示例数据，可直接覆盖或删除", ""),
        ("", "5.", "鼠标悬停列头可查看字段说明", ""),
        ("", "", "", ""),
        ("section", "字段说明", "格式要求", "示例"),
        ("req", "product_name *", "产品英文名称（必填）", "FROZEN SHRIMP 26/30"),
        ("", "product_name_jp", "产品日文名称", "冷凍エビ 26/30"),
        ("", "product_code", "产品代码/SKU（同国家+港口下唯一）", "FRZ-SHR-001"),
        ("req", "country_id *", "国家 ID（必填，见参考表）", "9 = JAPAN"),
        ("req", "category_id *", "类别 ID（必填，见参考表）", "21 = SEAFOOD"),
        ("", "port_id", "港口 ID（见参考表）", "19 = 横浜 大さん橋"),
        ("", "supplier_id", "供应商 ID（见参考表）", "2 = 株式会社 松武"),
        ("", "unit", "计量单位", "KG / L / PCS / CT"),
        ("", "price", "单价（纯数字，不含货币符号）", "1250 / 25.50"),
        ("", "currency", "货币代码（默认 JPY）", "JPY / USD / AUD"),
        ("", "pack_size", "包装规格", "6-10ct/10kg"),
        ("", "unit_size", "单位规格", "25LB"),
        ("", "brand", "品牌", "Oceanfresh"),
        ("", "country_of_origin", "产地/原産地", "Thailand"),
        ("req", "effective_from *", "生效日期（必填，YYYY-MM-DD）", "2025-01-15"),
        ("", "effective_to", "失效日期（YYYY-MM-DD）", "2025-12-31"),
        ("", "", "", ""),
        ("section", "唯一性约束", "", ""),
        ("", "•", "product_code + country_id + port_id 组合必须唯一", ""),
        ("", "•", "product_name + country_id + port_id 组合必须唯一", ""),
        ("", "•", "重复数据将被跳过并提示错误", ""),
    ]

    for row_idx, (style, a, b, c) in enumerate(lines, 1):
        ws2.cell(row=row_idx, column=1, value=a)
        ws2.cell(row=row_idx, column=2, value=b)
        ws2.cell(row=row_idx, column=3, value=c)
        if style == "title":
            ws2.cell(row=row_idx, column=1).font = _TITLE_FONT
            ws2.row_dimensions[row_idx].height = 28
        elif style == "section":
            for col in range(1, 4):
                ws2.cell(row=row_idx, column=col).font = _SECTION_FONT
        elif style == "req":
            ws2.cell(row=row_idx, column=1).font = _BOLD_BODY
            ws2.cell(row=row_idx, column=2).font = _BODY_FONT
            ws2.cell(row=row_idx, column=3).font = _BODY_FONT
        else:
            for col in range(1, 4):
                ws2.cell(row=row_idx, column=col).font = _BODY_FONT

    wb.save(output_path)
