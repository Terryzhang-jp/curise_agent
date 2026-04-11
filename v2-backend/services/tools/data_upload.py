"""
Data upload tools — staging-based product upload with confidence matching,
reference creation, audit logging, atomic execution, and rollback.

Tools:
  1. parse_file: Parse Excel, create UploadBatch + StagingProduct rows
  2. analyze_columns: Cross-reference unmapped columns against DB ref tables
  3. resolve_and_validate: Code match + LLM fuzzy match with confidence scoring
  4. create_references: Create missing suppliers/countries in dependency order
  5. preview_changes: Generate change preview report
  6. execute_upload: Atomic write to products table + ProductChangeLog
  7. audit_data: Structural + LLM semantic data quality audit
  8. prepare_upload: One-step validation + audit + preview
  9. rollback_batch: Undo a completed batch import

Tools use the closure pattern: create_data_upload_tools(registry, ctx).
Conditional registration: completed batches only get rollback_batch.
"""

from __future__ import annotations

import hashlib
import json
import logging
import re
from datetime import datetime
from decimal import Decimal, InvalidOperation
from difflib import SequenceMatcher
from io import BytesIO

from services.tools.registry_loader import ToolMetaInfo

logger = logging.getLogger(__name__)

TOOL_META = {
    "parse_upload": ToolMetaInfo(
        display_name="解析上传文件",
        group="business",
        description="解析上传的 Excel/CSV 文件，创建暂存数据",
        prompt_description="解析上传文件（Excel/CSV → 暂存数据）",
        summary="解析上传文件",
        auto_register=False,
    ),
    "manage_upload": ToolMetaInfo(
        display_name="上传管理",
        group="business",
        description="产品数据上传管道: 准备→执行→回滚→预览→审计→列分析→引用创建",
        prompt_description="产品上传管理（准备/执行/回滚/预览/审计）",
        summary="管理上传",
        auto_register=False,
    ),
}


def register(registry, ctx=None):
    """Auto-discovery compatible — only registers if upload context exists."""
    if has_upload_context(ctx) or (ctx and ctx.file_bytes):
        create_data_upload_tools(registry, ctx)


class _NoChangeSignal(Exception):
    """Internal signal: raised inside a savepoint to rollback when no fields actually changed."""
    pass


# ── Module-level helpers ───────────────────────────────────────


def _col_letter(col_num: int) -> str:
    """Convert 1-based column number to letter (A, B, ..., Z, AA, ...)."""
    from openpyxl.utils import get_column_letter
    return get_column_letter(col_num)


def _parse_price(value) -> float | None:
    """Parse a price value from cell, handling '$12.50', '12,500', etc."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return round(float(value), 2)
    s = str(value).strip()
    for ch in "$¥€£,，":
        s = s.replace(ch, "")
    s = s.strip()
    if not s:
        return None
    try:
        return round(float(Decimal(s)), 2)
    except (InvalidOperation, ValueError):
        return None


def _heuristic_column_mapping(labels: list[str], cols: list[str]) -> dict:
    """Fallback heuristic column mapping when Gemini is unavailable."""
    mapping: dict[str, str | None] = {}
    name_kw = ["name", "品名", "description", "product", "商品", "item"]
    code_kw = ["code", "sku", "コード", "品番", "item no", "item code"]
    price_kw = ["price", "単価", "unit price", "金額", "amount"]
    unit_kw = ["unit", "単位", "uom"]
    pack_kw = ["pack", "規格", "spec", "size"]
    brand_kw = ["brand", "ブランド"]
    currency_kw = ["currency", "通貨", "cur"]
    origin_kw = ["origin", "原産", "産地", "country of origin"]

    keyword_map = [
        ("product_name", name_kw),
        ("product_code", code_kw),
        ("price", price_kw),
        ("unit", unit_kw),
        ("pack_size", pack_kw),
        ("brand", brand_kw),
        ("currency", currency_kw),
        ("country_of_origin", origin_kw),
        # ID columns
        ("country_id", ["country_id"]),
        ("supplier_id", ["supplier_id"]),
        ("port_id", ["port_id"]),
        ("category_id", ["category_id", "cat_id"]),
    ]

    for label, col in zip(labels, cols):
        lower = label.lower()
        # Normalize separators: "country id", "country-id", "country.id" → "country_id"
        normalized = re.sub(r"[\s\-\.]+", "_", lower)
        for field, keywords in keyword_map:
            if field not in mapping and any(k in lower or k in normalized for k in keywords):
                mapping[field] = col
                break

    return mapping


def _get_active_batch_id(ctx) -> int | None:
    """Get active batch ID from session data."""
    return ctx.session_data.get("_upload_batch_id")


def _load_batch(ctx):
    """Load the active UploadBatch. Returns (batch, error_str)."""
    batch_id = _get_active_batch_id(ctx)
    if not batch_id:
        batch_id = _recover_batch_id(ctx)
    if not batch_id:
        return None, "Error: 没有活跃的上传批次。请先调用 parse_file 解析文件。"

    from core.models import UploadBatch
    batch = ctx.db.query(UploadBatch).filter(UploadBatch.id == batch_id).first()
    if not batch:
        return None, f"Error: 上传批次 {batch_id} 不存在"
    return batch, None


def _recover_batch_id(ctx) -> int | None:
    """Recover batch ID from DB if session_data lost it (e.g. cross-turn)."""
    session_id = ctx.pipeline_session_id
    if not session_id:
        return None
    from core.models import UploadBatch
    batch = (
        ctx.db.query(UploadBatch)
        .filter(
            UploadBatch.session_id == session_id,
            UploadBatch.status.notin_(["completed", "failed", "rolled_back"]),
        )
        .order_by(UploadBatch.created_at.desc())
        .first()
    )
    if batch:
        ctx.session_data["_upload_batch_id"] = batch.id
        return batch.id
    return None


def _recover_any_batch_id(ctx) -> int | None:
    """Recover most recent batch (including completed) for rollback availability."""
    session_id = ctx.pipeline_session_id
    if not session_id:
        return None
    from core.models import UploadBatch
    batch = (
        ctx.db.query(UploadBatch)
        .filter(
            UploadBatch.session_id == session_id,
            UploadBatch.status.notin_(["failed", "rolled_back"]),
        )
        .order_by(UploadBatch.created_at.desc())
        .first()
    )
    return batch.id if batch else None


def _make_new_match_result(sp, batch, ctx) -> dict:
    """Build a match_result dict for a staging row that has no match in the target port.

    If batch.port_id is set, check whether the product code exists in other ports
    (same country). If so, mark as 'new_at_port' and record source_product_id for
    field inheritance in execute_upload.
    """
    result = {
        "action": "new",
        "confidence": 0,
        "matched_product_id": None,
        "match_method": "none",
    }
    if batch.port_id and sp.product_code:
        from core.models import Product
        try:
            existing = (
                ctx.db.query(Product)
                .filter(
                    Product.code == sp.product_code.upper(),
                    Product.country_id == batch.country_id,
                    Product.status == True,
                )
                .first()
            )
            if existing:
                result["match_method"] = "new_at_port"
                result["source_product_id"] = existing.id
        except Exception:
            pass  # Graceful fallback — still a "new" product
    return result


def has_upload_context(ctx) -> bool:
    """Check if there's any upload batch (active or completed) for this session."""
    if _get_active_batch_id(ctx):
        return True
    if _recover_batch_id(ctx) is not None:
        return True
    return _recover_any_batch_id(ctx) is not None


def _auto_resolve_id_columns(batch, products: list[dict], column_mapping: dict, ctx) -> list[str]:
    """Auto-resolve ID columns (country_id, supplier_id, port_id, category_id)
    from column_mapping. Sets resolved IDs on batch and returns info lines."""
    from sqlalchemy import text

    id_fields = {
        "country_id": ("countries", "country_id", "country_name"),
        "supplier_id": ("suppliers", "supplier_id", "supplier_name"),
        "port_id": ("ports", "port_id", "port_name"),
        "category_id": ("categories", "category_id", None),
    }

    info_lines = []

    for field_key, (table_name, batch_id_attr, batch_name_attr) in id_fields.items():
        col = column_mapping.get(field_key)
        if not col:
            continue

        # Collect unique integer values from this column across all products
        raw_values = set()
        for p in products:
            raw_cells = p.get("raw_cells", {})
            v = raw_cells.get(col, "")
            if v is not None and str(v).strip():
                try:
                    raw_values.add(int(float(str(v).strip())))
                except (ValueError, TypeError):
                    pass

        if not raw_values:
            continue

        # Resolve IDs from DB
        id_list = sorted(raw_values)
        try:
            with ctx.db.begin_nested():
                rows = ctx.db.execute(
                    text(f"SELECT id, name FROM {table_name} WHERE id = ANY(:ids)"),
                    {"ids": id_list},
                ).fetchall()
            resolved = {r[0]: r[1] for r in rows}
        except Exception as e:
            logger.warning("Auto-resolve %s failed: %s", field_key, e)
            continue

        if len(id_list) == 1:
            # Single unique value → auto-set on batch
            single_id = id_list[0]
            name = resolved.get(single_id)
            if name:
                setattr(batch, batch_id_attr, single_id)
                if batch_name_attr:
                    setattr(batch, batch_name_attr, name)
                ctx.db.commit()
                info_lines.append(f"{table_name}: {name} (id={single_id}) [自动识别]")
            else:
                info_lines.append(f"{field_key} 列值 {single_id} 在 {table_name} 表中未找到")
        else:
            # Multiple unique values → report for Agent
            resolved_strs = []
            for vid in id_list[:10]:
                vname = resolved.get(vid)
                if vname:
                    resolved_strs.append(f"{vid}: {vname}")
                else:
                    resolved_strs.append(f"{vid}: (未找到)")
            info_lines.append(f"{field_key} 列有多个值 {{{', '.join(resolved_strs)}}}")

    return info_lines


# ── Tool Registration ─────────────────────────────────────────


def create_data_upload_tools(registry, ctx):
    """Register data upload tools onto the registry.

    If the session's batch is completed, only register rollback_batch.
    Otherwise register the full tool set.
    """
    active_batch_id = _get_active_batch_id(ctx) or _recover_batch_id(ctx)
    completed_batch_id = None
    if not active_batch_id:
        completed_batch_id = _recover_any_batch_id(ctx)

    # rollback always registered when any batch exists
    if active_batch_id or completed_batch_id:
        _register_rollback_tool(registry, ctx)

    # workflow tools only when active batch or new file
    if active_batch_id or ctx.file_bytes:
        _register_workflow_tools(registry, ctx)


def _register_rollback_tool(registry, ctx):
    """Register the rollback_batch tool."""

    @registry.tool(
        description=(
            "Rollback a completed batch import. Deletes created products and "
            "restores updated products to original values. Use when import was "
            "done with wrong parameters (e.g., wrong supplier). Only works on "
            "completed batches. Pass batch_id."
        ),
        parameters={
            "batch_id": {
                "type": "NUMBER",
                "description": "要回滚的批次ID",
            },
        },
        group="data_upload",
    )
    def rollback_batch(batch_id: int = 0) -> str:
        from core.models import UploadBatch, Product, ProductChangeLog

        if not batch_id:
            return "Error: 请提供 batch_id 参数。"
        batch_id = int(batch_id)

        batch = ctx.db.query(UploadBatch).filter(UploadBatch.id == batch_id).first()
        if not batch:
            return f"Error: 批次 {batch_id} 不存在。"
        if batch.status == "rolled_back":
            return f"Error: 批次 {batch_id} 已经被回滚过了。"
        if batch.status != "completed":
            return f"Error: 只能回滚已完成的批次（当前状态: {batch.status}）。"

        # Load changelog entries in reverse order
        logs = (
            ctx.db.query(ProductChangeLog)
            .filter(
                ProductChangeLog.batch_id == batch_id,
                ProductChangeLog.change_type.in_(["created", "updated"]),
            )
            .order_by(ProductChangeLog.id.desc())
            .all()
        )

        if not logs:
            return f"Error: 批次 {batch_id} 没有变更日志记录，无法回滚。"

        deleted_count = 0
        restored_count = 0
        failed_items = []
        user_id = batch.user_id

        for log_entry in logs:
            try:
                with ctx.db.begin_nested():
                    product = ctx.db.query(Product).filter(Product.id == log_entry.product_id).first()

                    if log_entry.change_type == "created":
                        if product:
                            ctx.db.delete(product)
                            deleted_count += 1
                        else:
                            # Product already deleted
                            deleted_count += 1

                    elif log_entry.change_type == "updated":
                        if not product:
                            failed_items.append(f"产品 id={log_entry.product_id}: 已不存在，无法恢复")
                            continue

                        field_changes = log_entry.field_changes or []
                        for change in field_changes:
                            field = change.get("field")
                            old_value = change.get("old_value")
                            if not field or field == "all":
                                continue
                            # Handle type conversion
                            if field == "price":
                                old_value = Decimal(str(old_value)) if old_value is not None else None
                            setattr(product, field, old_value)
                        restored_count += 1

                    # Audit log for rollback
                    rollback_log = ProductChangeLog(
                        product_id=log_entry.product_id,
                        batch_id=batch_id,
                        change_type="rolled_back",
                        field_changes=[{
                            "original_change_type": log_entry.change_type,
                            "original_log_id": log_entry.id,
                        }],
                        changed_by=user_id,
                    )
                    ctx.db.add(rollback_log)

            except Exception as e:
                failed_items.append(f"产品 id={log_entry.product_id}: {str(e)}")

        # Mark batch as rolled back
        batch.status = "rolled_back"
        batch.rolled_back_at = datetime.utcnow()
        batch.rolled_back_by = user_id
        ctx.db.commit()

        lines = [
            f"批次 #{batch_id} 回滚完成:",
            f"- 已删除新增产品: {deleted_count} 个",
            f"- 已恢复更新产品: {restored_count} 个",
        ]
        if failed_items:
            lines.append(f"- 回滚失败: {len(failed_items)} 个")
            for msg in failed_items[:10]:
                lines.append(f"  - {msg}")

        return "\n".join(lines)


def _register_workflow_tools(registry, ctx):
    """Register the full set of data upload workflow tools."""

    # ── Tool 1: parse_file ────────────────────────────────────

    @registry.tool(
        description=(
            "Parse an uploaded Excel/CSV file, auto-detect column mappings, "
            "and create staging rows. Call this first when the user uploads "
            "a price list or quotation. No parameters — reads from uploaded "
            "file. Returns product count, column mapping, and sample preview."
        ),
        parameters={},
        group="data_upload",
    )
    def parse_file() -> str:
        if not ctx.file_bytes:
            return "Error: 没有检测到上传的文件。请先上传一份 Excel 文件。"

        # 1. File hash
        file_hash = hashlib.sha256(ctx.file_bytes).hexdigest()

        # 2. Parse Excel
        try:
            from services.excel.excel_parser import parse_excel_file
            parsed = parse_excel_file(ctx.file_bytes)
        except Exception as e:
            return f"Error: 解析 Excel 文件失败 — {str(e)}"

        sheets = parsed.get("sheets", [])
        if not sheets:
            return "Error: Excel 文件中没有有效的工作表"

        sheet = sheets[0]
        headers = sheet.get("headers", [])
        sample_rows = sheet.get("sample_rows", [])
        data_start = sheet.get("data_start_row", 2)
        header_row = sheet.get("header_row", 1)

        if not headers:
            return "Error: 无法识别表头行"

        # 3. LLM column mapping (with heuristic fallback)
        header_labels = [h["label"] for h in headers]
        header_cols = [h["column"] for h in headers]

        sample_text = ""
        for i, row in enumerate(sample_rows[:3]):
            sample_text += f"  行{i+1}: {row}\n"

        prompt = f"""分析这份 Excel 报价单的列结构。

表头（第{header_row}行）: {json.dumps(dict(zip(header_cols, header_labels)), ensure_ascii=False)}

样例数据:
{sample_text}

请以 JSON 格式返回列映射，识别哪些列对应以下字段（没有的填 null）:
- product_name: 产品名称（英文或日文）
- product_code: 产品代码/SKU
- price: 单价/价格
- unit: 单位 (KG, EA, CS 等)
- pack_size: 包装规格
- brand: 品牌
- currency: 货币
- country_of_origin: 原产地
- country_id: 国家ID（整数，如 9）
- supplier_id: 供应商ID（整数，如 2）
- port_id: 港口ID（整数，如 19）
- category_id: 类别ID（整数，如 14）

只返回 JSON，格式如: {{"product_name": "B", "product_code": "A", "price": "D", "country_id": "E", ...}}
"""
        try:
            from core.config import settings
            import google.generativeai as genai

            genai.configure(api_key=settings.GOOGLE_API_KEY)
            model = genai.GenerativeModel("gemini-3-flash-preview")
            response = model.generate_content(prompt)
            mapping_text = response.text.strip()

            if "```" in mapping_text:
                m = re.search(r"```(?:json)?\s*(.*?)```", mapping_text, re.DOTALL)
                if m:
                    mapping_text = m.group(1).strip()

            column_mapping = json.loads(mapping_text)
        except Exception as e:
            logger.warning("Gemini column mapping failed: %s, falling back to heuristic", e)
            column_mapping = _heuristic_column_mapping(header_labels, header_cols)

        # 4. Extract products with openpyxl
        try:
            from openpyxl import load_workbook
            wb = load_workbook(BytesIO(ctx.file_bytes), data_only=True)
            ws = wb.worksheets[0]

            products = []
            for row_idx, row in enumerate(ws.iter_rows(min_row=data_start, values_only=False), start=1):
                cells = {}
                for cell in row:
                    if cell.column is not None:
                        cells[_col_letter(cell.column)] = cell.value

                name_col = column_mapping.get("product_name")
                name = str(cells.get(name_col, "") or "").strip() if name_col else ""
                if not name:
                    continue

                code_col = column_mapping.get("product_code")
                price_col = column_mapping.get("price")
                unit_col = column_mapping.get("unit")
                pack_col = column_mapping.get("pack_size")
                brand_col = column_mapping.get("brand")
                currency_col = column_mapping.get("currency")
                origin_col = column_mapping.get("country_of_origin")

                products.append({
                    "row_number": row_idx,
                    "product_name": name,
                    "product_code": str(cells.get(code_col, "") or "").strip() if code_col else "",
                    "price": _parse_price(cells.get(price_col)) if price_col else None,
                    "unit": str(cells.get(unit_col, "") or "").strip() if unit_col else "",
                    "pack_size": str(cells.get(pack_col, "") or "").strip() if pack_col else "",
                    "brand": str(cells.get(brand_col, "") or "").strip() if brand_col else "",
                    "currency": str(cells.get(currency_col, "") or "").strip() if currency_col else "",
                    "country_of_origin": str(cells.get(origin_col, "") or "").strip() if origin_col else "",
                    "raw_cells": {k: str(v) if v is not None else "" for k, v in cells.items()},
                })

            wb.close()
        except Exception as e:
            return f"Error: 提取产品数据失败 — {str(e)}"

        if not products:
            return "Error: 未能从文件中提取到任何产品数据"

        # 5. Create UploadBatch
        from core.models import UploadBatch, StagingProduct

        file_name = getattr(ctx, "file_name", None) or "uploaded.xlsx"
        user_id = getattr(ctx, "user_id", 0) or 0
        session_id = ctx.pipeline_session_id or ""

        batch = UploadBatch(
            session_id=session_id,
            user_id=user_id,
            file_name=file_name,
            file_hash=file_hash,
            status="staging",
            column_mapping=column_mapping,
            summary={"header_labels": dict(zip(header_cols, header_labels))},
        )
        ctx.db.add(batch)
        ctx.db.flush()  # Get batch.id

        # 6. Bulk insert StagingProduct rows
        for i, p in enumerate(products):
            staging = StagingProduct(
                batch_id=batch.id,
                row_number=p["row_number"],
                raw_data=p.get("raw_cells", {}),
                product_name=p["product_name"][:200],
                product_code=(p["product_code"] or "")[:100] or None,
                price=Decimal(str(p["price"])) if p.get("price") is not None else None,
                unit=(p["unit"] or "")[:50] or None,
                pack_size=(p["pack_size"] or "")[:100] or None,
                brand=(p["brand"] or "")[:100] or None,
                currency=(p["currency"] or "")[:20] or None,
                country_of_origin=(p["country_of_origin"] or "")[:100] or None,
                validation_status="pending",
            )
            ctx.db.add(staging)
            if (i + 1) % 100 == 0:
                ctx.db.flush()

        ctx.db.commit()

        # 7. Store batch ID in session data
        ctx.session_data["_upload_batch_id"] = batch.id

        # 7b. Auto-resolve ID columns from column_mapping
        id_resolve_info = _auto_resolve_id_columns(batch, products, column_mapping, ctx)

        # 8. Build summary
        with_price = sum(1 for p in products if p.get("price") is not None)
        with_code = sum(1 for p in products if p.get("product_code"))

        lines = [
            f"文件解析完成（批次 #{batch.id}）:",
            f"- 工作表: {sheet['name']}",
            f"- 产品数量: {len(products)} 个",
            f"- 有产品代码: {with_code} 个",
            f"- 有价格: {with_price} 个",
            f"- 列映射: {json.dumps(column_mapping, ensure_ascii=False)}",
        ]

        # Show auto-resolved ID columns
        if id_resolve_info:
            lines.append("")
            for info_line in id_resolve_info:
                lines.append(f"- {info_line}")

        lines.append("")
        lines.append("前 5 个产品预览:")
        for p in products[:5]:
            price_str = f"${p['price']}" if p.get("price") is not None else "无价格"
            lines.append(f"  行{p['row_number']}: {p['product_name']} [{p.get('product_code', '')}] {price_str}")

        return "\n".join(lines)

    # ── Tool 1b: analyze_columns ──────────────────────────────

    @registry.tool(
        description=(
            "Analyze unmapped columns by cross-referencing values against "
            "DB reference tables (suppliers, countries, ports, categories). "
            "Call immediately after parse_file to detect hidden reference "
            "columns like supplier_id or country_id. Critical for preventing "
            "wrong-supplier imports. No parameters needed."
        ),
        parameters={},
        group="data_upload",
    )
    def analyze_columns() -> str:
        from sqlalchemy import text
        from core.models import StagingProduct

        batch, err = _load_batch(ctx)
        if err:
            return err

        # Load column mapping from batch
        col_mapping = batch.column_mapping or {}
        mapped_cols = set(v for v in col_mapping.values() if v)

        # Load first 200 staging rows' raw_data
        staging_rows = (
            ctx.db.query(StagingProduct)
            .filter(StagingProduct.batch_id == batch.id)
            .order_by(StagingProduct.row_number)
            .limit(200)
            .all()
        )
        if not staging_rows:
            return "Error: 暂存表中没有数据。"

        # Reference tables to check against: (table_name, field_name)
        ref_tables = [
            ("suppliers", "supplier_id"),
            ("countries", "country_id"),
            ("ports", "port_id"),
            ("categories", "category_id"),
        ]

        lines = [f"## 列分析（批次 #{batch.id}）\n"]
        found_supplier_col = False

        # ── Report already-mapped ID columns ──
        id_fields = ["country_id", "supplier_id", "port_id", "category_id"]
        mapped_id_cols = {f: col_mapping[f] for f in id_fields if col_mapping.get(f)}
        if mapped_id_cols:
            lines.append("### 已映射的 ID 列\n")
            for field, col in mapped_id_cols.items():
                if "supplier" in field:
                    found_supplier_col = True
                # Collect unique values for this column
                id_values = set()
                for sp in staging_rows:
                    raw = sp.raw_data or {}
                    v = raw.get(col, "")
                    if v is not None and str(v).strip():
                        try:
                            id_values.add(int(float(str(v).strip())))
                        except (ValueError, TypeError):
                            pass

                if id_values:
                    table_name = next((t for t, f in ref_tables if f == field), None)
                    if table_name:
                        try:
                            with ctx.db.begin_nested():
                                rows = ctx.db.execute(
                                    text(f"SELECT id, name FROM {table_name} WHERE id = ANY(:ids)"),
                                    {"ids": sorted(id_values)},
                                ).fetchall()
                            resolved = {r[0]: r[1] for r in rows}
                            display = ", ".join(f"{vid}={resolved.get(vid, '?')}" for vid in sorted(id_values)[:10])
                            lines.append(f"- Column {col} → {field}: {display}")
                        except Exception:
                            lines.append(f"- Column {col} → {field}: 值={sorted(id_values)[:10]}")
                    else:
                        lines.append(f"- Column {col} → {field}: 值={sorted(id_values)[:10]}")
                else:
                    lines.append(f"- Column {col} → {field}: (无有效值)")
            lines.append("")

        # Also count batch-level resolved supplier as found
        if batch.supplier_id:
            found_supplier_col = True

        # ── Analyze unmapped columns ──
        all_cols = set()
        for sp in staging_rows:
            raw = sp.raw_data or {}
            all_cols.update(raw.keys())

        unmapped_cols = sorted(all_cols - mapped_cols)
        if not unmapped_cols:
            if not found_supplier_col:
                lines.append("⚠️ 未检测到 supplier 列。请确认数据中的供应商信息。")
            return "\n".join(lines)

        # Gather unique values per unmapped column
        col_values: dict[str, list] = {}
        for col in unmapped_cols:
            values = []
            for sp in staging_rows:
                raw = sp.raw_data or {}
                v = raw.get(col, "")
                if v is not None and str(v).strip():
                    values.append(str(v).strip())
            col_values[col] = values

        # Load header_labels from batch.summary for header-based detection
        header_labels = (batch.summary or {}).get("header_labels", {})

        lines.append("### 未映射列分析\n")

        for col in unmapped_cols:
            values = col_values[col]
            if not values:
                continue

            unique_values = list(dict.fromkeys(values))  # preserve order, dedupe
            col_findings = []

            # ── Header-based detection (primary signal) ──
            header = header_labels.get(col, "").lower().replace("*", "").strip()
            header_match = None
            if header:
                for table_name, field_name in ref_tables:
                    # Match "country_id", "supplier_id", etc. or stripped form like "country", "supplier"
                    base_name = field_name.replace("_id", "")
                    if field_name in header or (base_name == header):
                        header_match = (table_name, field_name)
                        break

            if header_match:
                table_name, field_name = header_match
                if "supplier" in field_name:
                    found_supplier_col = True

                # Resolve values from DB
                int_values = []
                for v in unique_values:
                    try:
                        int_values.append(int(float(v)))
                    except (ValueError, TypeError):
                        pass

                if int_values:
                    id_set = list(set(int_values))
                    try:
                        with ctx.db.begin_nested():
                            rows = ctx.db.execute(
                                text(f"SELECT id, name FROM {table_name} WHERE id = ANY(:ids)"),
                                {"ids": id_set},
                            ).fetchall()
                        matched = {r[0]: r[1] for r in rows}
                        display = ", ".join(f"{vid}={matched.get(vid, '?')}" for vid in id_set[:10])
                        lines.append(f"### Column {col} (列头: \"{header_labels.get(col, '')}\") → {field_name}")
                        lines.append(f"  已解析: {display}")
                    except Exception:
                        lines.append(f"### Column {col} (列头: \"{header_labels.get(col, '')}\") → {field_name}")
                        lines.append(f"  值: {id_set[:10]}")
                else:
                    lines.append(f"### Column {col} (列头: \"{header_labels.get(col, '')}\") → {field_name}")
                    lines.append(f"  值预览: {unique_values[:8]}")
                lines.append("")
                continue

            # ── Value-based detection (fallback) ──
            # Try integer detection
            int_values = []
            for v in unique_values:
                try:
                    int_values.append(int(float(v)))
                except (ValueError, TypeError):
                    pass

            int_ratio = len(int_values) / len(unique_values) if unique_values else 0

            if int_ratio >= 0.8 and int_values:
                # Integer column — check against ID-based reference tables
                id_set = list(set(int_values))

                for table_name, field_name in ref_tables:
                    try:
                        with ctx.db.begin_nested():
                            rows = ctx.db.execute(
                                text(f"SELECT id, name FROM {table_name} WHERE id = ANY(:ids)"),
                                {"ids": id_set},
                            ).fetchall()
                        matched = {r[0]: r[1] for r in rows}
                        hit_rate = len(matched) / len(id_set) if id_set else 0
                        if hit_rate >= 0.5:
                            col_findings.append((field_name, hit_rate, matched, id_set))
                            if "supplier" in field_name:
                                found_supplier_col = True
                    except Exception as e:
                        logger.warning("analyze_columns %s lookup failed: %s", table_name, e)

            else:
                # String column — check against name-based reference tables
                sample_names = unique_values[:20]
                like_patterns = [f"%{n}%" for n in sample_names]

                for table_name, field_name in ref_tables:
                    try:
                        with ctx.db.begin_nested():
                            rows = ctx.db.execute(
                                text(f"SELECT id, name FROM {table_name} WHERE name ILIKE ANY(:names)"),
                                {"names": like_patterns},
                            ).fetchall()
                        if rows:
                            matched_names = set()
                            for r in rows:
                                for sn in sample_names:
                                    if sn.lower() in (r[1] or "").lower():
                                        matched_names.add(sn)
                            hit_rate = len(matched_names) / len(sample_names) if sample_names else 0
                            if hit_rate >= 0.5:
                                matched = {r[0]: r[1] for r in rows}
                                col_findings.append((field_name, hit_rate, matched, sample_names))
                                if "supplier" in field_name:
                                    found_supplier_col = True
                    except Exception as e:
                        logger.warning("analyze_columns %s name lookup failed: %s", table_name, e)

            # Format output for this column
            if col_findings:
                col_findings.sort(key=lambda x: x[1], reverse=True)
                best = col_findings[0]
                field_name, hit_rate, matched, checked = best
                confidence = "很可能" if hit_rate >= 0.8 else "可能"

                lines.append(f"### Column {col} — {confidence}是 {field_name}（命中率 {hit_rate:.0%}）")
                display_matched = dict(list(matched.items())[:10])
                lines.append(f"  匹配: {display_matched}")
                if isinstance(checked, list):
                    if isinstance(checked[0], int):
                        unmatched = [v for v in checked if v not in matched]
                    else:
                        unmatched = checked
                else:
                    unmatched = [v for v in checked if v not in matched]
                if unmatched and isinstance(checked[0], int):
                    lines.append(f"  未匹配: {unmatched[:10]}")

                if field_name == "supplier_id" and len(matched) > 1:
                    lines.append(f"  ⚠️ 检测到多供应商数据（{len(matched)} 个供应商），请确认是否需要分别处理。")

                lines.append("")
            else:
                preview = unique_values[:8]
                lines.append(f"### Column {col} — 未匹配任何参考表")
                lines.append(f"  值预览: {preview}")
                lines.append("")

        if not found_supplier_col:
            lines.append("⚠️ 未检测到 supplier 列。请确认数据中的供应商信息。")

        return "\n".join(lines)

    # ── Tool 2: resolve_and_validate ──────────────────────────

    @registry.tool(
        description=(
            "Validate staging data with code-exact + fuzzy name + LLM matching "
            "and confidence scoring. Pass supplier_name/supplier_id, country_name/country_id, "
            "port_name/port_id, and effective dates. ID params take priority over name. "
            "Prefer prepare_upload which combines this with audit and preview."
        ),
        parameters={
            "supplier_name": {
                "type": "STRING",
                "description": "供应商名称（可选，用于筛选 DB 产品范围）",
                "required": False,
            },
            "supplier_id": {
                "type": "NUMBER",
                "description": "供应商ID（可选，优先于 supplier_name）",
                "required": False,
            },
            "country_name": {
                "type": "STRING",
                "description": "国家名称（可选，用于筛选 DB 产品范围）",
                "required": False,
            },
            "country_id": {
                "type": "NUMBER",
                "description": "国家ID（可选，优先于 country_name）",
                "required": False,
            },
            "port_name": {
                "type": "STRING",
                "description": "目标港口名称（如 横浜、Bangkok），决定更新哪个港口的价格记录",
                "required": False,
            },
            "port_id": {
                "type": "NUMBER",
                "description": "港口ID（可选，优先于 port_name）",
                "required": False,
            },
            "effective_from": {
                "type": "STRING",
                "description": "价格生效开始日期 YYYY-MM-DD",
                "required": False,
            },
            "effective_to": {
                "type": "STRING",
                "description": "价格生效结束日期 YYYY-MM-DD",
                "required": False,
            },
        },
        group="data_upload",
    )
    def resolve_and_validate(supplier_name: str = "", supplier_id: int = 0, country_name: str = "", country_id: int = 0, port_name: str = "", port_id: int = 0, effective_from: str = "", effective_to: str = "") -> str:
        from sqlalchemy import text
        from core.models import UploadBatch, StagingProduct, Product

        batch, err = _load_batch(ctx)
        if err:
            return err

        # ── Required field validation ──
        # Check both batch state (from previous calls) and current params
        has_country = batch.country_id or country_id or country_name.strip()
        has_port = batch.port_id or port_id or port_name.strip()
        if has_country and not has_port:
            return "Error: 请提供目标港口名称（port_name 参数）。例如：横浜、Bangkok 等。这决定了更新哪个港口的价格记录。"
        if not batch.effective_from and not effective_from.strip():
            return "Error: 请提供价格生效开始日期（effective_from 参数）。格式：YYYY-MM-DD。"
        if not batch.effective_to and not effective_to.strip():
            return "Error: 请提供价格生效结束日期（effective_to 参数）。格式：YYYY-MM-DD。"

        staging_rows = (
            ctx.db.query(StagingProduct)
            .filter(StagingProduct.batch_id == batch.id)
            .order_by(StagingProduct.row_number)
            .all()
        )
        if not staging_rows:
            return "Error: 暂存表中没有数据"

        # ── Phase A: Resolve supplier & country (ID params take priority) ──
        if supplier_id and not batch.supplier_id:
            supplier_id = int(supplier_id)
            try:
                with ctx.db.begin_nested():
                    row = ctx.db.execute(
                        text("SELECT id, name FROM suppliers WHERE id = :id"),
                        {"id": supplier_id},
                    ).fetchone()
                if row:
                    batch.supplier_id = row[0]
                    batch.supplier_name = row[1]
            except Exception as e:
                logger.warning("Supplier ID lookup failed: %s", e)
        elif supplier_name.strip() and not batch.supplier_id:
            try:
                sp_savepoint = ctx.db.begin_nested()
                rows = ctx.db.execute(
                    text("SELECT id, name FROM suppliers WHERE name ILIKE :p LIMIT 5"),
                    {"p": f"%{supplier_name.strip()}%"},
                ).fetchall()
                if rows:
                    batch.supplier_id = rows[0][0]
                    batch.supplier_name = rows[0][1]
                sp_savepoint.commit()
            except Exception as e:
                sp_savepoint.rollback()
                logger.warning("Supplier lookup failed: %s", e)

        if country_id and not batch.country_id:
            country_id = int(country_id)
            try:
                with ctx.db.begin_nested():
                    row = ctx.db.execute(
                        text("SELECT id, name FROM countries WHERE id = :id"),
                        {"id": country_id},
                    ).fetchone()
                if row:
                    batch.country_id = row[0]
                    batch.country_name = row[1]
            except Exception as e:
                logger.warning("Country ID lookup failed: %s", e)
        elif country_name.strip() and not batch.country_id:
            try:
                ct_savepoint = ctx.db.begin_nested()
                rows = ctx.db.execute(
                    text("SELECT id, name FROM countries WHERE name ILIKE :p OR code ILIKE :p LIMIT 5"),
                    {"p": f"%{country_name.strip()}%"},
                ).fetchall()
                if rows:
                    batch.country_id = rows[0][0]
                    batch.country_name = rows[0][1]
                ct_savepoint.commit()
            except Exception as e:
                ct_savepoint.rollback()
                logger.warning("Country lookup failed: %s", e)

        # ── Port lookup (ID takes priority) ──
        if port_id and not batch.port_id:
            port_id = int(port_id)
            try:
                with ctx.db.begin_nested():
                    row = ctx.db.execute(
                        text("SELECT id, name FROM ports WHERE id = :id"),
                        {"id": port_id},
                    ).fetchone()
                if row:
                    batch.port_id = row[0]
                    batch.port_name = row[1]
            except Exception as e:
                logger.warning("Port ID lookup failed: %s", e)
        elif port_name.strip() and not batch.port_id:
            try:
                pt_savepoint = ctx.db.begin_nested()
                rows = ctx.db.execute(
                    text("SELECT id, name FROM ports WHERE name ILIKE :p LIMIT 5"),
                    {"p": f"%{port_name.strip()}%"},
                ).fetchall()
                if rows:
                    batch.port_id = rows[0][0]
                    batch.port_name = rows[0][1]
                pt_savepoint.commit()
            except Exception as e:
                pt_savepoint.rollback()
                logger.warning("Port lookup failed: %s", e)

        # ── Effective dates ──
        if effective_from.strip() and not batch.effective_from:
            try:
                batch.effective_from = datetime.strptime(effective_from.strip(), "%Y-%m-%d").date()
            except ValueError:
                pass
        if effective_to.strip() and not batch.effective_to:
            try:
                batch.effective_to = datetime.strptime(effective_to.strip(), "%Y-%m-%d").date()
            except ValueError:
                pass

        # ── Load DB products for matching ──
        query = ctx.db.query(Product).filter(Product.status == True)
        if batch.supplier_id:
            query = query.filter(Product.supplier_id == batch.supplier_id)
        if batch.country_id:
            query = query.filter(Product.country_id == batch.country_id)
        if batch.port_id:
            query = query.filter(Product.port_id == batch.port_id)
        db_products = query.all()

        # Index by code and name
        by_code: dict[str, Product] = {}
        by_name: dict[str, Product] = {}
        for dbp in db_products:
            if dbp.code:
                by_code[dbp.code.upper()] = dbp
            if dbp.product_name_en:
                by_name[dbp.product_name_en.upper()] = dbp

        # ── Phase A: Code-level matching ──
        unmatched_rows = []
        stats = {"new": 0, "update": 0, "no_change": 0}

        for sp in staging_rows:
            code = (sp.product_code or "").upper()
            name = (sp.product_name or "").upper()
            new_price = float(sp.price) if sp.price is not None else None

            matched_db = None
            confidence = 0.0
            match_method = ""

            # 1. Exact code
            if code and code in by_code:
                matched_db = by_code[code]
                confidence = 1.0
                match_method = "code_exact"
            # 2. Exact name
            elif name and name in by_name:
                matched_db = by_name[name]
                confidence = 0.95
                match_method = "name_exact"
            # 3. Fuzzy name
            else:
                best_score = 0.0
                for dbp in db_products:
                    if not dbp.product_name_en:
                        continue
                    sim = SequenceMatcher(None, name, dbp.product_name_en.upper()).ratio()
                    if sim > best_score and sim >= 0.6:
                        best_score = sim
                        matched_db = dbp
                        match_method = "name_fuzzy"
                confidence = best_score

            if matched_db and confidence >= 0.6:
                old_price = float(matched_db.price) if matched_db.price is not None else None
                price_change_pct = None
                action = "update"

                if new_price is not None and old_price is not None and old_price > 0:
                    price_change_pct = round((new_price - old_price) / old_price * 100, 1)
                    if abs(price_change_pct) < 0.01:
                        action = "no_change"
                elif new_price is None and old_price is not None:
                    action = "no_change"

                sp.match_result = {
                    "action": action,
                    "confidence": round(confidence, 3),
                    "matched_product_id": matched_db.id,
                    "match_method": match_method,
                    "old_price": old_price,
                    "price_change_pct": price_change_pct,
                    "db_product_name": matched_db.product_name_en,
                }
                sp.validation_status = "quarantined" if confidence < 0.7 else "valid"
                stats[action] += 1
            else:
                # No match or very low confidence → mark for LLM
                if confidence < 0.6:
                    matched_db = None
                    confidence = 0.0
                unmatched_rows.append(sp)

        # ── Phase B: LLM fuzzy matching for unmatched items ──
        if unmatched_rows and db_products:
            try:
                llm_results = _llm_fuzzy_match(unmatched_rows, db_products, batch, ctx)
                for sp in unmatched_rows:
                    row_key = str(sp.row_number)
                    if row_key in llm_results:
                        lr = llm_results[row_key]
                        matched_id = lr.get("matched_product_id")
                        llm_conf = lr.get("confidence", 0)

                        if matched_id and llm_conf >= 0.5:
                            # Find the matched product
                            matched_product = None
                            for dbp in db_products:
                                if dbp.id == matched_id:
                                    matched_product = dbp
                                    break

                            if matched_product:
                                old_price = float(matched_product.price) if matched_product.price is not None else None
                                new_price = float(sp.price) if sp.price is not None else None
                                price_change_pct = None
                                action = "update"

                                if new_price is not None and old_price is not None and old_price > 0:
                                    price_change_pct = round((new_price - old_price) / old_price * 100, 1)
                                    if abs(price_change_pct) < 0.01:
                                        action = "no_change"

                                sp.match_result = {
                                    "action": action,
                                    "confidence": round(llm_conf, 3),
                                    "matched_product_id": matched_product.id,
                                    "match_method": "llm_fuzzy",
                                    "old_price": old_price,
                                    "price_change_pct": price_change_pct,
                                    "db_product_name": matched_product.product_name_en,
                                }
                                sp.validation_status = "quarantined" if llm_conf < 0.7 else "valid"
                                stats[action] += 1
                                continue

                    # Still unmatched
                    if not sp.match_result:
                        sp.match_result = _make_new_match_result(sp, batch, ctx)
                        sp.validation_status = "valid"
                        stats["new"] += 1
            except Exception as e:
                logger.warning("LLM fuzzy match failed: %s, marking remaining as new", e)
                for sp in unmatched_rows:
                    if not sp.match_result:
                        sp.match_result = _make_new_match_result(sp, batch, ctx)
                        sp.validation_status = "valid"
                        stats["new"] += 1
        else:
            # No DB products to match against — all new
            for sp in unmatched_rows:
                sp.match_result = _make_new_match_result(sp, batch, ctx)
                sp.validation_status = "valid"
                stats["new"] += 1

        # Update resolved IDs on staging rows
        for sp in staging_rows:
            sp.resolved_supplier_id = batch.supplier_id
            sp.resolved_country_id = batch.country_id

        batch.status = "validating"
        ctx.db.commit()

        # Build report
        total = len(staging_rows)
        lines = [
            f"## 验证结果（批次 #{batch.id}）",
            f"- 总行数: {total}",
            f"- 新增: {stats['new']}",
            f"- 更新: {stats['update']}",
            f"- 无变化: {stats['no_change']}",
        ]

        if batch.supplier_name:
            lines.append(f"- 供应商: {batch.supplier_name} (id={batch.supplier_id})")
        elif supplier_name:
            lines.append(f"- 未找到供应商 '{supplier_name}'（可能需要创建）")

        if batch.country_name:
            lines.append(f"- 国家: {batch.country_name} (id={batch.country_id})")
        elif country_name:
            lines.append(f"- 未找到国家 '{country_name}'（可能需要创建）")

        if batch.port_name:
            lines.append(f"- 港口: {batch.port_name} (id={batch.port_id})")
        elif port_name:
            lines.append(f"- 未找到港口 '{port_name}'")

        if batch.effective_from or batch.effective_to:
            lines.append(f"- 有效期: {batch.effective_from or '?'} ~ {batch.effective_to or '?'}")

        # Cross-port new count
        new_at_port = sum(
            1 for sp in staging_rows
            if sp.match_result and sp.match_result.get("match_method") == "new_at_port"
        )
        if new_at_port:
            lines.append(f"- 港口新增(其他港口已有): {new_at_port}")

        # Confidence distribution
        high = sum(1 for sp in staging_rows if sp.match_result and sp.match_result.get("confidence", 0) >= 0.9)
        mid = sum(1 for sp in staging_rows if sp.match_result and 0.7 <= sp.match_result.get("confidence", 0) < 0.9)
        low = sum(1 for sp in staging_rows if sp.match_result and 0 < sp.match_result.get("confidence", 0) < 0.7)
        none_ = sum(1 for sp in staging_rows if sp.match_result and sp.match_result.get("confidence", 0) == 0)

        lines.append("")
        lines.append("### 置信度分布")
        lines.append(f"- 高 (≥90%): {high}")
        lines.append(f"- 中 (70-89%): {mid}")
        lines.append(f"- 低 (<70%): {low}")
        lines.append(f"- 新增(无匹配): {none_}")

        # Show quarantined items
        quarantined = [sp for sp in staging_rows if sp.validation_status == "quarantined"]
        if quarantined:
            lines.append("")
            lines.append(f"### 需确认项 ({len(quarantined)} 个)")
            for sp in quarantined[:10]:
                mr = sp.match_result or {}
                conf = mr.get("confidence", 0)
                action = mr.get("action", "?")
                db_name = mr.get("db_product_name", "")
                pct = mr.get("price_change_pct")
                pct_str = f" ({pct:+.1f}%)" if pct is not None else ""
                lines.append(f"  行{sp.row_number}: {sp.product_name} → {db_name} [置信度:{conf:.0%}, {action}{pct_str}]")

        # Build structured data for frontend
        structured = {
            "card_type": "upload_validation",
            "tool": "resolve_and_validate",
            "batch_id": batch.id,
            "stats": dict(stats),
            "total": total,
            "supplier": {"name": batch.supplier_name, "id": batch.supplier_id},
            "country": {"name": batch.country_name, "id": batch.country_id},
            "port": {"name": batch.port_name, "id": batch.port_id},
            "effective_from": str(batch.effective_from) if batch.effective_from else None,
            "effective_to": str(batch.effective_to) if batch.effective_to else None,
            "confidence": {"high": high, "mid": mid, "low": low, "new": none_},
            "quarantined": [
                {
                    "row": sp.row_number,
                    "name": sp.product_name,
                    "code": sp.product_code,
                    "db_name": (sp.match_result or {}).get("db_product_name", ""),
                    "confidence": (sp.match_result or {}).get("confidence", 0),
                    "action": (sp.match_result or {}).get("action", "?"),
                    "price_change_pct": (sp.match_result or {}).get("price_change_pct"),
                }
                for sp in quarantined[:20]
            ] if quarantined else [],
            "missing_supplier": bool(supplier_name and not batch.supplier_id),
            "missing_country": bool(country_name and not batch.country_id),
        }
        text_report = "\n".join(lines)
        return text_report + "\n__STRUCTURED__\n" + json.dumps(structured, ensure_ascii=False)

    # ── Tool 3: create_references ─────────────────────────────

    @registry.tool(
        description=(
            "Create missing supplier or country entities in the database. "
            "Pass a JSON string with entities to create. Call when "
            "prepare_upload reports missing_supplier or missing_country, "
            "then re-run prepare_upload."
        ),
        parameters={
            "entities": {
                "type": "STRING",
                "description": 'JSON 格式，如 {"suppliers": [{"name": "ABC Co"}], "countries": [{"name": "Thailand", "code": "TH"}]}',
                "required": False,
            },
        },
        group="data_upload",
    )
    def create_references(entities: str = "") -> str:
        from core.models import UploadBatch, StagingProduct, Country, Supplier
        from sqlalchemy import text

        batch, err = _load_batch(ctx)
        if err:
            return err

        created = []

        # Parse entities parameter
        specs: dict = {}
        if entities.strip():
            try:
                specs = json.loads(entities)
            except json.JSONDecodeError:
                return "Error: entities 参数不是有效的 JSON"

        # ── Countries ──
        countries_to_create = specs.get("countries", [])
        for c in countries_to_create:
            name = c.get("name", "").strip()
            if not name:
                continue
            # Check if already exists
            existing = ctx.db.execute(
                text("SELECT id FROM countries WHERE name ILIKE :n LIMIT 1"),
                {"n": name},
            ).fetchone()
            if existing:
                created.append(f"国家 '{name}' 已存在 (id={existing[0]})")
                if not batch.country_id:
                    batch.country_id = existing[0]
                    batch.country_name = name
                continue

            new_country = Country(
                name=name,
                code=c.get("code", ""),
                status=True,
            )
            ctx.db.add(new_country)
            ctx.db.flush()
            batch.country_id = new_country.id
            batch.country_name = name
            created.append(f"已创建国家: {name} (id={new_country.id})")

        # ── Suppliers ──
        suppliers_to_create = specs.get("suppliers", [])
        for s in suppliers_to_create:
            name = s.get("name", "").strip()
            if not name:
                continue
            existing = ctx.db.execute(
                text("SELECT id FROM suppliers WHERE name ILIKE :n LIMIT 1"),
                {"n": name},
            ).fetchone()
            if existing:
                created.append(f"供应商 '{name}' 已存在 (id={existing[0]})")
                if not batch.supplier_id:
                    batch.supplier_id = existing[0]
                    batch.supplier_name = name
                continue

            new_supplier = Supplier(
                name=name,
                country_id=batch.country_id,
                status=True,
            )
            ctx.db.add(new_supplier)
            ctx.db.flush()
            batch.supplier_id = new_supplier.id
            batch.supplier_name = name
            created.append(f"已创建供应商: {name} (id={new_supplier.id})")

        # Update resolved IDs on staging rows
        if batch.supplier_id or batch.country_id:
            staging_rows = (
                ctx.db.query(StagingProduct)
                .filter(StagingProduct.batch_id == batch.id)
                .all()
            )
            for sp in staging_rows:
                if batch.supplier_id:
                    sp.resolved_supplier_id = batch.supplier_id
                if batch.country_id:
                    sp.resolved_country_id = batch.country_id

        ctx.db.commit()

        if not created:
            return "没有需要创建的引用数据。"

        return "引用数据创建完成:\n" + "\n".join(f"- {c}" for c in created)

    # ── Tool 4: preview_changes ───────────────────────────────

    @registry.tool(
        description=(
            "Generate a change preview report (new/update/no-change). "
            "Prefer prepare_upload which includes this step. Only call "
            "directly for re-preview without re-running validation."
        ),
        parameters={},
        group="data_upload",
    )
    def preview_changes() -> str:
        from core.models import StagingProduct

        batch, err = _load_batch(ctx)
        if err:
            return err

        staging_rows = (
            ctx.db.query(StagingProduct)
            .filter(
                StagingProduct.batch_id == batch.id,
                StagingProduct.validation_status.in_(["valid", "quarantined"]),
            )
            .order_by(StagingProduct.row_number)
            .all()
        )

        if not staging_rows:
            return "Error: 没有可预览的数据（所有行都无效）"

        # Group by action
        groups: dict[str, list] = {"new": [], "update": [], "no_change": []}
        for sp in staging_rows:
            mr = sp.match_result or {}
            action = mr.get("action", "new")
            groups.setdefault(action, []).append(sp)

        batch.status = "previewing"
        ctx.db.commit()

        lines = [
            f"## 变更预览（批次 #{batch.id}）",
            f"- 供应商: {batch.supplier_name or '未设置'}" + (f" (id={batch.supplier_id})" if batch.supplier_id else ""),
            f"- 国家: {batch.country_name or '未设置'}" + (f" (id={batch.country_id})" if batch.country_id else ""),
            f"- 港口: {batch.port_name or '未设置'}" + (f" (id={batch.port_id})" if batch.port_id else ""),
            f"- 有效期: {batch.effective_from or '?'} ~ {batch.effective_to or '?'}",
            "",
            f"### 统计",
            f"- 新增: {len(groups['new'])} 个",
            f"- 更新: {len(groups['update'])} 个",
            f"- 无变化: {len(groups['no_change'])} 个",
        ]

        if groups["new"]:
            lines.append("")
            lines.append(f"### 新增产品 (前10/{len(groups['new'])})")
            for sp in groups["new"][:10]:
                price_str = f"${sp.price}" if sp.price is not None else "无价格"
                lines.append(f"  行{sp.row_number}: {sp.product_name} [{sp.product_code or ''}] {price_str}")

        if groups["update"]:
            lines.append("")
            lines.append(f"### 价格更新 (前10/{len(groups['update'])})")
            for sp in groups["update"][:10]:
                mr = sp.match_result or {}
                old = mr.get("old_price", "?")
                new = sp.price
                pct = mr.get("price_change_pct")
                pct_str = f" ({pct:+.1f}%)" if pct is not None else ""
                lines.append(f"  行{sp.row_number}: {sp.product_name} ${old} → ${new}{pct_str}")

        lines.append("")
        lines.append("如需排除某些行，请告知行号。确认后调用 execute_upload 执行。")

        # Build structured data for frontend
        structured = {
            "card_type": "upload_preview",
            "tool": "preview_changes",
            "batch_id": batch.id,
            "supplier": {"name": batch.supplier_name, "id": batch.supplier_id},
            "country": {"name": batch.country_name, "id": batch.country_id},
            "port": {"name": batch.port_name, "id": batch.port_id},
            "effective_from": str(batch.effective_from) if batch.effective_from else None,
            "effective_to": str(batch.effective_to) if batch.effective_to else None,
            "stats": {
                "new": len(groups["new"]),
                "update": len(groups["update"]),
                "no_change": len(groups["no_change"]),
            },
            "new_items": [
                {"row": sp.row_number, "name": sp.product_name, "code": sp.product_code, "price": float(sp.price) if sp.price is not None else None}
                for sp in groups["new"][:15]
            ],
            "updates": [
                {
                    "row": sp.row_number,
                    "name": sp.product_name,
                    "old_price": (sp.match_result or {}).get("old_price"),
                    "new_price": float(sp.price) if sp.price is not None else None,
                    "change_pct": (sp.match_result or {}).get("price_change_pct"),
                }
                for sp in groups["update"][:15]
            ],
        }
        text_report = "\n".join(lines)
        return text_report + "\n__STRUCTURED__\n" + json.dumps(structured, ensure_ascii=False)

    # ── Tool 5: execute_upload ────────────────────────────────

    @registry.tool(
        description=(
            "Execute product import atomically: insert new products, update "
            "existing ones, write change logs. Only call AFTER user explicitly "
            "confirms the preview. Pass exclude_rows to skip rows. Use "
            "rollback_batch to undo if import was wrong."
        ),
        parameters={
            "exclude_rows": {
                "type": "STRING",
                "description": "要排除的行号，逗号分隔（如 '3,7,12'）。不传则全部执行。",
                "required": False,
            },
        },
        group="data_upload",
    )
    def execute_upload(exclude_rows: str = "") -> str:
        from core.models import StagingProduct, Product, ProductChangeLog

        batch, err = _load_batch(ctx)
        if err:
            return err

        if batch.status not in ("previewing", "validating"):
            return f"Error: 批次状态为 '{batch.status}'，需要先调用 preview_changes 预览变更"

        # Parse excluded rows
        excluded = set()
        if exclude_rows.strip():
            for part in exclude_rows.split(","):
                part = part.strip()
                if part.isdigit():
                    excluded.add(int(part))

        staging_rows = (
            ctx.db.query(StagingProduct)
            .filter(
                StagingProduct.batch_id == batch.id,
                StagingProduct.validation_status.in_(["valid", "quarantined"]),
            )
            .order_by(StagingProduct.row_number)
            .all()
        )

        batch.status = "executing"
        ctx.db.flush()

        inserted = 0
        updated = 0
        skipped = 0
        excluded_count = 0
        failed = []
        user_id = batch.user_id

        try:
            for sp in staging_rows:
                if sp.row_number in excluded:
                    excluded_count += 1
                    continue

                mr = sp.match_result or {}
                action = mr.get("action", "new")
                matched_id = mr.get("matched_product_id")

                # Route by whether a DB match exists, not by action label.
                # This ensures products with action="no_change" (set by price-only
                # check in resolve_and_validate) still go through full field
                # comparison for dates, unit, pack_size, brand, currency.
                # _NoChangeSignal handles truly unchanged items.

                if action == "new" and not matched_id:
                    # INSERT new product — use savepoint so one failure doesn't kill the session
                    try:
                        with ctx.db.begin_nested():
                            # Cross-port: inherit fields from existing product in other port
                            source_id = mr.get("source_product_id")
                            source_product = None
                            if source_id:
                                source_product = ctx.db.query(Product).filter(Product.id == source_id).first()

                            new_product = Product(
                                product_name_en=sp.product_name,
                                product_name_jp=source_product.product_name_jp if source_product else None,
                                code=sp.product_code or None,
                                price=sp.price,
                                unit=sp.unit or None,
                                pack_size=sp.pack_size or None,
                                brand=sp.brand or None,
                                currency=sp.currency or None,
                                country_of_origin=sp.country_of_origin or None,
                                supplier_id=sp.resolved_supplier_id or batch.supplier_id or (source_product.supplier_id if source_product else None),
                                country_id=sp.resolved_country_id or batch.country_id,
                                category_id=source_product.category_id if source_product else None,
                                port_id=batch.port_id,
                                status=True,
                                effective_from=batch.effective_from or datetime.utcnow().date(),
                                effective_to=batch.effective_to,
                            )
                            ctx.db.add(new_product)
                            ctx.db.flush()

                            # Audit log
                            changelog = ProductChangeLog(
                                product_id=new_product.id,
                                batch_id=batch.id,
                                change_type="created",
                                field_changes=[{
                                    "field": "all",
                                    "old_value": None,
                                    "new_value": {
                                        "product_name_en": sp.product_name,
                                        "code": sp.product_code,
                                        "price": str(sp.price) if sp.price is not None else None,
                                    },
                                }],
                                changed_by=user_id,
                            )
                            ctx.db.add(changelog)
                        inserted += 1
                    except Exception as e:
                        failed.append(f"行{sp.row_number} {sp.product_name}: {str(e)}")

                elif matched_id:
                    # UPDATE existing product — all matched products go through
                    # full field comparison regardless of action label.
                    try:
                        with ctx.db.begin_nested():
                            db_product = ctx.db.query(Product).filter(Product.id == matched_id).first()
                            if not db_product:
                                raise ValueError(f"产品 id={matched_id} 不存在")

                            field_changes = []

                            # Price update
                            if sp.price is not None:
                                old_price = float(db_product.price) if db_product.price is not None else None
                                new_price = float(sp.price)
                                if old_price != new_price:
                                    field_changes.append({
                                        "field": "price",
                                        "old_value": old_price,
                                        "new_value": new_price,
                                    })
                                    db_product.price = sp.price

                            # Unit update
                            if sp.unit and sp.unit != db_product.unit:
                                field_changes.append({
                                    "field": "unit",
                                    "old_value": db_product.unit,
                                    "new_value": sp.unit,
                                })
                                db_product.unit = sp.unit

                            # Pack size update
                            if sp.pack_size and sp.pack_size != db_product.pack_size:
                                field_changes.append({
                                    "field": "pack_size",
                                    "old_value": db_product.pack_size,
                                    "new_value": sp.pack_size,
                                })
                                db_product.pack_size = sp.pack_size

                            # Currency update
                            if sp.currency and sp.currency != db_product.currency:
                                field_changes.append({
                                    "field": "currency",
                                    "old_value": db_product.currency,
                                    "new_value": sp.currency,
                                })
                                db_product.currency = sp.currency

                            # Brand update
                            if sp.brand and sp.brand != db_product.brand:
                                field_changes.append({
                                    "field": "brand",
                                    "old_value": db_product.brand,
                                    "new_value": sp.brand,
                                })
                                db_product.brand = sp.brand

                            # Effective dates update
                            if batch.effective_from:
                                old_from = db_product.effective_from
                                # Compare as date (effective_from in products is DateTime, batch is Date)
                                old_from_date = old_from.date() if hasattr(old_from, 'date') and old_from else old_from
                                if old_from_date != batch.effective_from:
                                    field_changes.append({
                                        "field": "effective_from",
                                        "old_value": str(old_from) if old_from else None,
                                        "new_value": str(batch.effective_from),
                                    })
                                    db_product.effective_from = batch.effective_from

                            if batch.effective_to:
                                old_to = db_product.effective_to
                                old_to_date = old_to.date() if hasattr(old_to, 'date') and old_to else old_to
                                if old_to_date != batch.effective_to:
                                    field_changes.append({
                                        "field": "effective_to",
                                        "old_value": str(old_to) if old_to else None,
                                        "new_value": str(batch.effective_to),
                                    })
                                    db_product.effective_to = batch.effective_to

                            if not field_changes:
                                # No actual changes — treat as no_change
                                raise _NoChangeSignal()

                            ctx.db.flush()

                            changelog = ProductChangeLog(
                                product_id=db_product.id,
                                batch_id=batch.id,
                                change_type="updated",
                                field_changes=field_changes,
                                changed_by=user_id,
                            )
                            ctx.db.add(changelog)
                        updated += 1
                    except _NoChangeSignal:
                        skipped += 1
                    except Exception as e:
                        failed.append(f"行{sp.row_number} {sp.product_name}: {str(e)}")

                else:
                    # No match and not new — safe fallback
                    skipped += 1

            # Commit all
            batch.status = "completed"
            batch.completed_at = datetime.utcnow()
            batch.summary = {
                "inserted": inserted,
                "updated": updated,
                "skipped": skipped,
                "excluded": excluded_count,
                "failed": len(failed),
            }
            ctx.db.commit()

        except Exception as e:
            ctx.db.rollback()
            try:
                batch.status = "failed"
                ctx.db.commit()
            except Exception:
                ctx.db.rollback()
            return f"Error: 执行失败，已全部回滚 — {str(e)}"

        # Build result
        lines = [
            f"产品上传完成（批次 #{batch.id}）:",
            f"- 新增: {inserted} 个",
            f"- 更新: {updated} 个",
            f"- 无变化跳过: {skipped} 个",
        ]
        if excluded_count:
            lines.append(f"- 手动排除: {excluded_count} 个")
        if failed:
            lines.append(f"- 失败: {len(failed)} 个")
            for f_msg in failed[:10]:
                lines.append(f"  - {f_msg}")

        # Build structured data for frontend
        structured = {
            "card_type": "upload_result",
            "tool": "execute_upload",
            "batch_id": batch.id,
            "status": "completed" if not failed else "partial",
            "stats": {
                "inserted": inserted,
                "updated": updated,
                "skipped": skipped,
                "excluded": excluded_count,
                "failed": len(failed),
            },
            "failures": failed[:10],
        }
        text_report = "\n".join(lines)
        return text_report + "\n__STRUCTURED__\n" + json.dumps(structured, ensure_ascii=False)

    # ── Tool 6: audit_data ────────────────────────────────────

    @registry.tool(
        description=(
            "Run data quality audit: structural checks (missing columns, format, "
            "duplicates) plus LLM semantic audit (price/unit/name reasonableness). "
            "Prefer prepare_upload which includes this. Call directly only for "
            "standalone audit."
        ),
        parameters={},
        group="data_upload",
    )
    def audit_data() -> str:
        from core.models import StagingProduct

        batch, err = _load_batch(ctx)
        if err:
            return err

        staging_rows = (
            ctx.db.query(StagingProduct)
            .filter(StagingProduct.batch_id == batch.id)
            .order_by(StagingProduct.row_number)
            .all()
        )
        if not staging_rows:
            return "Error: 暂存表中没有数据，请先调用 parse_file 解析文件。"

        # ── Layer 1: Code-level structural checks ──
        findings = _code_audit(batch, staging_rows)

        # ── Layer 2: LLM semantic audit ──
        llm_findings = _llm_semantic_audit(batch, staging_rows, ctx)
        if llm_findings is not None:
            # Deduplicate: skip LLM findings with same category+rows as code findings
            existing_keys = {(f["category"], tuple(f["rows"][:5])) for f in findings}
            for lf in llm_findings:
                if not isinstance(lf, dict) or not lf.get("message"):
                    continue
                # Normalize: ensure all required fields exist
                lf.setdefault("severity", "info")
                lf.setdefault("category", "other")
                if not isinstance(lf.get("rows"), list):
                    lf["rows"] = []
                lf["rows"] = [r for r in lf["rows"] if isinstance(r, (int, float))]
                lf.setdefault("suggestion", "")
                # Validate severity
                if lf["severity"] not in ("error", "warning", "info"):
                    lf["severity"] = "info"
                key = (lf["category"], tuple(lf["rows"][:5]))
                if key not in existing_keys:
                    findings.append(lf)

        # ── Stats ──
        stats = {"error": 0, "warning": 0, "info": 0}
        for f in findings:
            sev = f.get("severity", "info")
            if sev in stats:
                stats[sev] += 1

        total = len(staging_rows)
        summary_parts = []
        if stats["error"]:
            summary_parts.append(f"{stats['error']} 个错误")
        if stats["warning"]:
            summary_parts.append(f"{stats['warning']} 个警告")
        if stats["info"]:
            summary_parts.append(f"{stats['info']} 个提示")
        if not summary_parts:
            summary = f"审计通过，{total} 行数据未发现问题"
        else:
            summary = f"审计完成: {', '.join(summary_parts)}"

        # Build text report
        lines = [
            f"## 数据质量审计（批次 #{batch.id}, {total} 行）",
            "",
        ]
        for sev in ("error", "warning", "info"):
            sev_findings = [f for f in findings if f["severity"] == sev]
            if not sev_findings:
                continue
            sev_label = {"error": "错误", "warning": "警告", "info": "提示"}[sev]
            lines.append(f"### {sev_label} ({len(sev_findings)})")
            for f in sev_findings:
                rows_str = ""
                if f["rows"]:
                    rows_str = f" (行: {', '.join(str(r) for r in f['rows'][:10])}{'...' if len(f['rows']) > 10 else ''})"
                lines.append(f"- {f['message']}{rows_str}")
                if f.get("suggestion"):
                    lines.append(f"  建议: {f['suggestion']}")

        lines.append("")
        lines.append(summary)

        # Build structured data for frontend
        structured = {
            "card_type": "data_audit",
            "batch_id": batch.id,
            "total_rows": total,
            "findings": findings,
            "summary": summary,
            "stats": stats,
        }
        text_report = "\n".join(lines)
        return text_report + "\n__STRUCTURED__\n" + json.dumps(structured, ensure_ascii=False)

    # ── Tool 7: prepare_upload ────────────────────────────────

    @registry.tool(
        description=(
            "One-step upload preparation: validation + audit + preview in a "
            "single unified review card. Recommended tool for the validation "
            "step — replaces separate resolve_and_validate + audit_data + "
            "preview_changes. Pass supplier_name/supplier_id, country_name/country_id, "
            "port_name/port_id, effective_from, effective_to. ID params take priority over names."
        ),
        parameters={
            "supplier_name": {"type": "STRING", "description": "供应商名称", "required": False},
            "supplier_id": {"type": "NUMBER", "description": "供应商ID（优先于 supplier_name）", "required": False},
            "country_name": {"type": "STRING", "description": "国家名称", "required": False},
            "country_id": {"type": "NUMBER", "description": "国家ID（优先于 country_name）", "required": False},
            "port_name": {"type": "STRING", "description": "目标港口名称", "required": False},
            "port_id": {"type": "NUMBER", "description": "港口ID（优先于 port_name）", "required": False},
            "effective_from": {"type": "STRING", "description": "生效开始日期 YYYY-MM-DD", "required": False},
            "effective_to": {"type": "STRING", "description": "生效结束日期 YYYY-MM-DD", "required": False},
        },
        group="data_upload",
    )
    def prepare_upload(supplier_name: str = "", supplier_id: int = 0, country_name: str = "", country_id: int = 0, port_name: str = "", port_id: int = 0, effective_from: str = "", effective_to: str = "") -> str:
        from sqlalchemy import text as sa_text
        from core.models import UploadBatch, StagingProduct, Product

        batch, err = _load_batch(ctx)
        if err:
            return err

        # ── Required field validation ──
        has_country = batch.country_id or country_id or country_name.strip()
        has_port = batch.port_id or port_id or port_name.strip()
        if has_country and not has_port:
            return "Error: 请提供目标港口名称（port_name 参数）。例如：横浜、Bangkok 等。"
        if not batch.effective_from and not effective_from.strip():
            return "Error: 请提供价格生效开始日期（effective_from 参数）。格式：YYYY-MM-DD。"
        if not batch.effective_to and not effective_to.strip():
            return "Error: 请提供价格生效结束日期（effective_to 参数）。格式：YYYY-MM-DD。"

        staging_rows = (
            ctx.db.query(StagingProduct)
            .filter(StagingProduct.batch_id == batch.id)
            .order_by(StagingProduct.row_number)
            .all()
        )
        if not staging_rows:
            return "Error: 暂存表中没有数据"

        # ═══════════════════════════════════════════════════════════
        # Phase A: Resolve supplier / country / port / dates
        #          (ID params take priority over name params)
        # ═══════════════════════════════════════════════════════════
        if supplier_id and not batch.supplier_id:
            supplier_id = int(supplier_id)
            try:
                with ctx.db.begin_nested():
                    row = ctx.db.execute(
                        sa_text("SELECT id, name FROM suppliers WHERE id = :id"),
                        {"id": supplier_id},
                    ).fetchone()
                if row:
                    batch.supplier_id = row[0]
                    batch.supplier_name = row[1]
            except Exception as e:
                logger.warning("Supplier ID lookup failed: %s", e)
        elif supplier_name.strip() and not batch.supplier_id:
            try:
                sp_sv = ctx.db.begin_nested()
                rows = ctx.db.execute(
                    sa_text("SELECT id, name FROM suppliers WHERE name ILIKE :p LIMIT 5"),
                    {"p": f"%{supplier_name.strip()}%"},
                ).fetchall()
                if rows:
                    batch.supplier_id = rows[0][0]
                    batch.supplier_name = rows[0][1]
                sp_sv.commit()
            except Exception as e:
                sp_sv.rollback()
                logger.warning("Supplier lookup failed: %s", e)

        if country_id and not batch.country_id:
            country_id = int(country_id)
            try:
                with ctx.db.begin_nested():
                    row = ctx.db.execute(
                        sa_text("SELECT id, name FROM countries WHERE id = :id"),
                        {"id": country_id},
                    ).fetchone()
                if row:
                    batch.country_id = row[0]
                    batch.country_name = row[1]
            except Exception as e:
                logger.warning("Country ID lookup failed: %s", e)
        elif country_name.strip() and not batch.country_id:
            try:
                ct_sv = ctx.db.begin_nested()
                rows = ctx.db.execute(
                    sa_text("SELECT id, name FROM countries WHERE name ILIKE :p OR code ILIKE :p LIMIT 5"),
                    {"p": f"%{country_name.strip()}%"},
                ).fetchall()
                if rows:
                    batch.country_id = rows[0][0]
                    batch.country_name = rows[0][1]
                ct_sv.commit()
            except Exception as e:
                ct_sv.rollback()
                logger.warning("Country lookup failed: %s", e)

        if port_id and not batch.port_id:
            port_id = int(port_id)
            try:
                with ctx.db.begin_nested():
                    row = ctx.db.execute(
                        sa_text("SELECT id, name FROM ports WHERE id = :id"),
                        {"id": port_id},
                    ).fetchone()
                if row:
                    batch.port_id = row[0]
                    batch.port_name = row[1]
            except Exception as e:
                logger.warning("Port ID lookup failed: %s", e)
        elif port_name.strip() and not batch.port_id:
            try:
                pt_sv = ctx.db.begin_nested()
                rows = ctx.db.execute(
                    sa_text("SELECT id, name FROM ports WHERE name ILIKE :p LIMIT 5"),
                    {"p": f"%{port_name.strip()}%"},
                ).fetchall()
                if rows:
                    batch.port_id = rows[0][0]
                    batch.port_name = rows[0][1]
                pt_sv.commit()
            except Exception as e:
                pt_sv.rollback()
                logger.warning("Port lookup failed: %s", e)

        if effective_from.strip() and not batch.effective_from:
            try:
                batch.effective_from = datetime.strptime(effective_from.strip(), "%Y-%m-%d").date()
            except ValueError:
                pass
        if effective_to.strip() and not batch.effective_to:
            try:
                batch.effective_to = datetime.strptime(effective_to.strip(), "%Y-%m-%d").date()
            except ValueError:
                pass

        # ── Load DB products for matching ──
        query = ctx.db.query(Product).filter(Product.status == True)
        if batch.supplier_id:
            query = query.filter(Product.supplier_id == batch.supplier_id)
        if batch.country_id:
            query = query.filter(Product.country_id == batch.country_id)
        if batch.port_id:
            query = query.filter(Product.port_id == batch.port_id)
        db_products = query.all()

        by_code: dict[str, Product] = {}
        by_name: dict[str, Product] = {}
        db_product_by_id: dict[int, Product] = {}
        for dbp in db_products:
            db_product_by_id[dbp.id] = dbp
            if dbp.code:
                by_code[dbp.code.upper()] = dbp
            if dbp.product_name_en:
                by_name[dbp.product_name_en.upper()] = dbp

        # ═══════════════════════════════════════════════════════════
        # Phase A-2: Code + LLM matching (same logic as resolve_and_validate)
        # ═══════════════════════════════════════════════════════════
        unmatched_rows = []
        match_stats = {"new": 0, "update": 0, "no_change": 0}

        for sp in staging_rows:
            code = (sp.product_code or "").upper()
            name = (sp.product_name or "").upper()
            new_price = float(sp.price) if sp.price is not None else None

            matched_db = None
            confidence = 0.0
            match_method = ""

            if code and code in by_code:
                matched_db = by_code[code]
                confidence = 1.0
                match_method = "code_exact"
            elif name and name in by_name:
                matched_db = by_name[name]
                confidence = 0.95
                match_method = "name_exact"
            else:
                best_score = 0.0
                for dbp in db_products:
                    if not dbp.product_name_en:
                        continue
                    sim = SequenceMatcher(None, name, dbp.product_name_en.upper()).ratio()
                    if sim > best_score and sim >= 0.6:
                        best_score = sim
                        matched_db = dbp
                        match_method = "name_fuzzy"
                confidence = best_score

            if matched_db and confidence >= 0.6:
                old_price = float(matched_db.price) if matched_db.price is not None else None
                price_change_pct = None
                action = "update"

                if new_price is not None and old_price is not None and old_price > 0:
                    price_change_pct = round((new_price - old_price) / old_price * 100, 1)
                    if abs(price_change_pct) < 0.01:
                        action = "no_change"
                elif new_price is None and old_price is not None:
                    action = "no_change"

                sp.match_result = {
                    "action": action,
                    "confidence": round(confidence, 3),
                    "matched_product_id": matched_db.id,
                    "match_method": match_method,
                    "old_price": old_price,
                    "price_change_pct": price_change_pct,
                    "db_product_name": matched_db.product_name_en,
                }
                sp.validation_status = "quarantined" if confidence < 0.7 else "valid"
                match_stats[action] += 1
            else:
                if confidence < 0.6:
                    matched_db = None
                unmatched_rows.append(sp)

        # LLM fuzzy matching for unmatched items
        if unmatched_rows and db_products:
            try:
                llm_results = _llm_fuzzy_match(unmatched_rows, db_products, batch, ctx)
                for sp in unmatched_rows:
                    row_key = str(sp.row_number)
                    if row_key in llm_results:
                        lr = llm_results[row_key]
                        matched_id = lr.get("matched_product_id")
                        llm_conf = lr.get("confidence", 0)

                        if matched_id and llm_conf >= 0.5:
                            matched_product = db_product_by_id.get(matched_id)
                            if matched_product:
                                old_price = float(matched_product.price) if matched_product.price is not None else None
                                new_price = float(sp.price) if sp.price is not None else None
                                price_change_pct = None
                                action = "update"

                                if new_price is not None and old_price is not None and old_price > 0:
                                    price_change_pct = round((new_price - old_price) / old_price * 100, 1)
                                    if abs(price_change_pct) < 0.01:
                                        action = "no_change"

                                sp.match_result = {
                                    "action": action,
                                    "confidence": round(llm_conf, 3),
                                    "matched_product_id": matched_product.id,
                                    "match_method": "llm_fuzzy",
                                    "old_price": old_price,
                                    "price_change_pct": price_change_pct,
                                    "db_product_name": matched_product.product_name_en,
                                }
                                sp.validation_status = "quarantined" if llm_conf < 0.7 else "valid"
                                match_stats[action] += 1
                                continue

                    if not sp.match_result:
                        sp.match_result = _make_new_match_result(sp, batch, ctx)
                        sp.validation_status = "valid"
                        match_stats["new"] += 1
            except Exception as e:
                logger.warning("LLM fuzzy match failed in prepare_upload: %s", e)
                for sp in unmatched_rows:
                    if not sp.match_result:
                        sp.match_result = _make_new_match_result(sp, batch, ctx)
                        sp.validation_status = "valid"
                        match_stats["new"] += 1
        else:
            for sp in unmatched_rows:
                sp.match_result = _make_new_match_result(sp, batch, ctx)
                sp.validation_status = "valid"
                match_stats["new"] += 1

        # Update resolved IDs
        for sp in staging_rows:
            sp.resolved_supplier_id = batch.supplier_id
            sp.resolved_country_id = batch.country_id

        # ═══════════════════════════════════════════════════════════
        # Phase B: Audit
        # ═══════════════════════════════════════════════════════════
        audit_findings_raw = _code_audit(batch, staging_rows)
        llm_audit = _llm_semantic_audit(batch, staging_rows, ctx)
        if llm_audit is not None:
            existing_keys = {(f["category"], tuple(f["rows"][:5])) for f in audit_findings_raw}
            for lf in llm_audit:
                if not isinstance(lf, dict) or not lf.get("message"):
                    continue
                lf.setdefault("severity", "info")
                lf.setdefault("category", "other")
                if not isinstance(lf.get("rows"), list):
                    lf["rows"] = []
                lf["rows"] = [r for r in lf["rows"] if isinstance(r, (int, float))]
                lf.setdefault("suggestion", "")
                if lf["severity"] not in ("error", "warning", "info"):
                    lf["severity"] = "info"
                key = (lf["category"], tuple(lf["rows"][:5]))
                if key not in existing_keys:
                    audit_findings_raw.append(lf)

        # ═══════════════════════════════════════════════════════════
        # Phase C: Build unified review data
        # ═══════════════════════════════════════════════════════════
        new_items = []
        updates = []
        no_change_count = 0

        # Collect audit error rows for cross-referencing
        audit_error_rows: set[int] = set()
        for af in audit_findings_raw:
            if af.get("severity") == "error":
                for r in af.get("rows", []):
                    audit_error_rows.add(int(r))

        for sp in staging_rows:
            mr = sp.match_result or {}
            action = mr.get("action", "new")
            confidence = mr.get("confidence", 0)
            matched_id = mr.get("matched_product_id")
            db_name = mr.get("db_product_name", "")

            # Build field diffs for matched items
            diffs = []
            if matched_id and action in ("update", "no_change"):
                db_prod = db_product_by_id.get(matched_id)
                if db_prod:
                    # Price diff (round to 2dp to avoid float imprecision)
                    if sp.price is not None:
                        old_p = round(float(db_prod.price), 2) if db_prod.price is not None else None
                        new_p = round(float(sp.price), 2)
                        if old_p != new_p:
                            diffs.append({"field": "price", "old": old_p, "new": new_p})
                    # Unit diff
                    if sp.unit and sp.unit != (db_prod.unit or ""):
                        diffs.append({"field": "unit", "old": db_prod.unit, "new": sp.unit})
                    # Pack size diff
                    if sp.pack_size and sp.pack_size != (db_prod.pack_size or ""):
                        diffs.append({"field": "pack_size", "old": db_prod.pack_size, "new": sp.pack_size})
                    # Brand diff
                    if sp.brand and sp.brand != (db_prod.brand or ""):
                        diffs.append({"field": "brand", "old": db_prod.brand, "new": sp.brand})
                    # Currency diff
                    if sp.currency and sp.currency != (db_prod.currency or ""):
                        diffs.append({"field": "currency", "old": db_prod.currency, "new": sp.currency})
                    # Country of origin diff
                    if sp.country_of_origin and sp.country_of_origin != (db_prod.country_of_origin if hasattr(db_prod, 'country_of_origin') else ""):
                        diffs.append({"field": "country_of_origin", "old": getattr(db_prod, 'country_of_origin', None), "new": sp.country_of_origin})
                    # Effective date diffs
                    if batch.effective_from:
                        old_from = db_prod.effective_from
                        old_from_date = old_from.date() if hasattr(old_from, 'date') and old_from else old_from
                        if old_from_date != batch.effective_from:
                            diffs.append({"field": "effective_from", "old": str(old_from) if old_from else None, "new": str(batch.effective_from)})
                    if batch.effective_to:
                        old_to = db_prod.effective_to
                        old_to_date = old_to.date() if hasattr(old_to, 'date') and old_to else old_to
                        if old_to_date != batch.effective_to:
                            diffs.append({"field": "effective_to", "old": str(old_to) if old_to else None, "new": str(batch.effective_to)})

            # Build warning string for items needing attention
            warnings = []
            if action in ("update",) and confidence < 0.7:
                warnings.append(f"低置信度({confidence:.0%})")
            if sp.row_number in audit_error_rows:
                for af in audit_findings_raw:
                    if af.get("severity") == "error" and sp.row_number in af.get("rows", []):
                        warnings.append(af.get("message", "审计错误"))
                        break
            warning = " | ".join(warnings) if warnings else None

            # Classify — no more "problematic" group
            effective_action = action
            if action == "no_change" and diffs:
                effective_action = "update"

            if effective_action == "new":
                new_items.append({
                    "row": sp.row_number,
                    "name": sp.product_name,
                    "code": sp.product_code,
                    "price": float(sp.price) if sp.price is not None else None,
                    "unit": sp.unit,
                    "pack_size": sp.pack_size,
                    "brand": sp.brand,
                })
            elif effective_action == "update" and diffs:
                updates.append({
                    "row": sp.row_number,
                    "name": sp.product_name,
                    "code": sp.product_code,
                    "confidence": confidence,
                    "match_method": mr.get("match_method", ""),
                    "db_name": db_name,
                    "diffs": diffs,
                    "warning": warning,
                })
            else:
                no_change_count += 1

        # Set batch status
        batch.status = "previewing"
        ctx.db.commit()

        # ── Build text report ──
        total = len(staging_rows)
        lines = [
            f"## 上传审查（批次 #{batch.id}）",
            f"- 总数: {total} | 新增: {len(new_items)} | 更新: {len(updates)} | 无变化: {no_change_count}",
        ]
        if batch.supplier_name:
            lines.append(f"- 供应商: {batch.supplier_name}")
        elif supplier_name:
            lines.append(f"- 未找到供应商 '{supplier_name}'")
        if batch.country_name:
            lines.append(f"- 国家: {batch.country_name}")
        if batch.port_name:
            lines.append(f"- 港口: {batch.port_name}")
        if batch.effective_from or batch.effective_to:
            lines.append(f"- 有效期: {batch.effective_from or '?'} ~ {batch.effective_to or '?'}")

        lines.append("\n请审查后执行导入。")

        # ── Build audit_findings for card (non-row-level) ──
        card_audit_findings = []
        for af in audit_findings_raw:
            if not af.get("rows"):
                card_audit_findings.append({
                    "severity": af.get("severity", "info"),
                    "message": af.get("message", ""),
                    "suggestion": af.get("suggestion", ""),
                })

        # ── Build structured data ──
        structured = {
            "card_type": "upload_review",
            "batch_id": batch.id,
            "supplier": {"name": batch.supplier_name, "id": batch.supplier_id},
            "country": {"name": batch.country_name, "id": batch.country_id},
            "port": {"name": batch.port_name, "id": batch.port_id},
            "effective_from": str(batch.effective_from) if batch.effective_from else None,
            "effective_to": str(batch.effective_to) if batch.effective_to else None,
            "stats": {
                "new": len(new_items),
                "update": len(updates),
                "no_change": no_change_count,
                "total": total,
            },
            "new_items": new_items[:50],
            "updates": updates[:50],
            "audit_findings": card_audit_findings,
            "missing_supplier": bool(supplier_name and not batch.supplier_id),
            "missing_country": bool(country_name and not batch.country_id),
        }
        text_report = "\n".join(lines)
        return text_report + "\n__STRUCTURED__\n" + json.dumps(structured, ensure_ascii=False, default=str)


# ── Audit Helpers ──────────────────────────────────────────────


def _code_audit(batch, staging_rows) -> list[dict]:
    """Deterministic structural checks — no LLM needed."""
    findings = []
    mapping = batch.column_mapping or {}

    # 1. Key column missing
    if not mapping.get("product_name"):
        findings.append({
            "severity": "error", "category": "column_missing",
            "rows": [],
            "message": "关键列缺失：未映射到 product_name（产品名称）",
            "suggestion": "检查 Excel 表头是否包含产品名称列，或手动指定列映射",
        })
    if not mapping.get("price"):
        findings.append({
            "severity": "warning", "category": "column_missing",
            "rows": [],
            "message": "价格列未映射：所有产品将没有价格信息",
            "suggestion": "检查是否有 Price/単価/Unit Price 等列",
        })

    # 2. Required fields empty
    no_price_rows = [sp.row_number for sp in staging_rows if sp.price is None]
    no_name_rows = [sp.row_number for sp in staging_rows if not sp.product_name]
    if no_price_rows and mapping.get("price"):
        findings.append({
            "severity": "warning", "category": "empty_field",
            "rows": no_price_rows[:20],
            "message": f"{len(no_price_rows)} 行缺少价格",
            "suggestion": "检查这些行的价格单元格是否为空或格式不正确",
        })
    if no_name_rows:
        findings.append({
            "severity": "error", "category": "empty_field",
            "rows": no_name_rows[:20],
            "message": f"{len(no_name_rows)} 行缺少产品名称",
            "suggestion": "这些行可能是空行或表尾注释，建议排除",
        })

    # 3. Duplicate detection
    code_counts: dict[str, list[int]] = {}
    name_counts: dict[str, list[int]] = {}
    for sp in staging_rows:
        if sp.product_code:
            code_counts.setdefault(sp.product_code, []).append(sp.row_number)
        if sp.product_name:
            name_counts.setdefault(sp.product_name.upper(), []).append(sp.row_number)
    dup_codes = {k: v for k, v in code_counts.items() if len(v) > 1}
    dup_names = {k: v for k, v in name_counts.items() if len(v) > 1}
    if dup_codes:
        all_rows = [r for rows in dup_codes.values() for r in rows]
        findings.append({
            "severity": "warning", "category": "duplicate",
            "rows": all_rows[:20],
            "message": f"{len(dup_codes)} 个产品代码重复: {', '.join(list(dup_codes.keys())[:5])}",
            "suggestion": "同一代码的多行可能是重复录入，建议保留最新价格",
        })
    if dup_names:
        # Only flag name dups not already covered by code dups
        dup_code_names = set()
        for code, rows in dup_codes.items():
            for sp in staging_rows:
                if sp.product_code == code and sp.product_name:
                    dup_code_names.add(sp.product_name.upper())
        pure_name_dups = {k: v for k, v in dup_names.items() if k not in dup_code_names}
        if pure_name_dups:
            all_rows = [r for rows in pure_name_dups.values() for r in rows][:20]
            findings.append({
                "severity": "info", "category": "duplicate",
                "rows": all_rows,
                "message": f"{len(pure_name_dups)} 个产品名称重复（无代码区分）",
                "suggestion": "检查是否为同一产品的不同规格",
            })

    # 4. Format errors (price column has raw value but parsed as None)
    if mapping.get("price"):
        price_col = mapping["price"]
        format_err_rows = []
        for sp in staging_rows:
            raw = (sp.raw_data or {}).get(price_col, "")
            if raw and str(raw).strip() and sp.price is None:
                format_err_rows.append(sp.row_number)
        if format_err_rows:
            findings.append({
                "severity": "error", "category": "format_error",
                "rows": format_err_rows[:20],
                "message": f"{len(format_err_rows)} 行价格格式无法解析（原始值非数字）",
                "suggestion": "这些行的价格单元格可能包含文本或特殊符号，需要手动修正",
            })

    # 5. Short product names (< 3 chars → likely garbage)
    short_name_rows = [
        sp.row_number for sp in staging_rows
        if sp.product_name and len(sp.product_name.strip()) < 3
    ]
    if short_name_rows:
        findings.append({
            "severity": "warning", "category": "name_quality",
            "rows": short_name_rows[:20],
            "message": f"{len(short_name_rows)} 行产品名称过短（<3字符）",
            "suggestion": "检查是否为缩写、错误数据或空行",
        })

    return findings


def _llm_semantic_audit(batch, staging_rows, ctx) -> list[dict] | None:
    """LLM-based semantic audit using world knowledge. Returns findings or None on failure."""
    try:
        from core.config import settings
        from google import genai
        from google.genai import types

        # Build comparison data (cap at 100 rows)
        if len(staging_rows) > 100:
            sample = staging_rows[:50] + staging_rows[-50:]
        else:
            sample = staging_rows

        comparison = []
        for sp in sample:
            mr = sp.match_result or {}
            row_data = {
                "row": sp.row_number,
                "product_name": sp.product_name,
                "product_code": sp.product_code or "",
                "price": float(sp.price) if sp.price is not None else None,
                "unit": sp.unit or "",
                "pack_size": sp.pack_size or "",
                "currency": sp.currency or "",
            }
            if mr.get("db_product_name"):
                row_data["db_product_name"] = mr["db_product_name"]
            if mr.get("old_price") is not None:
                row_data["db_price"] = mr["old_price"]
            comparison.append(row_data)

        mapping = batch.column_mapping or {}

        prompt = f"""你是产品数据质量审计专家。代码已经完成了基础检查，现在请你做深层语义审计。

## 审计重点
1. 列映射验证: column_mapping 是否合理（如 price 列的实际内容看起来像价格吗？）
2. 单位一致性: 上传单位与 DB 单位是否匹配（KG vs LB, EA vs PC 等）
3. 价格合理性: 基于产品类型和行业常识，价格是否在合理范围
4. 名称质量: 产品名称是否看起来像正常的产品名（非备注、非数字、非乱码）
5. 跨行一致性: 同一批次产品的单位/货币是否一致
6. 其他: 任何你觉得可疑的数据模式

## 列映射
{json.dumps(mapping, ensure_ascii=False)}

## 数据（上传 vs DB 对照, 共 {len(comparison)} 行 / 总 {len(staging_rows)} 行）
{json.dumps(comparison, ensure_ascii=False)}

返回 JSON（严格格式，不要 markdown）:
{{"findings": [{{"severity": "error"|"warning"|"info", "category": "mapping_error"|"unit_mismatch"|"price_anomaly"|"name_quality"|"consistency"|"other", "rows": [行号], "message": "中文描述", "suggestion": "中文建议"}}]}}

注意：
- 只返回有实质问题的 findings，不要为了凑数而生成
- 如果没有问题就返回空数组: {{"findings": []}}
- rows 数组最多 20 个行号
"""

        client = genai.Client(api_key=settings.GOOGLE_API_KEY)
        response = client.models.generate_content(
            model="gemini-3-flash-preview",
            contents=prompt,
            config=types.GenerateContentConfig(
                response_mime_type="application/json",
                thinking_config=types.ThinkingConfig(thinking_budget=2048),
                temperature=0.1,
            ),
        )

        text = response.text.strip()
        data = json.loads(text)
        return data.get("findings", [])

    except Exception as e:
        logger.warning("LLM semantic audit failed (graceful skip): %s", e)
        return None


# ── LLM Fuzzy Match Helper ────────────────────────────────────


def _llm_fuzzy_match(
    unmatched_rows: list,
    db_products: list,
    batch,
    ctx,
) -> dict:
    """Single LLM call to fuzzy-match unmatched staging rows against DB products.

    Returns: {row_number_str: {matched_product_id, confidence}}
    """
    # Build payload for LLM
    staging_items = []
    for sp in unmatched_rows[:50]:  # Cap at 50 items per LLM call
        staging_items.append({
            "row": sp.row_number,
            "name": sp.product_name,
            "code": sp.product_code or "",
            "brand": sp.brand or "",
        })

    # Sample DB products (cap at 200 for token limit)
    db_candidates = []
    for dbp in db_products[:200]:
        db_candidates.append({
            "id": dbp.id,
            "name": dbp.product_name_en,
            "code": dbp.code or "",
            "brand": dbp.brand or "",
        })

    if not db_candidates:
        return {}

    prompt = f"""你是产品匹配专家。将上传的产品与数据库产品进行模糊匹配。

## 上传产品 (待匹配)
{json.dumps(staging_items, ensure_ascii=False)}

## 数据库产品 (候选)
{json.dumps(db_candidates, ensure_ascii=False)}

## 规则
- 同一个产品可能名称略有不同（缩写、大小写、空格差异等）
- 返回每个上传产品最可能的匹配及置信度 (0-1)
- 如果没有合适的匹配，confidence 设为 0
- 只返回 JSON

## 返回格式
{{"results": [{{"row": 1, "matched_id": 123, "confidence": 0.85}}, ...]}}
"""

    try:
        from core.config import settings
        import google.generativeai as genai

        genai.configure(api_key=settings.GOOGLE_API_KEY)
        model = genai.GenerativeModel("gemini-3-flash-preview")
        response = model.generate_content(prompt)
        text = response.text.strip()

        if "```" in text:
            m = re.search(r"```(?:json)?\s*(.*?)```", text, re.DOTALL)
            if m:
                text = m.group(1).strip()

        data = json.loads(text)
        results = data.get("results", [])

        return {
            str(r["row"]): {
                "matched_product_id": r.get("matched_id"),
                "confidence": r.get("confidence", 0),
            }
            for r in results
            if r.get("row") is not None
        }
    except Exception as e:
        logger.warning("LLM fuzzy match failed: %s", e)
        return {}
