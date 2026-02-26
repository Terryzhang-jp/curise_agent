"""
Product upload tools — parse Excel price lists, resolve references,
check existing products, and execute inserts/updates.

Tools:
  - parse_price_list: Extract products from uploaded Excel
  - resolve_references: Fuzzy-match supplier/country/port names to DB IDs
  - check_existing_products: Diff against products table
  - execute_product_upload: Write new/updated products to DB

All tools use the closure pattern: create_product_upload_tools(registry, ctx).
"""

from __future__ import annotations

import json
import logging
import os
from datetime import datetime
from decimal import Decimal, InvalidOperation
from difflib import SequenceMatcher
from io import BytesIO

logger = logging.getLogger(__name__)

# ── State persistence helpers ────────────────────────────────
# Upload state is saved per-session as a JSON file so that it survives
# across multiple chat messages (each message creates a fresh ToolContext).

_STATE_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "uploads", "upload_state")


def _save_upload_state(session_id: str | None, data: dict):
    """Persist upload state (parsed_products, resolved_refs, upload_plan) to disk."""
    if not session_id:
        return
    os.makedirs(_STATE_DIR, exist_ok=True)
    path = os.path.join(_STATE_DIR, f"{session_id}.json")
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, default=str)


def load_upload_state(session_id: str | None) -> dict | None:
    """Load persisted upload state for a session. Returns None if no state."""
    if not session_id:
        return None
    path = os.path.join(_STATE_DIR, f"{session_id}.json")
    if not os.path.exists(path):
        return None
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return None


def _cleanup_upload_state(session_id: str | None):
    """Remove persisted state after successful upload."""
    if not session_id:
        return
    path = os.path.join(_STATE_DIR, f"{session_id}.json")
    try:
        os.remove(path)
    except OSError:
        pass


def create_product_upload_tools(registry, ctx):
    """Register product upload tools onto the registry."""

    session_id = ctx.pipeline_session_id

    def _persist():
        """Save relevant ctx.session_data keys to disk."""
        state = {}
        for key in ("parsed_products", "column_mapping", "resolved_refs", "upload_plan", "file_url"):
            if key in ctx.session_data:
                state[key] = ctx.session_data[key]
        if state:
            _save_upload_state(session_id, state)

    def _restore():
        """Load persisted state into ctx.session_data if empty."""
        if ctx.session_data.get("parsed_products") or ctx.session_data.get("upload_plan"):
            return  # Already populated (same turn)
        saved = load_upload_state(session_id)
        if saved:
            ctx.session_data.update(saved)

    # ── Tool 1: parse_price_list ──────────────────────────────

    @registry.tool(
        description="解析上传的 Excel 报价单/价格表，提取产品列表。无需参数，直接从上传文件读取。",
        parameters={},
        group="product_upload",
    )
    def parse_price_list() -> str:
        if not ctx.file_bytes:
            return "Error: 没有检测到上传的文件。请先上传一份 Excel 文件。"

        try:
            from services.excel_parser import parse_excel_file
            parsed = parse_excel_file(ctx.file_bytes)
        except Exception as e:
            return f"Error: 解析 Excel 文件失败 — {str(e)}"

        sheets = parsed.get("sheets", [])
        if not sheets:
            return "Error: Excel 文件中没有有效的工作表"

        # Use the first sheet with data
        sheet = sheets[0]
        headers = sheet.get("headers", [])
        sample_rows = sheet.get("sample_rows", [])
        total_rows = sheet.get("total_rows", 0)
        data_start = sheet.get("data_start_row", 2)
        header_row = sheet.get("header_row", 1)

        if not headers:
            return "Error: 无法识别表头行"

        # Use Gemini to analyze column mapping
        header_labels = [h["label"] for h in headers]
        header_cols = [h["column"] for h in headers]

        sample_text = ""
        for i, row in enumerate(sample_rows[:3]):
            sample_text += f"  行{i+1}: {row}\n"

        prompt = f"""分析这份 Excel 报价单的列结构。

表头（第{header_row}行）: {json.dumps(dict(zip(header_cols, header_labels)), ensure_ascii=False)}

样例数据:
{sample_text}

请以 JSON 格式返回列映射，识别哪些列对应以下字段（没有的填 null）:
- product_name: 产品名称（英文或日文）
- product_code: 产品代码/SKU
- price: 单价/价格
- unit: 单位 (KG, EA, CS 等)
- pack_size: 包装规格
- brand: 品牌
- currency: 货币
- country_of_origin: 原产地

只返回 JSON，格式如: {{"product_name": "B", "product_code": "A", "price": "D", ...}}
"""
        try:
            from config import settings
            import google.generativeai as genai

            genai.configure(api_key=settings.GOOGLE_API_KEY)
            model = genai.GenerativeModel("gemini-2.0-flash")
            response = model.generate_content(prompt)
            mapping_text = response.text.strip()

            # Extract JSON from possible markdown code block
            if "```" in mapping_text:
                import re
                m = re.search(r"```(?:json)?\s*(.*?)```", mapping_text, re.DOTALL)
                if m:
                    mapping_text = m.group(1).strip()

            column_mapping = json.loads(mapping_text)
        except Exception as e:
            logger.warning("Gemini column mapping failed: %s, falling back to heuristic", e)
            column_mapping = _heuristic_column_mapping(header_labels, header_cols)

        # Extract products using column mapping
        try:
            from openpyxl import load_workbook
            wb = load_workbook(BytesIO(ctx.file_bytes), data_only=True)  # non-read_only to avoid EmptyCell issues
            ws = wb.worksheets[0]

            products = []
            for row_idx, row in enumerate(ws.iter_rows(min_row=data_start, values_only=False), start=1):
                cells = {}
                for cell in row:
                    if cell.column is not None:
                        cells[_col_letter(cell.column)] = cell.value

                name_col = column_mapping.get("product_name")
                name = str(cells.get(name_col, "") or "").strip() if name_col else ""
                if not name:
                    continue  # Skip empty rows

                code_col = column_mapping.get("product_code")
                price_col = column_mapping.get("price")
                unit_col = column_mapping.get("unit")
                pack_col = column_mapping.get("pack_size")
                brand_col = column_mapping.get("brand")
                currency_col = column_mapping.get("currency")
                origin_col = column_mapping.get("country_of_origin")

                product = {
                    "row_num": row_idx,
                    "product_name": name,
                    "product_code": str(cells.get(code_col, "") or "").strip() if code_col else "",
                    "price": _parse_price(cells.get(price_col)) if price_col else None,
                    "unit": str(cells.get(unit_col, "") or "").strip() if unit_col else "",
                    "pack_size": str(cells.get(pack_col, "") or "").strip() if pack_col else "",
                    "brand": str(cells.get(brand_col, "") or "").strip() if brand_col else "",
                    "currency": str(cells.get(currency_col, "") or "").strip() if currency_col else "",
                    "country_of_origin": str(cells.get(origin_col, "") or "").strip() if origin_col else "",
                }
                products.append(product)

            wb.close()
        except Exception as e:
            return f"Error: 提取产品数据失败 — {str(e)}"

        if not products:
            return "Error: 未能从文件中提取到任何产品数据"

        # Store in session_data for subsequent tools
        ctx.session_data["parsed_products"] = products
        ctx.session_data["column_mapping"] = column_mapping

        # Build summary
        with_price = sum(1 for p in products if p.get("price") is not None)
        with_code = sum(1 for p in products if p.get("product_code"))

        lines = [
            f"文件解析完成:",
            f"- 工作表: {sheet['name']}",
            f"- 产品数量: {len(products)} 个",
            f"- 有产品代码: {with_code} 个",
            f"- 有价格: {with_price} 个",
            f"- 列映射: {json.dumps(column_mapping, ensure_ascii=False)}",
            "",
            "前 5 个产品预览:",
        ]
        for p in products[:5]:
            price_str = f"${p['price']}" if p.get("price") is not None else "无价格"
            lines.append(f"  - {p['product_name']} [{p.get('product_code', '')}] {price_str}")

        _persist()
        return "\n".join(lines)

    # ── Tool 2: resolve_references ────────────────────────────

    @registry.tool(
        description="根据供应商名称和国家名称查找数据库中的对应 ID。Agent 从文件或用户消息中提取名称后调用此工具。",
        parameters={
            "supplier_name": {
                "type": "STRING",
                "description": "供应商名称（可选，模糊匹配）",
                "required": False,
            },
            "country_name": {
                "type": "STRING",
                "description": "国家名称（可选，模糊匹配）",
                "required": False,
            },
        },
        group="product_upload",
    )
    def resolve_references(supplier_name: str = "", country_name: str = "") -> str:
        from sqlalchemy import text

        resolved = ctx.session_data.get("resolved_refs", {})
        lines = []

        # Resolve supplier
        if supplier_name.strip():
            try:
                rows = ctx.db.execute(
                    text("SELECT id, name FROM suppliers WHERE name ILIKE :pattern LIMIT 5"),
                    {"pattern": f"%{supplier_name.strip()}%"},
                ).fetchall()
                if rows:
                    best = rows[0]
                    resolved["supplier_id"] = best[0]
                    resolved["supplier_name"] = best[1]
                    lines.append(f"供应商: {best[1]} (id={best[0]})")
                    if len(rows) > 1:
                        others = ", ".join(f"{r[1]}(id={r[0]})" for r in rows[1:])
                        lines.append(f"  其他候选: {others}")
                else:
                    lines.append(f"未找到供应商 '{supplier_name}'。如需创建新供应商，请告知用户。")
            except Exception as e:
                ctx.db.rollback()
                lines.append(f"查询供应商失败: {str(e)}")

        # Resolve country
        if country_name.strip():
            try:
                rows = ctx.db.execute(
                    text("SELECT id, name, code FROM countries WHERE name ILIKE :pattern OR code ILIKE :pattern LIMIT 5"),
                    {"pattern": f"%{country_name.strip()}%"},
                ).fetchall()
                if rows:
                    best = rows[0]
                    resolved["country_id"] = best[0]
                    resolved["country_name"] = best[1]
                    lines.append(f"国家: {best[1]} (id={best[0]}, code={best[2]})")

                    # Also find ports for this country
                    ports = ctx.db.execute(
                        text("SELECT id, name FROM ports WHERE country_id = :cid LIMIT 10"),
                        {"cid": best[0]},
                    ).fetchall()
                    if ports:
                        port_list = ", ".join(f"{p[1]}(id={p[0]})" for p in ports)
                        lines.append(f"  该国家的港口: {port_list}")
                else:
                    lines.append(f"未找到国家 '{country_name}'")
            except Exception as e:
                ctx.db.rollback()
                lines.append(f"查询国家失败: {str(e)}")

        if not supplier_name.strip() and not country_name.strip():
            lines.append("请提供 supplier_name 或 country_name 参数")

        ctx.session_data["resolved_refs"] = resolved
        _persist()
        return "\n".join(lines) if lines else "无输入参数"

    # ── Tool 3: check_existing_products ───────────────────────

    @registry.tool(
        description="将解析出的产品与数据库现有产品比对，找出新增、价格更新、无变化和价格异常的产品。无需参数。",
        parameters={},
        group="product_upload",
    )
    def check_existing_products() -> str:
        _restore()
        products = ctx.session_data.get("parsed_products")
        if not products:
            return "Error: 请先调用 parse_price_list 解析文件"

        refs = ctx.session_data.get("resolved_refs", {})
        supplier_id = refs.get("supplier_id")
        country_id = refs.get("country_id")

        from models import ProductReadOnly

        # Fetch candidate DB products
        query = ctx.db.query(ProductReadOnly).filter(ProductReadOnly.status == True)
        if supplier_id:
            query = query.filter(ProductReadOnly.supplier_id == supplier_id)
        if country_id:
            query = query.filter(ProductReadOnly.country_id == country_id)
        db_products = query.all()

        # Index by code and name for fast lookup
        by_code = {}
        by_name = {}
        for dbp in db_products:
            if dbp.code:
                by_code[dbp.code.upper()] = dbp
            if dbp.product_name_en:
                by_name[dbp.product_name_en.upper()] = dbp

        new_products = []
        price_updates = []
        no_change = []
        anomalies = []

        for p in products:
            code = (p.get("product_code") or "").upper()
            name = (p.get("product_name") or "").upper()
            new_price = p.get("price")

            # Try to find existing product
            matched_db = None
            match_method = ""

            # 1. Exact code match
            if code and code in by_code:
                matched_db = by_code[code]
                match_method = "代码匹配"
            # 2. Exact name match
            elif name and name in by_name:
                matched_db = by_name[name]
                match_method = "名称精确匹配"
            # 3. Fuzzy name match
            else:
                best_score = 0.0
                for dbp in db_products:
                    if not dbp.product_name_en:
                        continue
                    sim = SequenceMatcher(None, name, dbp.product_name_en.upper()).ratio()
                    if sim > best_score and sim >= 0.85:
                        best_score = sim
                        matched_db = dbp
                        match_method = f"名称模糊匹配({sim:.0%})"

            entry = {
                "row_num": p["row_num"],
                "product_name": p["product_name"],
                "product_code": p.get("product_code", ""),
                "new_price": new_price,
                "unit": p.get("unit", ""),
                "pack_size": p.get("pack_size", ""),
                "brand": p.get("brand", ""),
                "currency": p.get("currency", ""),
                "country_of_origin": p.get("country_of_origin", ""),
            }

            if matched_db:
                entry["db_product_id"] = matched_db.id
                entry["db_product_name"] = matched_db.product_name_en
                entry["db_code"] = matched_db.code
                entry["match_method"] = match_method
                old_price = float(matched_db.price) if matched_db.price else None

                if new_price is not None and old_price is not None and old_price > 0:
                    change_pct = (new_price - old_price) / old_price * 100
                    entry["old_price"] = old_price
                    entry["price_change_pct"] = round(change_pct, 1)

                    if abs(change_pct) < 0.01:
                        entry["action"] = "no_change"
                        no_change.append(entry)
                    elif abs(change_pct) > 30:
                        entry["action"] = "anomaly"
                        anomalies.append(entry)
                    else:
                        entry["action"] = "update"
                        price_updates.append(entry)
                elif new_price is not None and old_price is None:
                    entry["action"] = "update"
                    entry["old_price"] = None
                    entry["price_change_pct"] = None
                    price_updates.append(entry)
                else:
                    entry["action"] = "no_change"
                    no_change.append(entry)
            else:
                entry["action"] = "new"
                new_products.append(entry)

        # Store upload plan
        upload_plan = {
            "new": new_products,
            "update": price_updates,
            "no_change": no_change,
            "anomaly": anomalies,
            "supplier_id": supplier_id,
            "country_id": country_id,
        }
        ctx.session_data["upload_plan"] = upload_plan

        # Build report
        lines = [
            "## 比对结果",
            f"- 新增产品: {len(new_products)} 个",
            f"- 价格更新: {len(price_updates)} 个",
            f"- 无变化: {len(no_change)} 个",
            f"- 价格异常 (涨跌幅 >30%): {len(anomalies)} 个",
            "",
        ]

        if new_products:
            lines.append("### 新增产品 (前10)")
            for p in new_products[:10]:
                price_str = f"${p['new_price']}" if p.get("new_price") is not None else "无价格"
                lines.append(f"  行{p['row_num']}: {p['product_name']} [{p.get('product_code', '')}] {price_str}")

        if price_updates:
            lines.append("")
            lines.append("### 价格更新 (前10)")
            for p in price_updates[:10]:
                old = p.get("old_price", "?")
                new = p.get("new_price", "?")
                pct = p.get("price_change_pct", "?")
                lines.append(f"  行{p['row_num']}: {p['product_name']} ${old} → ${new} ({pct:+.1f}%)" if isinstance(pct, (int, float)) else f"  行{p['row_num']}: {p['product_name']} ${old} → ${new}")

        if anomalies:
            lines.append("")
            lines.append("### ⚠️ 价格异常明细")
            for p in anomalies:
                old = p.get("old_price", "?")
                new = p.get("new_price", "?")
                pct = p.get("price_change_pct", "?")
                lines.append(f"  行{p['row_num']}: {p['product_name']} ${old} → ${new} ({pct:+.1f}%)" if isinstance(pct, (int, float)) else f"  行{p['row_num']}: {p['product_name']} ${old} → ${new}")

        lines.append("")
        lines.append("请确认后我将执行上传。如需排除某些行，请告知行号。")

        _persist()
        return "\n".join(lines)

    # ── Tool 4: execute_product_upload ────────────────────────

    @registry.tool(
        description="确认后执行产品数据导入（新增和更新）。可以排除指定行号的产品。",
        parameters={
            "exclude_rows": {
                "type": "STRING",
                "description": "要排除的行号，逗号分隔（如 '3,7,12'）。不传则全部执行。",
                "required": False,
            },
        },
        group="product_upload",
    )
    def execute_product_upload(exclude_rows: str = "") -> str:
        _restore()
        plan = ctx.session_data.get("upload_plan")
        if not plan:
            return "Error: 请先调用 check_existing_products 生成上传计划"

        # Parse excluded rows
        excluded = set()
        if exclude_rows.strip():
            for part in exclude_rows.split(","):
                part = part.strip()
                if part.isdigit():
                    excluded.add(int(part))

        supplier_id = plan.get("supplier_id")
        country_id = plan.get("country_id")

        from models import ProductReadOnly

        inserted = 0
        updated = 0
        failed = []

        # Process new products (INSERT)
        for p in plan.get("new", []):
            if p["row_num"] in excluded:
                continue
            try:
                new_product = ProductReadOnly(
                    product_name_en=p["product_name"],
                    code=p.get("product_code") or None,
                    price=Decimal(str(p["new_price"])) if p.get("new_price") is not None else None,
                    unit=p.get("unit") or None,
                    pack_size=p.get("pack_size") or None,
                    brand=p.get("brand") or None,
                    currency=p.get("currency") or None,
                    country_of_origin=p.get("country_of_origin") or None,
                    supplier_id=supplier_id,
                    country_id=country_id,
                    status=True,
                    effective_from=datetime.utcnow(),
                )
                ctx.db.add(new_product)
                ctx.db.flush()
                inserted += 1
            except Exception as e:
                ctx.db.rollback()
                failed.append(f"行{p['row_num']} {p['product_name']}: {str(e)}")

        # Process price updates (UPDATE)
        for p in plan.get("update", []):
            if p["row_num"] in excluded:
                continue
            try:
                db_product = ctx.db.query(ProductReadOnly).filter(
                    ProductReadOnly.id == p["db_product_id"]
                ).first()
                if db_product:
                    if p.get("new_price") is not None:
                        db_product.price = Decimal(str(p["new_price"]))
                    if p.get("unit"):
                        db_product.unit = p["unit"]
                    if p.get("pack_size"):
                        db_product.pack_size = p["pack_size"]
                    if p.get("currency"):
                        db_product.currency = p["currency"]
                    if p.get("brand"):
                        db_product.brand = p["brand"]
                    db_product.effective_from = datetime.utcnow()
                    ctx.db.flush()
                    updated += 1
                else:
                    failed.append(f"行{p['row_num']} {p['product_name']}: 数据库产品 id={p['db_product_id']} 不存在")
            except Exception as e:
                ctx.db.rollback()
                failed.append(f"行{p['row_num']} {p['product_name']}: {str(e)}")

        # Process anomalies that are NOT excluded (treat as updates)
        for p in plan.get("anomaly", []):
            if p["row_num"] in excluded:
                continue
            try:
                db_product = ctx.db.query(ProductReadOnly).filter(
                    ProductReadOnly.id == p["db_product_id"]
                ).first()
                if db_product:
                    if p.get("new_price") is not None:
                        db_product.price = Decimal(str(p["new_price"]))
                    if p.get("unit"):
                        db_product.unit = p["unit"]
                    if p.get("pack_size"):
                        db_product.pack_size = p["pack_size"]
                    if p.get("currency"):
                        db_product.currency = p["currency"]
                    db_product.effective_from = datetime.utcnow()
                    ctx.db.flush()
                    updated += 1
            except Exception as e:
                ctx.db.rollback()
                failed.append(f"行{p['row_num']} {p['product_name']}: {str(e)}")

        # Commit all changes
        try:
            ctx.db.commit()
        except Exception as e:
            ctx.db.rollback()
            return f"Error: 提交数据库失败 — {str(e)}"

        # Build result
        total = inserted + updated
        lines = [
            f"产品上传完成:",
            f"- 新增: {inserted} 个",
            f"- 更新: {updated} 个",
            f"- 总成功: {total} 个",
        ]
        if excluded:
            lines.append(f"- 已跳过行: {', '.join(str(r) for r in sorted(excluded))}")
        if failed:
            lines.append(f"- 失败: {len(failed)} 个")
            for f_msg in failed[:10]:
                lines.append(f"  - {f_msg}")

        _cleanup_upload_state(session_id)
        return "\n".join(lines)


# ── Helpers ───────────────────────────────────────────────────

def _col_letter(col_num: int) -> str:
    """Convert column number (1-based) to letter (A, B, ..., Z, AA, ...)."""
    from openpyxl.utils import get_column_letter
    return get_column_letter(col_num)


def _parse_price(value) -> float | None:
    """Parse a price value from cell, handling strings like '$12.50' or '12,500'."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    # Remove currency symbols and commas
    for ch in "$¥€£,，":
        s = s.replace(ch, "")
    s = s.strip()
    if not s:
        return None
    try:
        return float(Decimal(s))
    except (InvalidOperation, ValueError):
        return None


def _heuristic_column_mapping(labels: list[str], cols: list[str]) -> dict:
    """Fallback heuristic when Gemini is unavailable."""
    mapping = {}
    name_keywords = ["name", "品名", "description", "product", "商品", "item"]
    code_keywords = ["code", "sku", "コード", "品番", "item no", "item code"]
    price_keywords = ["price", "単価", "単価", "unit price", "金額", "amount"]
    unit_keywords = ["unit", "単位", "uom"]
    pack_keywords = ["pack", "規格", "spec", "size"]

    for label, col in zip(labels, cols):
        lower = label.lower()
        if not mapping.get("product_name") and any(k in lower for k in name_keywords):
            mapping["product_name"] = col
        elif not mapping.get("product_code") and any(k in lower for k in code_keywords):
            mapping["product_code"] = col
        elif not mapping.get("price") and any(k in lower for k in price_keywords):
            mapping["price"] = col
        elif not mapping.get("unit") and any(k in lower for k in unit_keywords):
            mapping["unit"] = col
        elif not mapping.get("pack_size") and any(k in lower for k in pack_keywords):
            mapping["pack_size"] = col

    return mapping
