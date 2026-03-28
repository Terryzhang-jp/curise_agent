"""Extract complete cell styles from an Excel template using openpyxl.

Produces a JSON-serializable dict that captures all visual formatting:
  - cell_styles: per-cell font, fill, border, alignment, number_format
  - merged_ranges: list of merged cell ranges
  - column_widths: {col_letter: width}
  - row_heights: {row_number: height}

This is the "style layer" that complements the AI "semantic layer" (cell_map).
Together they form a complete template_config for deterministic Excel generation.
"""

from __future__ import annotations

import io
import logging
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


def extract_template_styles(file_bytes: bytes) -> dict[str, Any]:
    """Extract all styles from an Excel template.

    Returns:
        {
            "cell_styles": {"A1": {...}, "B3": {...}, ...},
            "merged_ranges": ["A1:D1", "E5:G5", ...],
            "column_widths": {"A": 12.5, "B": 20.0, ...},
            "row_heights": {"1": 30.0, "5": 18.75, ...},
        }
    """
    wb = load_workbook(io.BytesIO(file_bytes), data_only=False)
    ws = wb.active

    result: dict[str, Any] = {
        "cell_styles": {},
        "merged_ranges": [],
        "column_widths": {},
        "row_heights": {},
    }

    # 1. Merged ranges
    result["merged_ranges"] = [str(r) for r in ws.merged_cells.ranges]

    # 2. Column widths
    for col_idx in range(1, (ws.max_column or 1) + 1):
        letter = get_column_letter(col_idx)
        dim = ws.column_dimensions.get(letter)
        if dim and dim.width is not None:
            result["column_widths"][letter] = round(dim.width, 2)

    # 3. Row heights
    for row_idx in range(1, (ws.max_row or 1) + 1):
        dim = ws.row_dimensions.get(row_idx)
        if dim and dim.height is not None:
            result["row_heights"][str(row_idx)] = round(dim.height, 2)

    # 4. Determine the content boundary (max col/row with actual data or merged ranges)
    max_content_col = 1
    max_content_row = 1
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.value is not None and not isinstance(cell, MergedCell):
                max_content_col = max(max_content_col, cell.column)
                max_content_row = max(max_content_row, cell.row)
    for mr in ws.merged_cells.ranges:
        max_content_col = max(max_content_col, mr.max_col)
        max_content_row = max(max_content_row, mr.max_row)
    # Add small buffer for surrounding styling
    max_content_col = min(max_content_col + 1, ws.max_column or max_content_col + 1)
    max_content_row = min(max_content_row + 2, ws.max_row or max_content_row + 2)

    # 5. Cell styles (only within content boundary, skip default-only cells)
    for row in ws.iter_rows(min_row=1, max_row=max_content_row, min_col=1, max_col=max_content_col):
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            style = _extract_cell_style(cell)
            if style:  # Only store if there's actual formatting
                pos = f"{get_column_letter(cell.column)}{cell.row}"
                result["cell_styles"][pos] = style

    logger.info(
        "Extracted styles: %d cells, %d merged ranges, %d col widths, %d row heights",
        len(result["cell_styles"]),
        len(result["merged_ranges"]),
        len(result["column_widths"]),
        len(result["row_heights"]),
    )

    return result


def _extract_cell_style(cell) -> dict[str, Any] | None:
    """Extract style from a single cell. Returns None if all defaults."""
    style: dict[str, Any] = {}

    # Font
    font = cell.font
    if font and (font.name or font.size or font.bold or font.italic or font.color):
        f: dict[str, Any] = {}
        if font.name:
            f["name"] = font.name
        if font.size:
            f["size"] = font.size
        if font.bold:
            f["bold"] = True
        if font.italic:
            f["italic"] = True
        if font.underline and font.underline != "none":
            f["underline"] = font.underline
        color_str = _color_to_str(font.color)
        if color_str:
            f["color"] = color_str
        if f:
            style["font"] = f

    # Fill
    fill = cell.fill
    if fill and fill.fill_type and fill.fill_type != "none":
        fl: dict[str, Any] = {"type": fill.fill_type}
        fg = _color_to_str(fill.fgColor)
        if fg:
            fl["fg_color"] = fg
        bg = _color_to_str(fill.bgColor)
        if bg and bg != "00000000":
            fl["bg_color"] = bg
        style["fill"] = fl

    # Border
    border = cell.border
    if border:
        b: dict[str, Any] = {}
        for side_name in ("left", "right", "top", "bottom"):
            side = getattr(border, side_name)
            if side and side.style:
                s: dict[str, Any] = {"style": side.style}
                sc = _color_to_str(side.color)
                if sc:
                    s["color"] = sc
                b[side_name] = s
        if b:
            style["border"] = b

    # Alignment
    align = cell.alignment
    if align:
        a: dict[str, Any] = {}
        if align.horizontal:
            a["horizontal"] = align.horizontal
        if align.vertical and align.vertical != "bottom":
            a["vertical"] = align.vertical
        if align.wrap_text:
            a["wrap_text"] = True
        if align.text_rotation:
            a["text_rotation"] = align.text_rotation
        if a:
            style["alignment"] = a

    # Number format
    if cell.number_format and cell.number_format != "General":
        style["number_format"] = cell.number_format

    return style if style else None


def _color_to_str(color) -> str | None:
    """Convert openpyxl Color object to a string."""
    if color is None:
        return None
    if color.type == "rgb" and color.rgb:
        rgb = str(color.rgb)
        # Skip default black "00000000" unless explicitly set
        if rgb == "00000000":
            return None
        return rgb
    if color.type == "theme":
        return f"theme:{color.theme}" + (f"+{color.tint}" if color.tint else "")
    if color.type == "indexed" and color.indexed is not None:
        return f"indexed:{color.indexed}"
    return None


def merge_semantic_and_styles(
    cell_map: dict[str, Any],
    styles: dict[str, Any],
    product_table_config: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """Merge AI semantic analysis (cell_map) with extracted styles into template_config.

    Only stores styles for:
      1. Cells in cell_map (semantic cells — headers, data fields, formulas)
      2. Product area first row (as row template for style cloning)
      3. Layout info (merged ranges, column widths, row heights)

    Returns a compact template_config dict ready to save as template_styles JSON.
    """
    config: dict[str, Any] = {
        "merged_ranges": styles.get("merged_ranges", []),
        "column_widths": styles.get("column_widths", {}),
        "row_heights": styles.get("row_heights", {}),
        "cells": {},
    }

    cell_styles = styles.get("cell_styles", {})

    # Determine product area first row for style template
    product_start_row = None
    if product_table_config:
        product_start_row = product_table_config.get("start_row")

    # Cells to include: cell_map keys + product first row cells
    relevant_positions = set(cell_map.keys())
    if product_start_row:
        for pos, _ in cell_styles.items():
            # Extract row number from position like "A22" → 22
            row_str = "".join(c for c in pos if c.isdigit())
            if row_str and int(row_str) == product_start_row:
                relevant_positions.add(pos)

    for pos in sorted(relevant_positions):
        entry: dict[str, Any] = {}

        # Semantic info from AI
        if pos in cell_map:
            semantic = cell_map[pos]
            entry["source_type"] = semantic.get("source_type")
            entry["writable"] = semantic.get("writable", False)
            entry["data_from"] = semantic.get("data_from")
            entry["field_key"] = semantic.get("field_key")
            entry["label"] = semantic.get("label")
            if semantic.get("formula"):
                entry["formula"] = semantic["formula"]

        # Style info from openpyxl
        if pos in cell_styles:
            entry["style"] = cell_styles[pos]

        if entry:
            config["cells"][pos] = entry

    # Also store product row style template separately for easy access
    if product_start_row:
        row_style: dict[str, Any] = {}
        for pos, st in cell_styles.items():
            row_str = "".join(c for c in pos if c.isdigit())
            if row_str and int(row_str) == product_start_row:
                col = "".join(c for c in pos if c.isalpha())
                row_style[col] = st
        if row_style:
            config["product_row_style"] = row_style

    return config
