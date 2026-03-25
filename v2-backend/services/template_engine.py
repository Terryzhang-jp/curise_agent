"""Deterministic template engine for inquiry Excel generation.

Given a template file + zone_config + order_data → fills and outputs Excel bytes.
Zero LLM calls. All cell positions, formulas, and data mappings are pre-computed
in zone_config (built by zone_config_builder.py at template analysis time).

Pipeline: load template → fill header → resize product zone → fill products
         → write formulas → update cross-refs → clone styles → return bytes
"""

from __future__ import annotations

import copy
import io
import logging
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.workbook.properties import CalcProperties
from openpyxl.utils import get_column_letter, column_index_from_string

from services.field_schema import _resolve_path

logger = logging.getLogger(__name__)


def fill_template(
    template_bytes: bytes,
    zone_config: dict[str, Any],
    order_data: dict[str, Any],
    supplier_id: str | int,
    field_overrides: dict[str, str] | None = None,
) -> bytes:
    """Fill a template with order data using zone config. Returns Excel bytes.

    Args:
        template_bytes: Raw bytes of the template .xlsx file
        zone_config: Complete zone configuration from zone_config_builder
        order_data: Full order data dict (from prepare_inquiry_workspace or equivalent)
        supplier_id: Supplier ID (string or int) for data path resolution
        field_overrides: Optional {cell_ref: value} overrides from user edits,
                         applied on top of resolved header field values

    Returns:
        Filled Excel file as bytes
    """
    sid = str(supplier_id)

    wb = load_workbook(io.BytesIO(template_bytes))
    ws = wb.active

    zones = zone_config["zones"]
    prod_zone = zones["product_data"]
    summ_zone = zones["summary"]

    products = _get_products(order_data, sid)
    if not products:
        raise ValueError(f"No products found for supplier {sid}")

    template_rows = prod_zone["end"] - prod_zone["start"] + 1
    actual_rows = len(products)
    row_delta = actual_rows - template_rows

    logger.info(
        "fill_template: supplier=%s, products=%d, template_rows=%d, delta=%d",
        sid, actual_rows, template_rows, row_delta,
    )

    # ── Step 0: Fill header fields ───────────────────────────────
    header_fields = zone_config.get("header_fields", {})
    for cell_ref, data_path in header_fields.items():
        col_letter = "".join(c for c in cell_ref if c.isalpha())
        row_num = int("".join(c for c in cell_ref if c.isdigit()))
        col_idx = column_index_from_string(col_letter)
        cell = ws.cell(row=row_num, column=col_idx)
        if isinstance(cell, MergedCell):
            continue
        value = _resolve_path(order_data, data_path, sid)
        if value is not None:
            cell.value = value

    # ── Step 0b: Apply user field overrides ────────────────────
    if field_overrides:
        for cell_ref, value in field_overrides.items():
            col_letter = "".join(c for c in cell_ref if c.isalpha())
            row_num = int("".join(c for c in cell_ref if c.isdigit()))
            if not col_letter or not row_num:
                continue
            col_idx = column_index_from_string(col_letter)
            cell = ws.cell(row=row_num, column=col_idx)
            if not isinstance(cell, MergedCell):
                cell.value = value
        logger.info("Applied %d field overrides", len(field_overrides))

    # ── Step 0c: Capture product-row merge pattern from template ──
    # Before row manipulation, detect which columns are merged per product
    # row (e.g. F:G merged for "description"). We'll re-apply after filling.
    row_merges = _detect_row_merge_pattern(ws, prod_zone["start"], col_map)

    # ── Step 1: Resize product zone ──────────────────────────────
    # Strategy: insert fresh rows at the START of the product zone, then
    # delete old template rows. This avoids openpyxl's merge inheritance
    # bug (inserting adjacent to merged summary rows propagates MergedCells).
    new_prod_end = prod_zone["start"] + actual_rows - 1

    if row_delta > 0:
        # Insert fresh rows at the beginning — pushes everything down
        ws.insert_rows(prod_zone["start"], actual_rows)
        # Delete old template rows (now shifted down by actual_rows)
        old_start = prod_zone["start"] + actual_rows
        ws.delete_rows(old_start, template_rows)
    elif row_delta < 0:
        # Fewer products than template rows — delete excess from bottom
        delete_start = prod_zone["start"] + actual_rows
        ws.delete_rows(delete_start, abs(row_delta))

    new_summ_start = new_prod_end + 1
    new_summ_end = new_summ_start + (summ_zone["end"] - summ_zone["start"])

    # ── Step 1b: Remove stale merge ranges in product zone ────────
    # openpyxl does NOT update merge range metadata during insert/delete.
    # Stale ranges from the original summary zone (e.g. I33:J33) stay at
    # old row numbers — now inside the product zone — and corrupt cells
    # on save/reload. Remove any range that overlaps the product zone.
    max_col = ws.max_column or 12
    merges_to_remove = [
        mr for mr in list(ws.merged_cells.ranges)
        if mr.min_row <= new_prod_end and mr.max_row >= prod_zone["start"]
    ]
    for mr in merges_to_remove:
        ws.merged_cells.remove(mr)
    if merges_to_remove:
        logger.info("Removed %d stale merge ranges from product zone", len(merges_to_remove))

    # Belt-and-suspenders: purge any MergedCell objects from _cells
    for row in range(prod_zone["start"], new_prod_end + 1):
        for col in range(1, max_col + 1):
            if (row, col) in ws._cells and isinstance(ws._cells[(row, col)], MergedCell):
                del ws._cells[(row, col)]

    # ── Step 2: Clear product zone values ─────────────────────────
    for row in range(prod_zone["start"], new_prod_end + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.value = None

    # ── Step 3: Clear stale summary data ─────────────────────────
    for row in range(new_summ_start, new_summ_end + 1):
        for col_letter in zone_config.get("stale_columns_in_summary", []):
            col_idx = column_index_from_string(col_letter)
            cell = ws.cell(row=row, column=col_idx)
            if not isinstance(cell, MergedCell):
                cell.value = None

    # ── Step 4: Fill product data ────────────────────────────────
    col_map = zone_config.get("product_columns", {})
    currency = order_data.get("currency") or "JPY"
    po_number = order_data.get("po_number", "")

    for i, product in enumerate(products):
        row = prod_zone["start"] + i
        for col_letter, field_name in col_map.items():
            col_idx = column_index_from_string(col_letter)
            cell = ws.cell(row=row, column=col_idx)
            if isinstance(cell, MergedCell):
                continue
            cell.value = _resolve_product_field(
                field_name, product, i, po_number, currency,
            )

    # ── Step 5: Product row formulas ─────────────────────────────
    for col_letter, formula_tpl in zone_config.get("product_row_formulas", {}).items():
        col_idx = column_index_from_string(col_letter)
        for row in range(prod_zone["start"], new_prod_end + 1):
            cell = ws.cell(row=row, column=col_idx)
            if isinstance(cell, MergedCell):
                continue
            cell.value = formula_tpl.replace("{row}", str(row))

    # ── Step 6: Summary formulas ─────────────────────────────────
    formula_cells = {}  # placeholders for cross-references

    for idx, sf in enumerate(zone_config.get("summary_formulas", [])):
        col_letter = "".join(c for c in sf["cell"] if c.isalpha())
        col_idx = column_index_from_string(col_letter)
        target_row = new_summ_start + idx

        if sf["type"] == "product_sum":
            col = sf.get("col", col_letter)
            formula = f"=SUM({col}{prod_zone['start']}:{col}{new_prod_end})"
            formula_cells["sum_cell"] = f"{col_letter}{target_row}"
        elif sf["type"] == "relative":
            formula = sf["formula_template"]
            for key, ref in formula_cells.items():
                formula = formula.replace(f"{{{key}}}", ref)
            # Track tax cell
            label = sf.get("label", "")
            if "tax" in label.lower() or "*0.08" in sf.get("formula_template", ""):
                formula_cells["tax_cell"] = f"{col_letter}{target_row}"
        else:
            continue

        ws.cell(row=target_row, column=col_idx).value = formula

        # Track grand total
        label = sf.get("label", "")
        if "grand" in label.lower() or "total" in label.upper():
            formula_cells["grand_total_cell"] = f"{col_letter}{target_row}"

    # ── Step 6b: External cross-references ───────────────────────
    for ext in zone_config.get("external_refs", []):
        cell_ref = ext["cell"]
        col_letter = "".join(c for c in cell_ref if c.isalpha())
        row_num = int("".join(c for c in cell_ref if c.isdigit()))
        col_idx = column_index_from_string(col_letter)
        formula = ext["formula_template"]
        for key, ref in formula_cells.items():
            formula = formula.replace(f"{{{key}}}", ref)
        ws.cell(row=row_num, column=col_idx).value = formula

    # ── Step 7: Restore summary static values ────────────────────
    for cell_ref, value in zone_config.get("summary_static_values", {}).items():
        col_letter = "".join(c for c in cell_ref if c.isalpha())
        orig_row = int("".join(c for c in cell_ref if c.isdigit()))
        offset = orig_row - summ_zone["start"]
        new_row = new_summ_start + offset
        col_idx = column_index_from_string(col_letter)
        ws.cell(row=new_row, column=col_idx).value = value

    # ── Step 8: Clone styles ─────────────────────────────────────
    src_row = prod_zone["start"]
    for row in range(src_row + 1, new_prod_end + 1):
        for col in range(1, max_col + 1):
            src = ws.cell(row=src_row, column=col)
            dst = ws.cell(row=row, column=col)
            if isinstance(src, MergedCell) or isinstance(dst, MergedCell):
                continue
            if src.has_style:
                dst.font = copy.copy(src.font)
                dst.fill = copy.copy(src.fill)
                dst.border = copy.copy(src.border)
                dst.alignment = copy.copy(src.alignment)
                dst.number_format = src.number_format

    # ── Step 8b: Re-apply product-row merges ──────────────────────
    # Restore per-row merges captured in Step 0c (e.g. F:G for description).
    if row_merges:
        for row in range(prod_zone["start"], new_prod_end + 1):
            for start_col, end_col in row_merges:
                start_letter = get_column_letter(start_col)
                end_letter = get_column_letter(end_col)
                ws.merge_cells(f"{start_letter}{row}:{end_letter}{row}")
        logger.info(
            "Re-applied %d column merge(s) per product row (%d rows)",
            len(row_merges), actual_rows,
        )

    # ── Step 9: Ensure minimum column widths ─────────────────────
    # Templates with merged headers often leave data columns too narrow.
    # After filling, enforce minimums based on field semantics so content
    # is always readable when opened in Excel.
    _ensure_column_widths(ws, col_map, zone_config.get("product_row_formulas", {}))

    # ── Force recalculation on open ──────────────────────────────
    wb.calculation = CalcProperties(fullCalcOnLoad=True)

    # ── Save to bytes ────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    logger.info(
        "fill_template complete: %d products, output=%d bytes",
        actual_rows, buf.getbuffer().nbytes,
    )
    return buf.read()


def verify_output(
    excel_bytes: bytes,
    zone_config: dict[str, Any],
    order_data: dict[str, Any],
    supplier_id: str | int,
) -> dict[str, Any]:
    """Round-trip verification: read back generated Excel and compare against source data.

    Returns: {"ok": bool, "errors": [...], "checks": int}
    """
    sid = str(supplier_id)
    wb = load_workbook(io.BytesIO(excel_bytes))
    ws = wb.active

    products = _get_products(order_data, sid)
    prod_start = zone_config["zones"]["product_data"]["start"]
    new_prod_end = prod_start + len(products) - 1
    new_summ_start = new_prod_end + 1

    errors: list[str] = []
    checks = 0

    # ── 1. Header fields ─────────────────────────────────────────
    for cell_ref, data_path in zone_config.get("header_fields", {}).items():
        col_letter = "".join(c for c in cell_ref if c.isalpha())
        row_num = int("".join(c for c in cell_ref if c.isdigit()))
        col_idx = column_index_from_string(col_letter)

        expected = _resolve_path(order_data, data_path, sid)
        actual = ws.cell(row=row_num, column=col_idx).value
        checks += 1

        if expected and actual != expected:
            # Allow string conversion mismatch (e.g., int vs str)
            if str(actual) != str(expected):
                errors.append(f"Header {cell_ref}: expected '{expected}', got '{actual}'")

    # ── 2. Product data ──────────────────────────────────────────
    col_map = zone_config.get("product_columns", {})
    po_number = order_data.get("po_number", "")
    currency = order_data.get("currency") or "JPY"

    for i, product in enumerate(products):
        row = prod_start + i
        for col_letter, field_name in col_map.items():
            col_idx = column_index_from_string(col_letter)
            cell = ws.cell(row=row, column=col_idx)
            if isinstance(cell, MergedCell):
                continue

            expected = _resolve_product_field(field_name, product, i, po_number, currency)
            actual = cell.value
            checks += 1

            if expected is None:
                continue
            if actual is None:
                errors.append(f"Row {row} {col_letter} ({field_name}): expected '{expected}', got None")
                continue

            # Numeric comparison
            if isinstance(expected, (int, float)) and isinstance(actual, (int, float)):
                if abs(float(expected) - float(actual)) > 0.001:
                    errors.append(f"Row {row} {col_letter}: expected {expected}, got {actual}")
            elif str(actual) != str(expected):
                errors.append(f"Row {row} {col_letter}: expected '{expected}', got '{actual}'")

    # ── 3. Product row formulas ──────────────────────────────────
    for col_letter, formula_tpl in zone_config.get("product_row_formulas", {}).items():
        col_idx = column_index_from_string(col_letter)
        for row in range(prod_start, new_prod_end + 1):
            cell = ws.cell(row=row, column=col_idx)
            if isinstance(cell, MergedCell):
                continue
            expected = formula_tpl.replace("{row}", str(row))
            checks += 1
            if cell.value != expected:
                errors.append(f"Formula {col_letter}{row}: expected '{expected}', got '{cell.value}'")
                if len(errors) > 20:
                    errors.append("... (truncated)")
                    break

    # ── 4. Summary formulas ──────────────────────────────────────
    for idx, sf in enumerate(zone_config.get("summary_formulas", [])):
        col_letter = "".join(c for c in sf["cell"] if c.isalpha())
        col_idx = column_index_from_string(col_letter)
        target_row = new_summ_start + idx
        actual = ws.cell(row=target_row, column=col_idx).value
        checks += 1
        if actual is None or not str(actual).startswith("="):
            errors.append(f"Summary {col_letter}{target_row}: expected formula, got '{actual}'")

    # ── 5. Stale data check ──────────────────────────────────────
    summ_zone = zone_config["zones"]["summary"]
    new_summ_end = new_summ_start + (summ_zone["end"] - summ_zone["start"])
    for row in range(new_summ_start, new_summ_end + 1):
        for col_letter in zone_config.get("stale_columns_in_summary", []):
            col_idx = column_index_from_string(col_letter)
            cell = ws.cell(row=row, column=col_idx)
            if isinstance(cell, MergedCell):
                continue
            checks += 1
            # Check if it's a known static value
            orig_row = summ_zone["start"] + (row - new_summ_start)
            orig_ref = f"{col_letter}{orig_row}"
            if orig_ref in zone_config.get("summary_static_values", {}):
                continue
            if cell.value is not None:
                errors.append(f"Stale data {col_letter}{row}: '{cell.value}'")

    # ── 6. Numeric type check ────────────────────────────────────
    numeric_fields = {"quantity", "unit_price"}
    for i in range(min(5, len(products))):
        row = prod_start + i
        for col_letter, field_name in col_map.items():
            if field_name not in numeric_fields:
                continue
            col_idx = column_index_from_string(col_letter)
            val = ws.cell(row=row, column=col_idx).value
            checks += 1
            if val is not None and isinstance(val, str):
                errors.append(f"Row {row} {col_letter} ({field_name}): should be numeric, got string '{val}'")

    result = {
        "ok": len(errors) == 0,
        "errors": errors,
        "checks": checks,
        "product_count": len(products),
    }

    if errors:
        logger.warning("verify_output: %d errors in %d checks", len(errors), checks)
        for e in errors[:5]:
            logger.warning("  %s", e)
    else:
        logger.info("verify_output: all %d checks passed", checks)

    return result


# ── Private helpers ──────────────────────────────────────────────


def _detect_row_merge_pattern(
    ws, first_row: int, col_map: dict[str, str],
) -> list[tuple[int, int]]:
    """Detect per-row column merges in the first product row of the template.

    Scans merge ranges that span exactly one row (the first product row) and
    cover columns used in the product table. Returns a list of (start_col, end_col)
    tuples representing merges to replicate per product row.

    Also infers merges from "gaps" in col_map: if column G has no mapping but
    F and H do, F:G is a merge (the template visually merges them).
    """
    merges: list[tuple[int, int]] = []

    # Method 1: explicit merge ranges on the first product row
    for mr in ws.merged_cells.ranges:
        if mr.min_row == first_row and mr.max_row == first_row:
            merges.append((mr.min_col, mr.max_col))

    # Method 2: infer from gaps in col_map
    if not merges and col_map:
        mapped_cols = sorted(column_index_from_string(c) for c in col_map)
        for i in range(len(mapped_cols) - 1):
            gap = mapped_cols[i + 1] - mapped_cols[i]
            if gap > 1:
                # Columns between mapped_cols[i] and mapped_cols[i+1] are skipped
                merges.append((mapped_cols[i], mapped_cols[i + 1] - 1))

    if merges:
        logger.info(
            "Detected %d product-row merge pattern(s): %s",
            len(merges),
            [(get_column_letter(s), get_column_letter(e)) for s, e in merges],
        )
    return merges


def _get_products(order_data: dict, sid: str) -> list[dict]:
    """Extract product list for a supplier from order_data."""
    supplier = order_data.get("suppliers", {}).get(sid, {})
    return supplier.get("products", [])



def _resolve_product_field(
    field_name: str,
    product: dict,
    index: int,
    po_number: str,
    currency: str,
) -> Any:
    """Resolve a product column value from the product dict."""
    if field_name == "line_number" or field_name == "__line_number__":
        return index + 1
    if field_name == "po_number" or field_name == "__po_number__":
        return po_number
    if field_name == "currency" or field_name == "__currency__":
        return product.get("currency") or currency

    if field_name == "product_code":
        return product.get("product_code", "")
    if field_name in ("product_name", "product_name_en"):
        return product.get("product_name", "")
    if field_name == "product_name_jp":
        return product.get("product_name_jp", "")
    if field_name == "description":
        return product.get("pack_size", "")

    if field_name == "quantity":
        val = product.get("quantity")
        return _to_number(val)
    if field_name == "unit_price":
        val = product.get("unit_price")
        return _to_number(val)
    if field_name == "unit":
        return product.get("unit", "CT")

    # Generic fallback
    return product.get(field_name, "")


def _to_number(val: Any) -> int | float | None:
    """Convert value to number, keeping None as None."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return val
    if isinstance(val, str):
        try:
            f = float(val)
            return int(f) if f == int(f) else f
        except (ValueError, TypeError):
            return val
    return val


# Minimum column widths by field type (in Excel character units).
# Templates with merged headers often leave data columns too narrow;
# after filling and unmerging, content overflows visually.
# These minimums ensure readability without disrupting intentional layout.
_FIELD_MIN_WIDTHS: dict[str, float] = {
    "line_number": 4.5,
    "po_number": 14,
    "product_code": 14,
    "product_name": 28,
    "product_name_en": 28,
    "product_name_jp": 20,
    "description": 12,
    "quantity": 8,
    "unit": 6,
    "unit_price": 10,
    "currency": 7,
}

# Formula columns (amounts, totals) also need readable width.
_FORMULA_MIN_WIDTH: float = 12


def _ensure_column_widths(
    ws,
    col_map: dict[str, str],
    formula_cols: dict[str, str],
) -> None:
    """Widen columns that are narrower than the minimum for their field type.

    Only increases width — never shrinks columns that are already wide enough.
    This preserves the template's intended layout while preventing overflow.
    """
    for col_letter, field_name in col_map.items():
        min_w = _FIELD_MIN_WIDTHS.get(field_name)
        if min_w is None:
            min_w = 10  # sensible default for unknown fields
        current = ws.column_dimensions[col_letter].width or 0
        if current < min_w:
            ws.column_dimensions[col_letter].width = min_w

    for col_letter in formula_cols:
        current = ws.column_dimensions[col_letter].width or 0
        if current < _FORMULA_MIN_WIDTH:
            ws.column_dimensions[col_letter].width = _FORMULA_MIN_WIDTH
