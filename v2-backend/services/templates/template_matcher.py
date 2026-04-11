"""
Template Matcher — match uploaded orders to known OrderFormatTemplates.

Cascading Multi-Signal Matcher (0 LLM):
  Phase 1: Fingerprint exact match (Excel only, precision=100%)
  Phase 2: source_company match (precision≈95%)
  Phase 3: IDF-weighted keyword scoring (precision≈80%)

Other functions:
- get_scannable_text: extract searchable text from file (0 LLM for Excel, pdfplumber for PDF)
- build_guided_prompt: construct template-guided extraction prompt
- extract_excel_deterministic: 0 LLM extraction when column_mapping is complete
"""

from __future__ import annotations

import io
import logging
from collections import Counter
from typing import Optional

from core.models import OrderFormatTemplate

logger = logging.getLogger(__name__)


# ─── Template Matching — Cascading Multi-Signal (0 LLM) ─────────

def find_matching_template(
    scannable_text: str,
    db,
    file_bytes: bytes | None = None,
    file_type: str | None = None,
) -> tuple[Optional[OrderFormatTemplate], Optional[str]]:
    """Cascading multi-signal template matching. Returns (template, method) or (None, None).

    Phase 1: fingerprint  — Excel header hash exact match (100% precision)
    Phase 2: source_company — company name substring match in document text
    Phase 3: keyword_idf  — IDF-weighted keyword scoring
    """
    all_templates = db.query(OrderFormatTemplate).filter(
        OrderFormatTemplate.is_active == True,
    ).all()

    if not all_templates:
        return None, None

    # ── Phase 1: Fingerprint exact match (Excel only) ──
    if file_bytes and file_type and file_type != "pdf":
        result = _phase1_fingerprint(file_bytes, all_templates)
        if result:
            return result

    # ── Phase 2: source_company match ──
    if scannable_text:
        result = _phase2_source_company(scannable_text, all_templates)
        if result:
            return result

    # ── Phase 3: IDF-weighted keyword scoring ──
    if scannable_text:
        result = _phase3_keyword_idf(scannable_text, all_templates)
        if result:
            return result

    return None, None


def _phase1_fingerprint(
    file_bytes: bytes, templates: list[OrderFormatTemplate],
) -> tuple[OrderFormatTemplate, str] | None:
    """Phase 1: Exact fingerprint match for Excel files."""
    from services.excel.excel_parser import compute_fingerprint
    from openpyxl import load_workbook

    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb[wb.sheetnames[0]]
        headers = []
        header_row = 1
        for cell in ws[header_row]:
            if cell.value is not None:
                headers.append(str(cell.value))
        if not headers:
            return None

        fp = compute_fingerprint(headers)

        for tpl in templates:
            if tpl.format_fingerprint and tpl.format_fingerprint == fp:
                logger.info(
                    "Phase 1 HIT: fingerprint '%s' → template '%s' (id=%d)",
                    fp, tpl.name, tpl.id,
                )
                return tpl, "fingerprint"

        logger.debug("Phase 1 MISS: fingerprint '%s' matched no template", fp)
    except Exception as e:
        logger.warning("Phase 1 fingerprint computation failed: %s", e)

    return None


def _phase2_source_company(
    scannable_text: str, templates: list[OrderFormatTemplate],
) -> tuple[OrderFormatTemplate, str] | None:
    """Phase 2: Match source_company name in document text."""
    text_upper = scannable_text.upper()

    candidates = []
    for tpl in templates:
        company = tpl.source_company
        if not company or len(company.strip()) < 2:
            continue
        if company.upper() in text_upper:
            candidates.append(tpl)

    if len(candidates) == 1:
        tpl = candidates[0]
        logger.info(
            "Phase 2 HIT: source_company '%s' → template '%s' (id=%d)",
            tpl.source_company, tpl.name, tpl.id,
        )
        return tpl, "source_company"

    if len(candidates) > 1:
        names = [c.name for c in candidates]
        logger.info("Phase 2 AMBIGUOUS: %d templates matched by source_company: %s → fall through to Phase 3", len(candidates), names)

    return None


def _phase3_keyword_idf(
    scannable_text: str, templates: list[OrderFormatTemplate],
) -> tuple[OrderFormatTemplate, str] | None:
    """Phase 3: IDF-weighted keyword scoring."""
    text_upper = scannable_text.upper()

    # Collect all keywords across templates and compute IDF weights
    # IDF(keyword) = 1 / (number of templates that contain this keyword)
    keyword_template_count: Counter[str] = Counter()
    templates_with_keywords = []

    for tpl in templates:
        keywords = tpl.match_keywords or []
        if not keywords:
            continue
        templates_with_keywords.append(tpl)
        seen = set()
        for kw in keywords:
            kw_upper = kw.upper()
            if kw_upper not in seen:
                keyword_template_count[kw_upper] += 1
                seen.add(kw_upper)

    if not templates_with_keywords:
        return None

    # Score each template
    best_tpl = None
    best_score = 0.0
    best_hits = 0

    for tpl in templates_with_keywords:
        keywords = tpl.match_keywords or []
        score = 0.0
        hits = 0
        for kw in keywords:
            kw_upper = kw.upper()
            if kw_upper in text_upper:
                idf = 1.0 / keyword_template_count[kw_upper]
                score += idf
                hits += 1

        if score > best_score:
            best_score = score
            best_hits = hits
            best_tpl = tpl

    # Require at least 1 keyword hit
    if best_tpl and best_hits >= 1:
        logger.info(
            "Phase 3 HIT: template '%s' (id=%d) — score=%.2f, %d keyword hits",
            best_tpl.name, best_tpl.id, best_score, best_hits,
        )
        return best_tpl, "keyword_idf"

    return None


# ─── Scannable Text Extraction ──────────────────────────────────

def get_scannable_text(file_bytes: bytes, file_type: str) -> str:
    """Extract searchable text from file for keyword matching. No LLM calls."""
    if file_type != "pdf":
        return _excel_to_scannable(file_bytes)
    return _pdf_to_scannable(file_bytes)


def _excel_to_scannable(file_bytes: bytes) -> str:
    """Read all cell values from Excel and join as text."""
    from openpyxl import load_workbook

    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
        parts = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) for c in row if c is not None]
                if cells:
                    parts.append(" ".join(cells))
        return "\n".join(parts)
    except Exception as e:
        logger.warning("Failed to extract Excel text: %s", e)
        return ""


def _pdf_to_scannable(file_bytes: bytes) -> str:
    """Extract text from PDF. Tries pdfplumber first, falls back to Vision for image-based PDFs."""
    text = ""

    # Try pdfplumber (instant, 0 LLM)
    try:
        import pdfplumber
        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            parts = []
            for page in pdf.pages[:3]:
                t = page.extract_text()
                if t:
                    parts.append(t)
            text = "\n".join(parts)
    except ImportError:
        logger.warning("pdfplumber not installed")
    except Exception as e:
        logger.warning("pdfplumber failed: %s", e)

    if text.strip():
        return text

    # Fallback: image-based PDF → Vision read first page only (~2-3s, 1 LLM call)
    logger.info("PDF has no extractable text, using Vision fallback on page 1")
    try:
        return _pdf_vision_scannable(file_bytes)
    except Exception as e:
        logger.warning("Vision fallback failed: %s", e)
        return ""


def _pdf_vision_scannable(file_bytes: bytes) -> str:
    """Read first page of image-based PDF with Gemini Vision for keyword extraction."""
    from pdf2image import convert_from_bytes
    from google import genai
    from google.genai import types
    from core.config import settings

    images = convert_from_bytes(file_bytes, dpi=150, first_page=1, last_page=1)
    if not images:
        return ""

    # Convert to JPEG bytes
    img_buf = io.BytesIO()
    images[0].convert("RGB").save(img_buf, format="JPEG", quality=80)
    img_bytes = img_buf.getvalue()

    client = genai.Client(api_key=settings.GOOGLE_API_KEY)
    response = client.models.generate_content(
        model="gemini-2.0-flash",
        contents=[
            types.Part.from_text(text="Read all visible text on this page. Return the raw text only, no formatting or analysis."),
            types.Part.from_bytes(data=img_bytes, mime_type="image/jpeg"),
        ],
        config=types.GenerateContentConfig(
            temperature=0.0,
            max_output_tokens=2000,
        ),
    )
    text = response.text or ""
    logger.info("Vision fallback extracted %d chars from page 1", len(text))
    return text


# ─── Guided Prompt Construction ─────────────────────────────────

def build_guided_prompt(template: OrderFormatTemplate, field_definitions=None) -> str:
    """Build a template-guided extraction prompt with known layout/column info."""
    parts = ["分析这个采购订单文档，提取所有可见信息。\n"]

    if template.source_company:
        parts.append(f"## 已知信息\n来源公司: {template.source_company}\n")

    if template.layout_prompt:
        parts.append(f"## 文档布局\n{template.layout_prompt}\n")

    if template.column_mapping:
        parts.append("## 列映射（已知各列含义）")
        for col, field_key in template.column_mapping.items():
            parts.append(f"- 列 {col} = {field_key}")
        parts.append("")

    if template.extracted_fields:
        parts.append("## 已知元数据字段")
        for f in template.extracted_fields:
            parts.append(f"- {f.get('label', '')}: 字段名 {f.get('key', '')}")
        parts.append("")

    if field_definitions:
        parts.append("## 标准字段名（请使用这些字段名）")
        for fd in field_definitions:
            hint = f" ({fd.extraction_hint})" if fd.extraction_hint else ""
            parts.append(f"- {fd.field_key}: {fd.field_label}{hint}")
        parts.append("")

    # Append standard metadata schema
    from services.orders.order_processor import ORDER_METADATA_SCHEMA

    parts.append("## 元数据字段（必须使用以下确切键名）")
    for k, v in ORDER_METADATA_SCHEMA.items():
        parts.append(f"- {k}: {v}")
    parts.append("- extra_fields: 其他可见元数据字段")
    parts.append("vendor_name 必须是纯字符串，不是对象。日期格式必须为 YYYY-MM-DD。total_amount 必须是数字。看不到的字段用 null。")
    parts.append("")

    parts.append("""返回 JSON：
{
  "order_metadata": {
    "po_number": "...",
    "ship_name": "...",
    "vendor_name": "...",
    "delivery_date": "YYYY-MM-DD",
    "order_date": "YYYY-MM-DD",
    "currency": "...",
    "destination_port": "...",
    "total_amount": 数字,
    "extra_fields": {}
  },
  "products": [
    { "line_number": N, "product_code": "...", "product_name": "...",
      "quantity": N, "unit": "...", "unit_price": N, "total_price": N }
  ]
}
只提取文档中可见的信息，不要编造。数字使用数值类型。返回纯 JSON，不要 markdown 代码块。""")

    return "\n".join(parts)


# ─── Deterministic Excel Extraction (0 LLM) ────────────────────

def extract_excel_deterministic(file_bytes: bytes, template: OrderFormatTemplate) -> dict:
    """Extract from Excel using template column_mapping — 0 LLM calls."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb[wb.sheetnames[0]]

    col_map = template.column_mapping  # {"A": "product_name", "B": "quantity", ...}
    if not col_map:
        raise ValueError("Template has no column_mapping")

    # Extract products from data rows
    products = []
    for row_idx in range(template.data_start_row, ws.max_row + 1):
        product = {}
        has_data = False
        for col_letter, field_key in col_map.items():
            cell_ref = f"{col_letter}{row_idx}"
            val = ws[cell_ref].value
            if val is not None:
                has_data = True
            product[field_key] = val
        if has_data and (product.get("product_name") or product.get("product_code")):
            product["line_number"] = row_idx - template.data_start_row + 1
            products.append(product)

    # Extract header metadata from rows above data
    metadata = _extract_header_metadata(ws, template)

    return {
        "order_metadata": metadata,
        "products": products,
        "extraction_method": "template_deterministic",
        "template_id": template.id,
    }


def _extract_header_metadata(ws, template: OrderFormatTemplate) -> dict:
    """Extract metadata from header area (rows 1 to header_row-1).

    Uses extracted_fields if available, otherwise scans for common patterns.
    """
    metadata = {}

    scan_max_row = max(template.data_start_row - 1, 1)

    # If template has extracted_fields with positions, use them
    if template.extracted_fields:
        for field in template.extracted_fields:
            key = field.get("key", "")
            label = field.get("label", "")
            if not key:
                continue
            # Try to find value in header area by scanning cells
            # extracted_fields typically has label/key but not exact positions,
            # so scan the header rows for the label text and grab the adjacent cell
            found = False
            for row in ws.iter_rows(min_row=1, max_row=scan_max_row, values_only=False):
                for cell in row:
                    if cell.value and label and str(cell.value).strip() == label.strip():
                        # Value is likely in the next cell to the right
                        next_col = cell.column + 1
                        if next_col <= ws.max_column:
                            val_cell = ws.cell(row=cell.row, column=next_col)
                            if val_cell.value is not None:
                                metadata[key] = val_cell.value
                                found = True
                                break
                if found:
                    break

    # Also scan for common metadata patterns in first few rows
    for row in ws.iter_rows(min_row=1, max_row=scan_max_row, values_only=False):
        for cell in row:
            val = str(cell.value).strip() if cell.value else ""
            val_lower = val.lower()
            # Common label → key mappings
            label_map = {
                "po number": "po_number", "po no": "po_number", "purchase order": "po_number",
                "delivery date": "delivery_date", "deliver on": "delivery_date",
                "ship name": "ship_name", "vessel": "ship_name",
                "currency": "currency",
                "destination": "destination_port",
            }
            for pattern, meta_key in label_map.items():
                if pattern in val_lower and meta_key not in metadata:
                    next_col = cell.column + 1
                    if next_col <= ws.max_column:
                        val_cell = ws.cell(row=cell.row, column=next_col)
                        if val_cell.value is not None:
                            metadata[meta_key] = val_cell.value

    return metadata
