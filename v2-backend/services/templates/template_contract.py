"""Template contract extraction for deterministic Excel validation.

The contract captures structure facts from the template itself so runtime
verification can validate against the workbook's own invariants rather than
hard-coded global business assumptions.
"""

from __future__ import annotations

import io
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def build_template_contract(file_bytes: bytes, zone_config: dict[str, Any]) -> dict[str, Any]:
    """Extract workbook structure invariants from the template file.

    The contract intentionally stores only facts that can be derived
    deterministically from the workbook and zone_config.
    """
    wb = load_workbook(io.BytesIO(file_bytes), data_only=False)
    ws = wb.active

    prod_zone = zone_config["zones"]["product_data"]
    summ_zone = zone_config["zones"]["summary"]
    prod_start = prod_zone["start"]
    prod_end = prod_zone["end"]
    summ_start = summ_zone["start"]
    summ_end = summ_zone["end"]

    summary_static = zone_config.get("summary_static_values", {})

    contract: dict[str, Any] = {
        "version": 2,
        "sheet_name": ws.title,
        "template_dimensions": ws.calculate_dimension(),
        "header_merged_ranges": [],
        "header_field_anchors": [],
        "product_header_cells": [],
        "product_row_merges": [],
        "summary_relative_merges": [],
        "summary_static_labels": [],
        "formula_anchors": [],
    }

    header_field_cells = set(zone_config.get("header_fields", {}).keys())
    for cell_ref in sorted(header_field_cells):
        anchor = _find_header_anchor(ws, cell_ref, header_field_cells)
        if anchor:
            contract["header_field_anchors"].append(anchor)

    header_row = prod_start - 1
    if header_row >= 1:
        product_cols = {
            *zone_config.get("product_columns", {}).keys(),
            *zone_config.get("product_row_formulas", {}).keys(),
        }
        for col_letter in sorted(product_cols, key=lambda col: _column_sort_key(col)):
            value = ws[f"{col_letter}{header_row}"].value
            if value is None:
                continue
            if isinstance(value, str) and not value.strip():
                continue
            contract["product_header_cells"].append(
                {
                    "col": col_letter,
                    "row_offset": -1,
                    "value": value,
                }
            )

    for mr in ws.merged_cells.ranges:
        if mr.max_row < prod_start:
            contract["header_merged_ranges"].append(str(mr))
            continue

        if mr.min_row >= prod_start and mr.max_row <= prod_end:
            if mr.min_row == mr.max_row == prod_start:
                contract["product_row_merges"].append(
                    {"start_col": mr.min_col, "end_col": mr.max_col}
                )
            continue

        if mr.min_row >= summ_start and mr.max_row <= summ_end:
            contract["summary_relative_merges"].append(
                {
                    "row_offset": mr.min_row - summ_start,
                    "start_col": mr.min_col,
                    "end_col": mr.max_col,
                }
            )

    summary_formula_cells = {
        sf["cell"] for sf in zone_config.get("summary_formulas", []) if sf.get("cell")
    }
    for cell_ref, value in summary_static.items():
        if cell_ref in summary_formula_cells:
            continue
        if isinstance(value, str) and value.strip():
            row_num = int("".join(c for c in cell_ref if c.isdigit()))
            col_letter = "".join(c for c in cell_ref if c.isalpha())
            contract["summary_static_labels"].append(
                {
                    "row_offset": row_num - summ_start,
                    "col": col_letter,
                    "value": value,
                }
            )

    for sf in zone_config.get("summary_formulas", []):
        cell_ref = sf.get("cell")
        if not cell_ref:
            continue
        row_num = int("".join(c for c in cell_ref if c.isdigit()))
        col_letter = "".join(c for c in cell_ref if c.isalpha())
        contract["formula_anchors"].append(
            {
                "row_offset": row_num - summ_start,
                "col": col_letter,
                "label": sf.get("label", ""),
                "type": sf.get("type", ""),
            }
        )

    contract["header_merged_ranges"].sort()
    contract["header_field_anchors"].sort(
        key=lambda item: (item["target_cell"], item["anchor_cell"])
    )
    contract["product_header_cells"].sort(
        key=lambda item: (_column_sort_key(item["col"]), item["row_offset"])
    )
    contract["product_row_merges"].sort(key=lambda item: (item["start_col"], item["end_col"]))
    contract["summary_relative_merges"].sort(
        key=lambda item: (item["row_offset"], item["start_col"], item["end_col"])
    )
    contract["summary_static_labels"].sort(
        key=lambda item: (item["row_offset"], item["col"], item["value"])
    )
    contract["formula_anchors"].sort(
        key=lambda item: (item["row_offset"], item["col"], item["label"])
    )

    return contract


def verify_template_contract(
    ws,
    template_contract: dict[str, Any],
    product_start_row: int,
    product_count: int,
) -> tuple[list[str], int]:
    """Verify workbook structure against the stored template contract."""
    errors: list[str] = []
    checks = 0

    if ws.title != template_contract.get("sheet_name"):
        errors.append(
            f"Sheet title mismatch: expected '{template_contract.get('sheet_name')}', got '{ws.title}'"
        )
    checks += 1

    actual_merges = {str(rng) for rng in ws.merged_cells.ranges}
    for merge_ref in template_contract.get("header_merged_ranges", []):
        checks += 1
        if merge_ref not in actual_merges:
            errors.append(f"Missing header merge: {merge_ref}")

    for anchor in template_contract.get("header_field_anchors", []):
        checks += 1
        actual = ws[anchor["anchor_cell"]].value
        if actual != anchor["value"]:
            errors.append(
                f"Header anchor {anchor['anchor_cell']}: expected '{anchor['value']}', got '{actual}'"
            )

    for header in template_contract.get("product_header_cells", []):
        row = product_start_row + header.get("row_offset", -1)
        cell_ref = f"{header['col']}{row}"
        checks += 1
        actual = ws[cell_ref].value
        if actual != header["value"]:
            errors.append(
                f"Product header {cell_ref}: expected '{header['value']}', got '{actual}'"
            )

    for merge in template_contract.get("product_row_merges", []):
        start_col = get_column_letter(merge["start_col"])
        end_col = get_column_letter(merge["end_col"])
        for row in range(product_start_row, product_start_row + product_count):
            checks += 1
            merge_ref = f"{start_col}{row}:{end_col}{row}"
            if merge_ref not in actual_merges:
                errors.append(f"Missing product-row merge: {merge_ref}")

    summary_start = product_start_row + product_count
    for merge in template_contract.get("summary_relative_merges", []):
        row = summary_start + merge["row_offset"]
        start_col = get_column_letter(merge["start_col"])
        end_col = get_column_letter(merge["end_col"])
        checks += 1
        merge_ref = f"{start_col}{row}:{end_col}{row}"
        if merge_ref not in actual_merges:
            errors.append(f"Missing summary merge: {merge_ref}")

    for label in template_contract.get("summary_static_labels", []):
        row = summary_start + label["row_offset"]
        cell_ref = f"{label['col']}{row}"
        checks += 1
        actual = ws[cell_ref].value
        if actual != label["value"]:
            errors.append(f"Summary label {cell_ref}: expected '{label['value']}', got '{actual}'")

    for anchor in template_contract.get("formula_anchors", []):
        row = summary_start + anchor["row_offset"]
        cell_ref = f"{anchor['col']}{row}"
        checks += 1
        actual = ws[cell_ref].value
        if actual is None or not str(actual).startswith("="):
            errors.append(f"Formula anchor {cell_ref}: expected formula, got '{actual}'")

    return errors, checks


def _find_header_anchor(ws, target_cell_ref: str, header_field_cells: set[str]) -> dict[str, Any] | None:
    col = "".join(c for c in target_cell_ref if c.isalpha())
    row = int("".join(c for c in target_cell_ref if c.isdigit()))
    col_idx = _column_sort_key(col)

    candidates = [
        (row, col_idx - 1),
        (row, col_idx - 2),
        (row - 1, col_idx),
    ]
    for candidate_row, candidate_col in candidates:
        if candidate_row < 1 or candidate_col < 1:
            continue
        anchor_cell = f"{get_column_letter(candidate_col)}{candidate_row}"
        if anchor_cell in header_field_cells:
            continue
        value = ws[anchor_cell].value
        if value is None:
            continue
        if isinstance(value, str) and not value.strip():
            continue
        return {
            "target_cell": target_cell_ref,
            "anchor_cell": anchor_cell,
            "value": value,
        }
    return None


def _column_sort_key(col_letter: str) -> int:
    key = 0
    for ch in col_letter:
        key = key * 26 + (ord(ch.upper()) - ord("A") + 1)
    return key
