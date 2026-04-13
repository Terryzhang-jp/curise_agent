"""Template engine v2: compose-from-scratch xlsx renderer.

Why this exists
===============
The original `template_engine.fill_template` mutates an openpyxl Workbook in
place: it calls `insert_rows`, `delete_rows`, and tries to clean up the
mess (stale merge ranges, lost cell formats). openpyxl's row manipulation
has known footguns — merge ranges don't auto-shift, cell formats can be
silently reset to General, and the cleanup code in fill_template is a long
list of workarounds that has accumulated bugs over time. On 2026-04-11 we
found 5 distinct production bugs in fill_template in a single day.

This module takes the opposite approach: **compose, never mutate**.

  1. Open the template as a READ-ONLY reference
  2. Build a NEW Workbook from scratch
  3. Walk the template's rows top-to-bottom and emit each cell at its
     final position in the destination, computing row offsets explicitly
  4. Re-create merges, formulas, and styles in the destination
  5. Save

Because we never touch openpyxl's row manipulation APIs, every footgun
that bit fill_template is structurally impossible here. The trade-off is
that we do more work per cell (explicit copies), but the time difference
on real templates is well under 1 second.

This module is the production version of the POC at
`tests/_poc_compose_renderer.py`. Both should produce byte-equivalent
output (modulo cell ordering); the POC is preserved as a reference.

Public API
----------
    compose_render(template_bytes, zone_config_dict, order_data, supplier_id) -> bytes

Validation
----------
The `zone_config_dict` is validated against `ZoneConfigV1` before use.
Any malformed config raises `ZoneConfigValidationError` with field paths.
"""

from __future__ import annotations

import copy
import io
import logging
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import column_index_from_string, get_column_letter

from services.templates.zone_config_schema import (
    ZoneConfigV1,
    ZoneConfigValidationError,
    parse_zone_config,
)

logger = logging.getLogger(__name__)


# Integer-only number formats that should be promoted to optional-decimal
# when the data is fractional. `#,##0.##` is the universal "show up to 2
# decimals only when needed" format — backward-compatible with integer data.
_INTEGER_ONLY_FORMATS = {
    "#,##0",
    "0",
    "General",
    "#,##0;-#,##0",
    "#,##0_ ;-#,##0_ ",
}


# OWASP CSV/Excel injection prevention: strings starting with these characters
# are treated as formulas/DDE commands by Excel on open. To block injection
# via supplier name, product name, address fields, etc., we prepend a single
# quote (Excel's text-literal marker) to any user-supplied string whose first
# char is one of these AND which is not a legitimate number.
_INJECTION_PREFIXES = ("=", "+", "-", "@", "\t", "\r")


def _sanitize_cell_input(value: Any) -> Any:
    """Escape potentially-dangerous strings to prevent CSV/Excel formula injection.

    Rules:
      - Non-strings pass through unchanged (numbers, dates, bools)
      - Empty strings pass through unchanged
      - Strings starting with dangerous prefixes (= + - @ \\t \\r):
          * If it parses as a number (e.g. "-50", "+1.5"), leave as-is —
            Excel will display it as a number
          * Otherwise prepend a single quote (') which Excel uses as a text-
            literal marker. The quote is stripped from the display in Excel
            but prevents formula evaluation.

    This is applied to every value written from user-controlled sources
    (header_fields resolved from order_data, product row cells resolved from
    products). It is NOT applied to intentional formulas written from the
    zone_config (product_row_formulas, summary_formulas, external_refs) —
    those must remain executable.
    """
    if not isinstance(value, str) or not value:
        return value
    if value[0] not in _INJECTION_PREFIXES:
        return value
    try:
        float(value)
        return value  # legitimate numeric string like "-50" or "+1.5"
    except (ValueError, TypeError):
        pass
    return "'" + value


# ──────────────────────────────────────────────────────────────────────
# Public API
# ──────────────────────────────────────────────────────────────────────


def compose_render(
    template_bytes: bytes,
    zone_config_dict: dict[str, Any],
    order_data: dict[str, Any],
    supplier_id: str | int,
    field_overrides: dict[str, str] | None = None,
) -> bytes:
    """Render a filled xlsx by composing from a template, never mutating it.

    Args:
        template_bytes: Raw bytes of the .xlsx template file.
        zone_config_dict: A dict matching `ZoneConfigV1` shape. Will be
            validated; ZoneConfigValidationError raised on bad input.
        order_data: The order's data, with the standard structure used by
            inquiry_agent. Must contain `suppliers[str(supplier_id)].products`.
        supplier_id: The supplier this inquiry is for.

    Returns:
        The filled xlsx file as bytes.

    Raises:
        ZoneConfigValidationError: zone_config_dict failed schema validation.
        ValueError: order_data is missing required structure (e.g. no products).
        Exception: openpyxl-level failures (rare, indicates corrupt template).
    """
    # ── Validate zone_config first — fail fast on bad config ──
    config: ZoneConfigV1 = parse_zone_config(zone_config_dict)

    sid = str(supplier_id)
    supplier_data = (order_data.get("suppliers") or {}).get(sid)
    if not supplier_data:
        raise ValueError(f"order_data has no supplier {sid}")
    products = supplier_data.get("products") or []
    if not products:
        raise ValueError(f"supplier {sid} has no products in order_data")

    n_products = len(products)

    # ── Load source template (read-only reference) ──
    src_wb = load_workbook(io.BytesIO(template_bytes))
    src_ws = src_wb.active

    # ── Create blank destination workbook ──
    dst_wb = Workbook()
    dst_ws = dst_wb.active
    dst_ws.title = src_ws.title

    # ── Zone definitions ──
    prod_start = config.zones.product_data.start
    prod_end = config.zones.product_data.end
    summ_start = config.zones.summary.start
    summ_end = config.zones.summary.end
    template_prod_rows = prod_end - prod_start + 1
    row_delta = n_products - template_prod_rows
    new_prod_end = prod_start + n_products - 1
    new_summ_start = prod_end + 1 + row_delta
    new_summ_end = new_summ_start + (summ_end - summ_start)

    # ── Build flat order context for both header_fields and product cells ──
    order_context = {
        "ship_name": order_data.get("ship_name", ""),
        "delivery_date": order_data.get("delivery_date", ""),
        "order_date": order_data.get("order_date", ""),
        "destination_port": order_data.get("destination_port", ""),
        "voyage": order_data.get("voyage", ""),
        "supplier_name": supplier_data.get("supplier_name", ""),
        "po_number": order_data.get("po_number", ""),
        "currency": order_data.get("currency", ""),
    }

    # ── Pre-scan: does this order contain ANY fractional numeric value? ──
    # If so, every integer-only format in the product zone (including formula
    # cells whose value we can't see at render time) gets promoted to
    # `#,##0.##`. This makes display precision consistent across the whole
    # filled workbook regardless of which cells happen to be integer.
    any_fractional = _scan_for_fractional_values(products)

    # ── Phase 1: copy header rows (rows 1 to prod_start - 1) ──
    for src_row in range(1, prod_start):
        _copy_row_cells(src_ws, dst_ws, src_row, src_row)

    # ── Phase 1b: apply header_fields (placeholder substitution) ──
    # If a mapped source resolves to None (order_data missing that field),
    # we EXPLICITLY clear the cell — otherwise the template's sample value
    # would survive Phase 1 copy and leak into the output (e.g. a stale
    # invoice number, stale delivery address, etc).
    for cell_ref, data_path in config.header_fields.items():
        col_letter, row_num = _split_cell_ref(cell_ref)
        if row_num >= prod_start:
            continue  # not in header zone
        col_idx = column_index_from_string(col_letter)
        cell = dst_ws.cell(row=row_num, column=col_idx)
        if isinstance(cell, MergedCell):
            continue
        value = _resolve_data_path(order_data, data_path, sid)
        value = _sanitize_cell_input(value)   # block formula injection
        cell.value = value  # explicit None clears any copied template sample
        if value is not None:
            cell.number_format = _format_for_value(
                cell.number_format, value, force_decimal=any_fractional
            )

    # ── Phase 1c: apply user field_overrides (highest priority) ──
    if field_overrides:
        for cell_ref, value in field_overrides.items():
            col_letter, row_num = _split_cell_ref(cell_ref)
            if row_num >= prod_start:
                continue  # only header zone
            col_idx = column_index_from_string(col_letter)
            cell = dst_ws.cell(row=row_num, column=col_idx)
            if isinstance(cell, MergedCell):
                continue
            cell.value = _sanitize_cell_input(value)

    # ── Phase 2: emit N product rows from the template's first product row ──
    template_row = prod_start
    po_number = order_data.get("po_number", "")
    currency = order_data.get("currency", "")

    for i, product in enumerate(products):
        dst_row = prod_start + i
        # Copy the template product row's cell formats
        _copy_row_cells(src_ws, dst_ws, template_row, dst_row)
        # Override values with product data (with order_context fallback).
        # If the field name is unrecognized, _resolve_product_field returns the
        # _USE_TEMPLATE sentinel — we preserve the template row's original value
        # in that cell (covers static per-row labels like "JPY" / "Submitted").
        for col_letter, field_name in config.product_columns.items():
            col_idx = column_index_from_string(col_letter)
            value = _resolve_product_field(
                field_name, product, i, po_number, currency, order_context
            )
            if value is _USE_TEMPLATE:
                continue  # leave the template's copied value in place
            cell = dst_ws.cell(row=dst_row, column=col_idx)
            if isinstance(cell, MergedCell):
                continue
            value = _sanitize_cell_input(value)   # block formula injection
            cell.value = value
            cell.number_format = _format_for_value(
                cell.number_format, value, force_decimal=any_fractional
            )
        # Apply per-row formulas — formula cells don't have a value at render
        # time, so we promote unconditionally if any product data is fractional
        for col_letter, formula_tpl in config.product_row_formulas.items():
            col_idx = column_index_from_string(col_letter)
            cell = dst_ws.cell(row=dst_row, column=col_idx)
            if isinstance(cell, MergedCell):
                continue
            cell.value = formula_tpl.replace("{row}", str(dst_row))
            if any_fractional:
                cell.number_format = _promote_int_format(cell.number_format)

    # ── Phase 3: copy rows below the product zone (summary + footer) ──
    # Each row gets shifted by row_delta. Formulas in summary cells are
    # NOT auto-adjusted from the template — instead we re-emit them from
    # `summary_formulas` in zone_config (the explicit declarative source
    # of truth). This avoids fragile regex-based formula parsing.
    for src_row in range(prod_end + 1, src_ws.max_row + 1):
        dst_row = src_row + row_delta
        _copy_row_cells(src_ws, dst_ws, src_row, dst_row)

    # ── Phase 3a: clear stale columns inside the summary zone ──
    for row in range(new_summ_start, new_summ_end + 1):
        for col_letter in config.stale_columns_in_summary:
            col_idx = column_index_from_string(col_letter)
            cell = dst_ws.cell(row=row, column=col_idx)
            if not isinstance(cell, MergedCell):
                cell.value = None

    # ── Phase 3b: restore summary static values at new positions ──
    for cell_ref, value in config.summary_static_values.items():
        col_letter, orig_row = _split_cell_ref(cell_ref)
        offset = orig_row - summ_start
        new_row = new_summ_start + offset
        col_idx = column_index_from_string(col_letter)
        cell = dst_ws.cell(row=new_row, column=col_idx)
        if not isinstance(cell, MergedCell):
            cell.value = value

    # ── Phase 3c: emit summary formulas at new positions ──
    formula_cell_refs: dict[str, str] = {}
    for sf in config.summary_formulas:
        col_letter, orig_row = _split_cell_ref(sf.cell)
        new_row = orig_row + row_delta
        col_idx = column_index_from_string(col_letter)

        if sf.type == "product_sum":
            formula = f"=SUM({col_letter}{prod_start}:{col_letter}{new_prod_end})"
            formula_cell_refs["sum_cell"] = f"{col_letter}{new_row}"
        elif sf.type == "relative":
            formula = sf.formula_template or ""
            for key, ref in formula_cell_refs.items():
                formula = formula.replace(f"{{{key}}}", ref)
            label = (sf.label or "").lower()
            if "tax" in label:
                formula_cell_refs["tax_cell"] = f"{col_letter}{new_row}"
            if "grand" in label or "total" in label:
                formula_cell_refs["grand_total_cell"] = f"{col_letter}{new_row}"
        else:
            continue

        cell = dst_ws.cell(row=new_row, column=col_idx)
        if isinstance(cell, MergedCell):
            continue
        cell.value = formula
        if any_fractional:
            cell.number_format = _promote_int_format(cell.number_format)

    # ── Phase 3d: external cross-references ──
    # zone_config["external_refs"] declares header (or other) cells whose
    # formulas reference summary cells (typically grand_total) by name.
    # Template authors use these to display the grand total in the header.
    # After row resizing, the underlying grand total cell has moved, so we
    # must rewrite these formulas with the NEW addresses.
    for ext in config.external_refs:
        col_letter, row_num = _split_cell_ref(ext.cell)
        col_idx = column_index_from_string(col_letter)
        formula = ext.formula_template
        for key, ref in formula_cell_refs.items():
            formula = formula.replace(f"{{{key}}}", ref)
        cell = dst_ws.cell(row=row_num, column=col_idx)
        if isinstance(cell, MergedCell):
            continue
        cell.value = formula
        if any_fractional:
            cell.number_format = _promote_int_format(cell.number_format)

    # ── Phase 4: copy column widths (visual fidelity) ──
    for col_letter, dim in src_ws.column_dimensions.items():
        if dim.width:
            dst_ws.column_dimensions[col_letter].width = dim.width

    # ── Phase 4b: copy row heights, shifting rows below the product zone ──
    # Rows in the product zone are replicated (all N products get the
    # template product row's height). Rows below prod_end shift by row_delta.
    template_prod_height = src_ws.row_dimensions[prod_start].height if prod_start in src_ws.row_dimensions else None
    for src_row_num, dim in src_ws.row_dimensions.items():
        if dim.height is None:
            continue
        if src_row_num < prod_start:
            dst_ws.row_dimensions[src_row_num].height = dim.height
        elif src_row_num > prod_end:
            dst_ws.row_dimensions[src_row_num + row_delta].height = dim.height
        # product-zone rows handled below
    if template_prod_height is not None:
        for i in range(n_products):
            dst_ws.row_dimensions[prod_start + i].height = template_prod_height

    # ── Phase 4c: copy sheet-level layout properties ──
    # These control printing, zoom, and default row metrics. Losing them
    # breaks PDF export and visual parity with the template.
    try:
        dst_ws.page_margins = copy.copy(src_ws.page_margins)
        dst_ws.page_setup = copy.copy(src_ws.page_setup)
        dst_ws.print_options = copy.copy(src_ws.print_options)
        dst_ws.sheet_format = copy.copy(src_ws.sheet_format)
        dst_ws.sheet_properties = copy.copy(src_ws.sheet_properties)
        # sheet_view has no setter — mutate in place via the sheet_views list
        if src_ws.sheet_view is not None and dst_ws.views.sheetView:
            dst_ws.views.sheetView[0] = copy.copy(src_ws.sheet_view)
        # Freeze panes (simple string like "A5") — safe to assign directly
        if src_ws.freeze_panes:
            dst_ws.freeze_panes = src_ws.freeze_panes
        # Page header/footer
        dst_ws.oddHeader = copy.copy(src_ws.oddHeader)
        dst_ws.oddFooter = copy.copy(src_ws.oddFooter)
    except Exception as exc:  # pragma: no cover
        logger.warning("Failed to copy sheet layout properties: %s", exc)

    # ── Phase 4d: shift print area / print titles / auto filter by row_delta ──
    # These reference absolute rows in the template; if the summary moved,
    # they need to move too, or printing cuts off the wrong region.
    if src_ws.print_area:
        dst_ws.print_area = _shift_range_rows(src_ws.print_area, prod_end, row_delta)
    if src_ws.print_title_rows:
        # Print title rows are usually header rows (above prod_start) — no shift needed
        dst_ws.print_title_rows = src_ws.print_title_rows
    if src_ws.auto_filter and src_ws.auto_filter.ref:
        dst_ws.auto_filter.ref = _shift_range_rows(
            src_ws.auto_filter.ref, prod_end, row_delta
        )

    # ── Phase 5: re-create merges, adjusting positions ──
    for merged in list(src_ws.merged_cells.ranges):
        if merged.max_row < prod_start:
            # Header merge: same position
            try:
                dst_ws.merge_cells(str(merged))
            except Exception as exc:  # pragma: no cover
                logger.warning("Failed to copy header merge %s: %s", merged, exc)

        elif merged.min_row >= prod_start and merged.max_row <= prod_end:
            # Product-zone merge: replicate per product row
            row_offset_in_zone = merged.min_row - prod_start
            for i in range(n_products):
                dst_row = prod_start + i + row_offset_in_zone
                start_letter = get_column_letter(merged.min_col)
                end_letter = get_column_letter(merged.max_col)
                try:
                    dst_ws.merge_cells(f"{start_letter}{dst_row}:{end_letter}{dst_row}")
                except Exception:  # pragma: no cover
                    pass

        elif merged.min_row > prod_end:
            # Summary or footer merge: shift down by row_delta
            new_min = merged.min_row + row_delta
            new_max = merged.max_row + row_delta
            start_letter = get_column_letter(merged.min_col)
            end_letter = get_column_letter(merged.max_col)
            try:
                dst_ws.merge_cells(f"{start_letter}{new_min}:{end_letter}{new_max}")
            except Exception as exc:  # pragma: no cover
                logger.warning("Failed to copy summary merge %s → %s%d:%s%d: %s",
                               merged, start_letter, new_min, end_letter, new_max, exc)
        else:
            # Boundary-straddling merge — rare, skip with warning
            logger.warning("Skipped boundary-straddling merge: %s", merged)

    # ── Save ──
    buf = io.BytesIO()
    dst_wb.save(buf)
    return buf.getvalue()


# ──────────────────────────────────────────────────────────────────────
# Helpers (private)
# ──────────────────────────────────────────────────────────────────────


def _split_cell_ref(cell_ref: str) -> tuple[str, int]:
    """Split 'B2' into ('B', 2)."""
    col_letter = "".join(c for c in cell_ref if c.isalpha())
    row_num = int("".join(c for c in cell_ref if c.isdigit()))
    return col_letter, row_num


def _shift_range_rows(range_ref: str, pivot_row: int, row_delta: int) -> str:
    """Shift rows > pivot_row in an Excel range reference by row_delta.

    Handles refs like 'A1:K30', '$A$1:$K$30', "'Sheet'!$A$1:$K$30".
    Rows <= pivot_row stay put (they're in the header/product zone).
    """
    import re
    if not range_ref or row_delta == 0:
        return range_ref

    def _shift_one(match: re.Match) -> str:
        col = match.group(1)
        row = int(match.group(2))
        if row > pivot_row:
            row += row_delta
        return f"{col}{row}"

    # Shift each cell ref (letters+digits) in the string
    return re.sub(r"(\$?[A-Z]+\$?)(\d+)", _shift_one, range_ref)


def _scan_for_fractional_values(products: list[dict]) -> bool:
    """Return True if any product has a non-integer numeric field."""
    for product in products:
        for v in product.values():
            if isinstance(v, float) and v != int(v):
                return True
    return False


def _copy_row_cells(src_ws, dst_ws, src_row: int, dst_row: int) -> None:
    """Copy all cells in src_row to dst_row, including value + style."""
    max_col = src_ws.max_column or 1
    for col_idx in range(1, max_col + 1):
        src_cell = src_ws.cell(row=src_row, column=col_idx)
        if isinstance(src_cell, MergedCell):
            continue
        dst_cell = dst_ws.cell(row=dst_row, column=col_idx)
        dst_cell.value = src_cell.value
        if src_cell.has_style:
            dst_cell.font = copy.copy(src_cell.font)
            dst_cell.fill = copy.copy(src_cell.fill)
            dst_cell.border = copy.copy(src_cell.border)
            dst_cell.alignment = copy.copy(src_cell.alignment)
            dst_cell.number_format = src_cell.number_format
            dst_cell.protection = copy.copy(src_cell.protection)


def _promote_int_format(template_fmt: str) -> str:
    """Force-promote an integer-only format to optional-decimal."""
    if template_fmt in _INTEGER_ONLY_FORMATS:
        return "#,##0.##"
    return template_fmt


def _format_for_value(template_fmt: str, value: Any, force_decimal: bool = False) -> str:
    """Promote integer-only number formats to optional-decimal when needed.

    Strings (non-formula) never get promoted — strings ignore number_format
    in Excel anyway. Formula strings (start with "=") are treated as numeric
    because Excel evaluates them.
    """
    if isinstance(value, str) and not value.startswith("="):
        return template_fmt
    if force_decimal:
        return _promote_int_format(template_fmt)
    if not isinstance(value, (int, float)):
        return template_fmt
    if value == int(value):
        return template_fmt
    return _promote_int_format(template_fmt)


def _resolve_data_path(order_data: dict, path: str, sid: str) -> Any:
    """Resolve a dotted data path like 'suppliers.{sid}.supplier_name'."""
    if not path:
        return None
    expanded = path.replace("{sid}", sid)
    cursor: Any = order_data
    for part in expanded.split("."):
        if isinstance(cursor, dict):
            cursor = cursor.get(part)
        else:
            return None
        if cursor is None:
            return None
    return cursor


# Sentinel: means "this field name is unknown — keep whatever the template had".
# Returned by _resolve_product_field so the caller can distinguish
# "known field, value is empty" (must clear cell) vs "unknown field name"
# (should preserve the template's per-row static value like "JPY" or "Submitted").
_USE_TEMPLATE = object()


def _resolve_product_field(
    field_name: str,
    product: dict,
    index: int,
    po_number: str,
    currency: str,
    order_context: dict | None = None,
) -> Any:
    """Resolve a product cell value, falling back to order-level context.

    Mirrors services/template_engine.py::_resolve_product_field so that flat-
    table templates (which repeat ship_name / supplier_name / delivery_date
    on every product row) work correctly.

    Returns `_USE_TEMPLATE` sentinel if the field name is completely
    unrecognized AND no value is found in product / order_context. The
    caller should treat this as "preserve the template row's original value".
    """
    if field_name in ("line_number", "__line_number__"):
        return index + 1
    if field_name in ("po_number", "__po_number__"):
        return po_number
    if field_name in ("currency", "__currency__"):
        return product.get("currency") or currency

    if field_name == "product_code":
        return product.get("product_code", "")
    if field_name in ("product_name", "product_name_en"):
        return product.get("product_name", "")
    if field_name == "product_name_jp":
        return product.get("product_name_jp", "")
    if field_name == "description":
        return product.get("pack_size", "")
    if field_name == "quantity":
        return product.get("quantity")
    if field_name == "unit_price":
        return product.get("unit_price")
    if field_name == "unit":
        return product.get("unit", "CT")

    # Try product dict first
    val = product.get(field_name)
    if val is not None and val != "":
        return val

    # Fallback to order-level context for repeated fields
    if order_context:
        val = order_context.get(field_name)
        if val is not None:
            return val

    # Unknown field and no value anywhere — let the template's copied value stand.
    # This covers per-row static labels like "JPY" (currency_label) or
    # "Submitted" (order_status) that the template author put in every row
    # but didn't want parameterized.
    return _USE_TEMPLATE


# Re-export for convenient imports
__all__ = ["compose_render", "ZoneConfigValidationError"]
