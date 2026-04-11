"""Excel template structure analysis using Gemini AI.

Analyzes an uploaded Excel template to discover field positions and product table layout,
producing a configuration that can be saved to SupplierTemplate for deterministic filling.
"""

import io
import logging
import re
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from services.documents.pdf_analyzer import _get_model, _parse_json_response

logger = logging.getLogger(__name__)

ANALYSIS_PROMPT = """你是 Excel 询价单/采购单模板分析专家。以下是一个 Excel 模板的所有非空单元格内容。
请分析模板结构，找出：

1. **头部字段位置**：哪些单元格是需要填入数据的位置（不是标签本身，而是标签旁边的值单元格）
2. **产品表格配置**：产品明细表的表头行、数据起始行、每列对应的字段

## Excel 单元格内容
{cell_text}

## 标准化字段名映射

头部字段（找到对应标签后，返回**值所在的单元格位置**，不是标签位置）：
- po_number: PO号/注文番号/Purchase Order No
- order_date: 下单日期/注文日/Order Date
- delivery_date: 交货日期/纳品日/Delivery Date
- ship_name: 船名/Ship Name/Vessel
- voyage: 航次号/Voyage No
- destination: 目的地/Destination
- port_name: 港口/Port
- supplier_name: 供应商名/Vendor/Supplier
- vendor_name: 供应商名（同 supplier_name）
- invoice: 发票号/Invoice No
- currency: 币种/Currency
- total_amount: 合计/Total（通常是公式，标注为公式列）
- payment_date: 付款日期/Payment Date
- payment_method: 付款方式/Payment Method
- contact_person: 联系人/Contact
- delivery_address: 交货地址/Delivery Address

产品表列：
- line_number: 行号/No./Item No
- po_number: PO号（每行重复）
- product_code: 商品代码/Product Code/Item Code
- product_name_en: 英文名/Product Name/Description
- product_name_jp: 日文名/品名
- description: 规格/包装/Pack Size/Specification
- quantity: 数量/Qty/Quantity
- unit: 单位/Unit/UOM
- unit_price: 单价/Unit Price/Price
- currency: 币种/Currency
- total_price: 金额/Amount/Total（通常是公式）

## 输出要求

返回纯 JSON（不要 markdown 代码块）：
{{
  "field_positions": {{
    "字段名": "单元格位置（如 B3）",
    ...
  }},
  "product_table_config": {{
    "header_row": 表头所在行号,
    "start_row": 数据起始行号（表头下一行）,
    "columns": {{
      "列字母": "字段名",
      ...
    }},
    "formula_columns": ["公式列的列字母，如 L"]
  }},
  "notes": "特殊备注（合并单元格、税率行等）"
}}

注意：
- field_positions 中的位置是**值应该填入的单元格**，不是标签单元格
- 如果标签在 A3，值通常在 B3 或 C3
- formula_columns 中列出包含公式的列（如金额列 = 数量 * 单价），这些列填充时应跳过
- 只返回你确定识别到的字段，不确定的不要返回
"""


def _build_cell_text(wb) -> str:
    """Build a text representation of all non-empty cells in the workbook."""
    lines = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if len(wb.sheetnames) > 1:
            lines.append(f"=== Sheet: {sheet_name} ===")
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if cell.value is not None:
                    col_letter = get_column_letter(cell.column)
                    pos = f"{col_letter}{cell.row}"
                    # Mark formula cells
                    val = cell.value
                    if isinstance(val, str) and val.startswith("="):
                        lines.append(f"{pos}: [FORMULA] {val}")
                    else:
                        lines.append(f"{pos}: {val}")
    return "\n".join(lines)


def analyze_excel_template(file_bytes: bytes) -> dict[str, Any]:
    """Analyze an Excel template to discover its structure and field positions.

    Returns:
        {
            "field_positions": {"po_number": "B3", ...},
            "product_table_config": {
                "header_row": 11,
                "start_row": 12,
                "columns": {"A": "line_number", "B": "product_code", ...},
                "formula_columns": ["L"]
            },
            "notes": "..."
        }
    """
    wb = load_workbook(io.BytesIO(file_bytes), data_only=False)
    cell_text = _build_cell_text(wb)

    if not cell_text.strip():
        return {
            "field_positions": {},
            "product_table_config": {},
            "notes": "Empty workbook",
        }

    logger.info("Analyzing Excel template (%d chars of cell text)", len(cell_text))

    model = _get_model()
    prompt = ANALYSIS_PROMPT.format(cell_text=cell_text[:15000])

    response = model.generate_content([prompt])
    response_text = response.text.strip()
    logger.info("Gemini response length: %d chars", len(response_text))

    result = _parse_json_response(response_text)

    result.setdefault("field_positions", {})
    result.setdefault("product_table_config", {})
    result.setdefault("notes", "")

    fp_count = len(result["field_positions"])
    col_count = len(result.get("product_table_config", {}).get("columns", {}))
    logger.info("Template analysis complete: %d field positions, %d product columns", fp_count, col_count)

    return result


# ── Enhanced analysis with order context ────────────────────────

ENHANCED_ANALYSIS_PROMPT = """你是 Excel 询价单/采购单模板分析专家。以下是一个 Excel 模板的所有非空单元格内容。
请分析模板结构，**重点将订单模板提供的字段与模板中的单元格位置进行匹配**。

## Excel 单元格内容
{cell_text}

## 订单模板提供的字段

### 头部字段（需要匹配到模板中的具体单元格）
{header_fields_text}

### 产品表列字段
{product_fields_text}

{source_company_text}

## 任务

1. **头部字段匹配**：对上面列出的每个头部字段，找到模板中对应的**值单元格位置**（不是标签位置）
   - 如果标签在 A3，值通常在 B3 或 C3
   - 返回 field_positions 映射
   - 对每个匹配的字段，也返回 field_mapping_preview 条目（含 current_value）

2. **产品表格配置**：找到产品明细表的表头行、数据起始行、每列对应的字段

3. **field_mapping_preview**：为每个订单头部字段生成一条预览记录

## 输出要求

返回纯 JSON（不要 markdown 代码块）：
{{
  "field_positions": {{
    "字段key": "单元格位置（如 B3）",
    ...
  }},
  "product_table_config": {{
    "header_row": 表头所在行号,
    "start_row": 数据起始行号（表头下一行）,
    "columns": {{
      "列字母": "字段key",
      ...
    }},
    "formula_columns": ["公式列的列字母，如 L"]
  }},
  "field_mapping_preview": [
    {{
      "order_field_key": "字段key",
      "order_field_label": "字段中文标签",
      "matched_position": "匹配到的单元格位置（如 H8）或 null",
      "current_cell_value": "该单元格当前的值（字符串）或 null",
      "confidence": "high/medium/low",
      "source": "ai"
    }},
    ...
  ],
  "notes": "特殊备注"
}}

注意：
- field_positions 中的位置是**值应该填入的单元格**，不是标签单元格
- formula_columns 中列出包含公式的列
- field_mapping_preview 必须为每个订单头部字段都生成一条记录，即使未匹配到也要返回（matched_position 设为 null）
- current_cell_value 是模板中该单元格当前已有的值
- 只返回你确定识别到的匹配，不确定的 matched_position 设为 null
"""


def _build_order_context_text(order_context: dict) -> tuple[str, str, str]:
    """Build text sections from order context for the enhanced prompt."""
    # Header fields
    header_fields = order_context.get("header_fields", [])
    if header_fields:
        lines = []
        for f in header_fields:
            lines.append(f"- {f['key']}: {f['label']}")
        header_text = "\n".join(lines)
    else:
        header_text = "（无头部字段）"

    # Product fields
    product_fields = order_context.get("product_fields", [])
    if product_fields:
        lines = []
        for f in product_fields:
            lines.append(f"- {f['key']}: {f['label']}")
        product_text = "\n".join(lines)
    else:
        product_text = "（无产品列字段）"

    # Source company
    source_company = order_context.get("source_company")
    if source_company:
        company_text = f"## 来源公司\n此订单来自: {source_company}"
    else:
        company_text = ""

    return header_text, product_text, company_text


def analyze_excel_template_with_order_context(
    file_bytes: bytes,
    order_context: dict,
) -> dict[str, Any]:
    """Analyze an Excel template with order context for targeted field matching.

    Args:
        file_bytes: The Excel file bytes
        order_context: {
            "header_fields": [{"key": str, "label": str}, ...],
            "product_fields": [{"key": str, "label": str}, ...],
            "source_company": str | None
        }

    Returns:
        Same structure as analyze_excel_template() plus:
        - field_mapping_preview: list of per-field matching results
    """
    wb = load_workbook(io.BytesIO(file_bytes), data_only=False)
    cell_text = _build_cell_text(wb)

    if not cell_text.strip():
        return {
            "field_positions": {},
            "product_table_config": {},
            "field_mapping_preview": [],
            "notes": "Empty workbook",
        }

    header_text, product_text, company_text = _build_order_context_text(order_context)

    logger.info(
        "Analyzing Excel template with order context (%d chars, %d header fields, %d product fields)",
        len(cell_text),
        len(order_context.get("header_fields", [])),
        len(order_context.get("product_fields", [])),
    )

    model = _get_model()
    prompt = ENHANCED_ANALYSIS_PROMPT.format(
        cell_text=cell_text[:15000],
        header_fields_text=header_text,
        product_fields_text=product_text,
        source_company_text=company_text,
    )

    response = model.generate_content([prompt])
    response_text = response.text.strip()
    logger.info("Gemini response length: %d chars", len(response_text))

    result = _parse_json_response(response_text)

    result.setdefault("field_positions", {})
    result.setdefault("product_table_config", {})
    result.setdefault("field_mapping_preview", [])
    result.setdefault("notes", "")

    fp_count = len(result["field_positions"])
    preview_count = len(result["field_mapping_preview"])
    logger.info(
        "Enhanced analysis complete: %d field positions, %d preview items",
        fp_count,
        preview_count,
    )

    return result


# ── HTML preview generation ─────────────────────────────────────


def generate_template_html(file_bytes: bytes, sheet: int = 0) -> str:
    """Convert an Excel file to an HTML table string using xlsx2html.

    The returned HTML contains only the <table>...</table> portion with
    ``data-cell-ref`` attributes added to each <td> for frontend targeting.

    Args:
        file_bytes: Raw bytes of the .xlsx file.
        sheet: 0-based sheet index (default first sheet).

    Returns:
        HTML string with a single <table> element.
    """
    from xlsx2html import xlsx2html as _xlsx2html

    src = io.BytesIO(file_bytes)
    dest = io.StringIO()
    _xlsx2html(src, dest, sheet=sheet)
    raw_html = dest.getvalue()

    # Extract just the <table>…</table> portion
    m = re.search(r"(<table[\s\S]*?</table>)", raw_html, re.IGNORECASE)
    if not m:
        logger.warning("xlsx2html produced no <table> element")
        return ""
    table_html = m.group(1)

    # Add data-cell-ref="A4" to every <td id="...!A4"> or <td id="A4">
    def _add_data_ref(match: re.Match) -> str:
        full = match.group(0)
        td_id = match.group(1)
        # id may be "Sheet1!A4" or just "A4"
        bare_ref = td_id.split("!")[-1] if "!" in td_id else td_id
        if "data-cell-ref" not in full:
            return full.replace(">", f' data-cell-ref="{bare_ref}">', 1)
        return full

    table_html = re.sub(
        r'<td[^>]*\bid=["\']([^"\']+)["\'][^>]*>',
        _add_data_ref,
        table_html,
    )

    logger.info("Generated template HTML (%d chars)", len(table_html))
    return table_html
