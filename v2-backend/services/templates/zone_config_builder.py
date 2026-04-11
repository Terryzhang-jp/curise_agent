"""Auto-generate zone config from template analysis + openpyxl scan.

Takes the AI-produced field_positions + product_table_config and enriches
with structural information extracted deterministically from the template file:
  - Product zone boundaries (start/end rows)
  - Summary zone (SUM, tax, grand total formulas)
  - Cross-zone formula references
  - Stale data columns in summary rows
  - Header field → order_data path mapping

The output is a complete zone_config ready for template_engine.fill_template().
"""

from __future__ import annotations

import io
import logging
import re
from collections import defaultdict
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter, column_index_from_string

logger = logging.getLogger(__name__)


# ── Field key → order_data path mapping ──────────────────────────
# This is a fixed business-logic lookup: field_key (from AI analysis)
# maps to a dotted path in order_data JSON.
# {sid} is replaced at runtime with the actual supplier_id.

FIELD_DATA_PATHS: dict[str, str] = {
    # Order fields
    "ship_name": "ship_name",
    "ship_name_alt": "ship_name",
    "po_number": "po_number",
    "order_date": "order_date",
    "delivery_date": "delivery_date",
    "delivery_address": "delivery_address",
    "delivery_contact": "delivery_location.contact_person",
    "delivery_time_notes": "delivery_location.delivery_notes",
    "destination": "destination_port",
    "destination_port": "destination_port",
    "voyage": "voyage",
    "invoice_number": "po_number",
    "currency": "currency",
    "payment_date": "suppliers.{sid}.supplier_info.default_payment_terms",
    "payment_method": "suppliers.{sid}.supplier_info.default_payment_method",
    # Supplier fields
    "supplier_name": "suppliers.{sid}.supplier_name",
    "supplier_contact": "suppliers.{sid}.supplier_info.contact",
    "supplier_tel": "suppliers.{sid}.supplier_info.phone",
    "supplier_fax": "suppliers.{sid}.supplier_info.fax",
    "supplier_email": "suppliers.{sid}.supplier_info.email",
    "supplier_address": "suppliers.{sid}.supplier_info.address",
    "supplier_zip_code": "suppliers.{sid}.supplier_info.zip_code",
    "supplier_bank": "suppliers.{sid}.supplier_info.bank_info",
    "supplier_account": "suppliers.{sid}.supplier_info.account_info",
}

# Product column → order_data product field mapping
PRODUCT_FIELD_MAP: dict[str, str] = {
    "line_number": "__line_number__",
    "po_number": "__po_number__",
    "product_code": "product_code",
    "product_name": "product_name",
    "product_name_en": "product_name",
    "product_name_jp": "product_name_jp",
    "description": "pack_size",
    "quantity": "quantity",
    "unit": "unit",
    "unit_price": "unit_price",
    "currency": "__currency__",
    "item_amount": "__formula__",
    "total_price": "__formula__",
    "amount": "__formula__",
}


def build_zone_config(
    file_bytes: bytes,
    field_positions: dict[str, str],
    product_table_config: dict[str, Any],
    cell_map: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """Build complete zone config from AI analysis + openpyxl structural scan.

    Supports two template layouts:
    1. Header + product table + summary (Japanese-style): has field_positions, SUM formulas
    2. Flat product table only (Korean-style): no header fields, no summary zone

    Args:
        file_bytes: Template Excel file content
        field_positions: AI-derived {field_key: cell_ref} mapping (can be empty)
        product_table_config: AI-derived product table structure
        cell_map: Optional full cell classification from AI

    Returns:
        Complete zone_config dict ready for template_engine.fill_template()
    """
    wb = load_workbook(io.BytesIO(file_bytes), data_only=False)
    ws = wb.active

    config: dict[str, Any] = {}

    # ── 1. Product zone from AI analysis ─────────────────────────
    prod_start = product_table_config.get("start_row")
    if not prod_start:
        raise ValueError("product_table_config missing start_row")

    prod_columns = product_table_config.get("columns", {})
    formula_details = product_table_config.get("formula_column_details", {})
    formula_cols = product_table_config.get("formula_columns", [])

    # ── 2. Find product zone end + summary zone via SUM formula scan ─
    has_summary = False
    try:
        prod_end, summary_info = _scan_summary_zone(ws, prod_start, formula_cols)
        has_summary = True
    except ValueError:
        # No SUM formula found — flat table template (e.g., Korean-style)
        # Product zone extends to last non-empty row
        prod_end = prod_start
        for row in range(prod_start, (ws.max_row or prod_start) + 1):
            has_data = any(
                ws.cell(row=row, column=col).value is not None
                for col in range(1, (ws.max_column or 12) + 1)
            )
            if has_data:
                prod_end = row
        logger.info(
            "No summary zone found — flat table template, product rows %d-%d",
            prod_start, prod_end,
        )

    if has_summary:
        config["zones"] = {
            "product_data": {"start": prod_start, "end": prod_end},
            "summary": summary_info["zone"],
        }
    else:
        config["zones"] = {
            "product_data": {"start": prod_start, "end": prod_end},
            "summary": {"start": prod_end + 1, "end": prod_end + 1},
        }

    # ── 3. Summary formulas ──────────────────────────────────────
    if has_summary:
        config["summary_formulas"] = summary_info["formulas"]
        config["summary_static_values"] = summary_info["static_values"]
    else:
        config["summary_formulas"] = []
        config["summary_static_values"] = {}

    # ── 4. Stale columns in summary rows ─────────────────────────
    if has_summary:
        config["stale_columns_in_summary"] = _detect_stale_columns(
            ws, summary_info["zone"]["start"], summary_info["zone"]["end"],
            summary_info["formula_positions"],
        )

        # ── 4b. Filter summary_static_values to exclude stale product data ─
        _row_counts: dict[int, int] = defaultdict(int)
        for _ref in summary_info["static_values"]:
            _row = int("".join(c for c in _ref if c.isdigit()))
            _row_counts[_row] += 1

        config["summary_static_values"] = {
            ref: val
            for ref, val in summary_info["static_values"].items()
            if _row_counts[int("".join(c for c in ref if c.isdigit()))] <= 3
        }
    else:
        config["stale_columns_in_summary"] = []

    # ── 5. Product row formulas ──────────────────────────────────
    prod_row_formulas = {}
    for col_letter, pattern in formula_details.items():
        # AI gives patterns like "=H*J" — convert to "=H{row}*J{row}"
        prod_row_formulas[col_letter] = _normalize_formula_pattern(pattern)

    # Fallback: scan template's first product row for formulas if AI didn't provide
    if not prod_row_formulas and prod_start:
        for col in range(1, (ws.max_column or 12) + 1):
            cell = ws.cell(row=prod_start, column=col)
            if isinstance(cell, MergedCell):
                continue
            val = cell.value
            if val and isinstance(val, str) and val.startswith("="):
                col_letter = get_column_letter(col)
                # Convert concrete formula (e.g., =H22*J22) to template pattern
                pattern = re.sub(r"(\d+)", "{row}", val)
                prod_row_formulas[col_letter] = pattern

    config["product_row_formulas"] = prod_row_formulas

    # ── 6. Product columns ───────────────────────────────────────
    config["product_columns"] = {
        col: field for col, field in prod_columns.items()
        if field not in ("item_amount", "total_price", "amount")  # skip formula columns
    }

    # ── 7. Header fields with data paths ─────────────────────────
    config["header_fields"] = _build_header_fields(field_positions, cell_map)

    # ── 7b. Field schema (schema-driven replacement for header_fields) ──
    from services.data.field_schema import build_field_schema
    config["field_schema"] = build_field_schema(field_positions, cell_map)

    # ── 8. Cross-zone references ─────────────────────────────────
    summ_end = config["zones"]["summary"]["end"]
    config["external_refs"] = _scan_external_refs(
        ws, prod_start, summ_end, cell_map,
    )

    summ_start = config["zones"]["summary"]["start"]
    logger.info(
        "Zone config built: product=%d-%d, summary=%d-%d, "
        "%d header fields, %d field_schema entries, %d external refs, %d summary formulas",
        prod_start, prod_end,
        summ_start, summ_end,
        len(config["header_fields"]),
        len(config["field_schema"]),
        len(config["external_refs"]),
        len(config["summary_formulas"]),
    )

    return config


# ── Internal helpers ─────────────────────────────────────────────


def _scan_summary_zone(
    ws, prod_start: int, formula_cols: list[str],
) -> tuple[int, dict]:
    """Scan for SUM formula to find product zone end and summary structure.

    Returns: (product_zone_end_row, summary_info_dict)
    """
    max_row = ws.max_row or prod_start + 200

    # Strategy: scan rows starting from prod_start looking for SUM formula
    sum_row = None
    sum_col = None
    sum_range_end = None

    for row in range(prod_start, max_row + 1):
        for col in range(1, (ws.max_column or 12) + 1):
            cell = ws.cell(row=row, column=col)
            if isinstance(cell, MergedCell):
                continue
            val = cell.value
            if val and isinstance(val, str):
                # Match SUM(X##:X##) pattern
                m = re.match(r"=SUM\(([A-Z]+)(\d+):([A-Z]+)(\d+)\)", val, re.IGNORECASE)
                if m:
                    sum_col = get_column_letter(col)
                    range_start = int(m.group(2))
                    sum_range_end = int(m.group(4))
                    # Verify the SUM range starts at product start
                    if range_start == prod_start:
                        sum_row = row
                        break
        if sum_row:
            break

    if not sum_row:
        raise ValueError(
            f"Could not find SUM formula referencing product start row {prod_start}. "
            "Template may not have a summary section."
        )

    product_zone_end = sum_range_end  # SUM(L22:L121) → product ends at 121

    # Scan summary rows: from SUM row, look for chained formulas
    summary_formulas = []
    static_values: dict[str, str] = {}
    formula_positions: set[str] = set()
    summary_end = sum_row

    for row_offset in range(0, 10):  # scan up to 10 rows
        row = sum_row + row_offset
        if row > max_row:
            break

        has_formula = False
        for col in range(1, (ws.max_column or 12) + 1):
            cell = ws.cell(row=row, column=col)
            if isinstance(cell, MergedCell):
                continue
            val = cell.value
            if val is None:
                continue

            col_letter = get_column_letter(col)
            cell_ref = f"{col_letter}{row}"

            if isinstance(val, str) and val.startswith("="):
                has_formula = True
                formula_positions.add(cell_ref)

                if row == sum_row and "SUM" in val.upper():
                    summary_formulas.append({
                        "cell": cell_ref,
                        "type": "product_sum",
                        "col": col_letter,
                        "label": "",
                    })
                else:
                    # Parse relative formula pattern
                    pattern = _extract_formula_pattern(val, sum_row, sum_col)
                    summary_formulas.append({
                        "cell": cell_ref,
                        "type": "relative",
                        "formula_template": pattern,
                        "label": "",
                    })

                    # Detect tax vs grand total by formula content
                    if "*0.08" in val or "*0.1" in val:
                        summary_formulas[-1]["label"] = "tax"
                    elif "TOTAL" in str(ws.cell(row=row, column=9).value or "").upper():
                        summary_formulas[-1]["label"] = "GRAND TOTAL"

            elif isinstance(val, str) and val.strip():
                # Static text in summary row (e.g. "GRAND TOTAL", "JPY")
                static_values[cell_ref] = val.strip()

        if has_formula:
            summary_end = row
        elif row_offset > 0:
            # No formula in this row and we're past the SUM row — summary zone ended
            break

    # Label enrichment from static values
    for sf in summary_formulas:
        row_num = int("".join(c for c in sf["cell"] if c.isdigit()))
        for sv_ref, sv_val in static_values.items():
            sv_row = int("".join(c for c in sv_ref if c.isdigit()))
            if sv_row == row_num and "TOTAL" in sv_val.upper():
                sf["label"] = sv_val

    return product_zone_end, {
        "zone": {"start": sum_row, "end": summary_end},
        "formulas": summary_formulas,
        "static_values": static_values,
        "formula_positions": formula_positions,
    }


def _extract_formula_pattern(
    formula: str, sum_row: int, sum_col: str | None,
) -> str:
    """Convert a concrete formula like '=L122*0.08' into a template pattern.

    Replaces cell references to summary rows with placeholders:
    - SUM row reference → {sum_cell}
    - SUM+1 (tax) row reference → {tax_cell}
    """
    pattern = formula

    # Replace references to the SUM row
    if sum_col:
        sum_ref = f"{sum_col}{sum_row}"
        if sum_ref in pattern:
            pattern = pattern.replace(sum_ref, "{sum_cell}")

        # Replace reference to tax row (sum_row + 1)
        tax_ref = f"{sum_col}{sum_row + 1}"
        if tax_ref in pattern:
            pattern = pattern.replace(tax_ref, "{tax_cell}")

    return pattern


def _detect_stale_columns(
    ws, summ_start: int, summ_end: int,
    formula_positions: set[str],
) -> list[str]:
    """Find columns in summary rows that have non-formula data (stale product data)."""
    stale_cols: set[str] = set()

    for row in range(summ_start, summ_end + 1):
        for col in range(1, (ws.max_column or 12) + 1):
            cell = ws.cell(row=row, column=col)
            if isinstance(cell, MergedCell):
                continue
            val = cell.value
            if val is None:
                continue

            col_letter = get_column_letter(col)
            cell_ref = f"{col_letter}{row}"

            # If this cell has a formula or is a known static value, skip
            if cell_ref in formula_positions:
                continue
            if isinstance(val, str) and val.startswith("="):
                continue

            # This cell has data but no formula — it's stale
            stale_cols.add(col_letter)

    return sorted(stale_cols)


def _normalize_formula_pattern(ai_pattern: str) -> str:
    """Convert AI formula pattern like '=H*J' to engine format '=H{row}*J{row}'.

    Also handles patterns already in correct format.
    """
    if "{row}" in ai_pattern:
        return ai_pattern  # Already in correct format

    # AI gives "=H*J" — insert {row} after each column letter
    result = ""
    i = 0
    s = ai_pattern
    while i < len(s):
        if s[i].isalpha() and s[i].isupper():
            # Collect full column letters
            col = ""
            while i < len(s) and s[i].isalpha() and s[i].isupper():
                col += s[i]
                i += 1
            # Skip if followed by digits (already has row number)
            if i < len(s) and s[i].isdigit():
                result += col
            else:
                result += col + "{row}"
        else:
            result += s[i]
            i += 1

    return result


def _build_header_fields(
    field_positions: dict[str, str],
    cell_map: dict[str, Any] | None,
) -> dict[str, str]:
    """Build header_fields: {cell_ref: data_path} from field_positions.

    Uses FIELD_DATA_PATHS lookup to map field_key → order_data path.
    """
    header_fields: dict[str, str] = {}

    for field_key, cell_ref in field_positions.items():
        data_path = FIELD_DATA_PATHS.get(field_key)
        if data_path:
            header_fields[cell_ref] = data_path

    # Also check cell_map for formula cells that reference summary
    if cell_map:
        for pos, info in cell_map.items():
            if info.get("source_type") == "formula" and info.get("field_key"):
                # Formula cells like total_amount are handled by external_refs
                pass

    return header_fields


def _scan_external_refs(
    ws,
    prod_start: int,
    summ_end: int,
    cell_map: dict[str, Any] | None,
) -> list[dict[str, str]]:
    """Find cells OUTSIDE product/summary zones that reference cells INSIDE.

    These formulas need updating when the product zone is resized.
    """
    external_refs: list[dict[str, str]] = []

    for row in range(1, (ws.max_row or 200) + 1):
        # Skip cells inside the managed zones
        if prod_start <= row <= summ_end:
            continue

        for col in range(1, (ws.max_column or 12) + 1):
            cell = ws.cell(row=row, column=col)
            if isinstance(cell, MergedCell):
                continue
            val = cell.value
            if not val or not isinstance(val, str) or not val.startswith("="):
                continue

            # Check if formula references any row in managed zone
            refs = re.findall(r"([A-Z]+)(\d+)", val)
            for ref_col, ref_row_str in refs:
                ref_row = int(ref_row_str)
                if prod_start <= ref_row <= summ_end:
                    cell_ref = f"{get_column_letter(col)}{row}"

                    # Determine what this references
                    # Build a template pattern based on what it references
                    formula_template = val
                    # Replace the concrete reference with a placeholder
                    target_ref = f"{ref_col}{ref_row}"

                    # Determine the placeholder name
                    placeholder = _classify_summary_ref(ref_row, summ_end - 2, summ_end)
                    if placeholder:
                        formula_template = formula_template.replace(
                            target_ref, f"{{{placeholder}}}"
                        )

                    external_refs.append({
                        "cell": cell_ref,
                        "formula_template": formula_template,
                    })
                    break  # One entry per cell

    return external_refs


def _classify_summary_ref(ref_row: int, summ_start: int, summ_end: int) -> str | None:
    """Classify a row reference as sum_cell, tax_cell, or grand_total_cell."""
    offset = ref_row - summ_start
    if offset == 0:
        return "sum_cell"
    elif offset == 1:
        return "tax_cell"
    elif offset == 2:
        return "grand_total_cell"
    return None
