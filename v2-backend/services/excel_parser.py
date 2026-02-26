"""Excel file parsing utilities for the Settings Center."""

import hashlib
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def parse_excel_file(file_bytes: bytes) -> dict:
    """Parse an Excel file and return sheet info, headers, and sample rows.

    Returns:
        {
            "sheets": [
                {
                    "name": "Sheet1",
                    "headers": ["Col A", "Col B", ...],
                    "header_row": 1,
                    "sample_rows": [[val, val, ...], ...],
                    "total_rows": 100,
                    "fingerprint": "sha256..."
                }
            ]
        }
    """
    wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    sheets = []

    for ws in wb.worksheets:
        header_row_idx = detect_header_row(ws)
        headers = []
        header_columns = []

        for cell in ws[header_row_idx]:
            val = cell.value
            col_letter = get_column_letter(cell.column)
            if val is not None:
                headers.append({"column": col_letter, "label": str(val).strip()})
                header_columns.append(cell.column)

        # Collect up to 5 sample rows after the header
        sample_rows = []
        data_start = header_row_idx + 1
        for row in ws.iter_rows(min_row=data_start, max_row=min(data_start + 4, ws.max_row or data_start)):
            row_data = []
            for cell in row:
                v = cell.value
                row_data.append(str(v) if v is not None else "")
            sample_rows.append(row_data)

        fingerprint = compute_fingerprint([h["label"] for h in headers])

        sheets.append({
            "name": ws.title,
            "headers": headers,
            "header_row": header_row_idx,
            "data_start_row": data_start,
            "sample_rows": sample_rows,
            "total_rows": ws.max_row or 0,
            "fingerprint": fingerprint,
        })

    wb.close()
    return {"sheets": sheets}


def parse_excel_cell_positions(file_bytes: bytes) -> dict:
    """Parse an Excel file and return all non-empty cell positions.

    Used for supplier template mapping — shows where each value lives in the template.

    Returns:
        {
            "sheets": [
                {
                    "name": "Sheet1",
                    "cells": [
                        {"position": "A1", "value": "PO Number", "row": 1, "col": "A"},
                        ...
                    ]
                }
            ]
        }
    """
    wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    sheets = []

    for ws in wb.worksheets:
        cells = []
        for row in ws.iter_rows(max_row=min(ws.max_row or 50, 50)):
            for cell in row:
                if cell.value is not None:
                    col_letter = get_column_letter(cell.column)
                    cells.append({
                        "position": f"{col_letter}{cell.row}",
                        "value": str(cell.value).strip(),
                        "row": cell.row,
                        "col": col_letter,
                    })
        sheets.append({"name": ws.title, "cells": cells})

    wb.close()
    return {"sheets": sheets}


def detect_header_row(ws) -> int:
    """Heuristically detect which row is the header row.

    Strategy: find the first row with >= 3 non-empty cells that contain text.
    Falls back to row 1.
    """
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row or 20, 20)):
        text_count = 0
        for cell in row:
            if cell.value is not None and isinstance(cell.value, str) and cell.value.strip():
                text_count += 1
        if text_count >= 3:
            return row[0].row
    return 1


def compute_fingerprint(header_labels: list[str]) -> str:
    """Compute a SHA256 fingerprint from sorted header labels."""
    normalized = "|".join(sorted(h.lower().strip() for h in header_labels if h))
    return hashlib.sha256(normalized.encode()).hexdigest()[:16]
