"""
Template Matcher — match uploaded orders to known OrderFormatTemplates.

Functions:
- find_matching_template: keyword-based matching (0 LLM)
- get_scannable_text: extract searchable text from file (0 LLM for Excel, pdfplumber for PDF)
- build_guided_prompt: construct template-guided extraction prompt
- extract_excel_deterministic: 0 LLM extraction when column_mapping is complete
"""

from __future__ import annotations

import io
import logging
from typing import Optional

from models import OrderFormatTemplate

logger = logging.getLogger(__name__)


# ─── Template Matching (0 LLM) ──────────────────────────────────

def find_matching_template(scannable_text: str, db) -> tuple[Optional[OrderFormatTemplate], Optional[str]]:
    """Find the best matching template by keyword search. Returns (template, method) or (None, None)."""
    templates = db.query(OrderFormatTemplate).filter(
        OrderFormatTemplate.is_active == True,
        OrderFormatTemplate.match_keywords.isnot(None),
    ).all()

    if not templates:
        return None, None

    text_upper = scannable_text.upper()
    best, best_hits = None, 0

    for tpl in templates:
        keywords = tpl.match_keywords or []
        if not keywords:
            continue
        hits = sum(1 for kw in keywords if kw.upper() in text_upper)
        if hits > best_hits:
            best, best_hits = tpl, hits

    if best:
        logger.info("Template matched: '%s' (id=%d) with %d keyword hits", best.name, best.id, best_hits)
        return best, "keyword"

    return None, None


# ─── Scannable Text Extraction ──────────────────────────────────

def get_scannable_text(file_bytes: bytes, file_type: str) -> str:
    """Extract searchable text from file for keyword matching. No LLM calls."""
    if file_type != "pdf":
        return _excel_to_scannable(file_bytes)
    return _pdf_to_scannable(file_bytes)


def _excel_to_scannable(file_bytes: bytes) -> str:
    """Read all cell values from Excel and join as text."""
    from openpyxl import load_workbook

    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
        parts = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) for c in row if c is not None]
                if cells:
                    parts.append(" ".join(cells))
        return "\n".join(parts)
    except Exception as e:
        logger.warning("Failed to extract Excel text: %s", e)
        return ""


def _pdf_to_scannable(file_bytes: bytes) -> str:
    """Extract text from PDF using pdfplumber (no LLM). Falls back gracefully."""
    try:
        import pdfplumber
        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            parts = []
            # Only scan first 3 pages for keywords (sufficient for header matching)
            for page in pdf.pages[:3]:
                text = page.extract_text()
                if text:
                    parts.append(text)
            return "\n".join(parts)
    except ImportError:
        logger.warning("pdfplumber not installed, skipping PDF template matching")
        return ""
    except Exception as e:
        logger.warning("Failed to extract PDF text: %s", e)
        return ""


# ─── Guided Prompt Construction ─────────────────────────────────

def build_guided_prompt(template: OrderFormatTemplate, field_definitions=None) -> str:
    """Build a template-guided extraction prompt with known layout/column info."""
    parts = ["分析这个采购订单文档，提取所有可见信息。\n"]

    if template.source_company:
        parts.append(f"## 已知信息\n来源公司: {template.source_company}\n")

    if template.layout_prompt:
        parts.append(f"## 文档布局\n{template.layout_prompt}\n")

    if template.column_mapping:
        parts.append("## 列映射（已知各列含义）")
        for col, field_key in template.column_mapping.items():
            parts.append(f"- 列 {col} = {field_key}")
        parts.append("")

    if template.extracted_fields:
        parts.append("## 已知元数据字段")
        for f in template.extracted_fields:
            parts.append(f"- {f.get('label', '')}: 字段名 {f.get('key', '')}")
        parts.append("")

    if field_definitions:
        parts.append("## 标准字段名（请使用这些字段名）")
        for fd in field_definitions:
            hint = f" ({fd.extraction_hint})" if fd.extraction_hint else ""
            parts.append(f"- {fd.field_key}: {fd.field_label}{hint}")
        parts.append("")

    # Append standard metadata schema
    from services.order_processor import ORDER_METADATA_SCHEMA

    parts.append("## 元数据字段（必须使用以下确切键名）")
    for k, v in ORDER_METADATA_SCHEMA.items():
        parts.append(f"- {k}: {v}")
    parts.append("- extra_fields: 其他可见元数据字段")
    parts.append("vendor_name 必须是纯字符串，不是对象。日期格式必须为 YYYY-MM-DD。total_amount 必须是数字。看不到的字段用 null。")
    parts.append("")

    parts.append("""返回 JSON：
{
  "order_metadata": {
    "po_number": "...",
    "ship_name": "...",
    "vendor_name": "...",
    "delivery_date": "YYYY-MM-DD",
    "order_date": "YYYY-MM-DD",
    "currency": "...",
    "destination_port": "...",
    "total_amount": 数字,
    "extra_fields": {}
  },
  "products": [
    { "line_number": N, "product_code": "...", "product_name": "...",
      "quantity": N, "unit": "...", "unit_price": N, "total_price": N }
  ]
}
只提取文档中可见的信息，不要编造。数字使用数值类型。返回纯 JSON，不要 markdown 代码块。""")

    return "\n".join(parts)


# ─── Deterministic Excel Extraction (0 LLM) ────────────────────

def extract_excel_deterministic(file_bytes: bytes, template: OrderFormatTemplate) -> dict:
    """Extract from Excel using template column_mapping — 0 LLM calls."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb[wb.sheetnames[0]]

    col_map = template.column_mapping  # {"A": "product_name", "B": "quantity", ...}
    if not col_map:
        raise ValueError("Template has no column_mapping")

    # Extract products from data rows
    products = []
    for row_idx in range(template.data_start_row, ws.max_row + 1):
        product = {}
        has_data = False
        for col_letter, field_key in col_map.items():
            cell_ref = f"{col_letter}{row_idx}"
            val = ws[cell_ref].value
            if val is not None:
                has_data = True
            product[field_key] = val
        if has_data and (product.get("product_name") or product.get("product_code")):
            product["line_number"] = row_idx - template.data_start_row + 1
            products.append(product)

    # Extract header metadata from rows above data
    metadata = _extract_header_metadata(ws, template)

    return {
        "order_metadata": metadata,
        "products": products,
        "extraction_method": "template_deterministic",
        "template_id": template.id,
    }


def _extract_header_metadata(ws, template: OrderFormatTemplate) -> dict:
    """Extract metadata from header area (rows 1 to header_row-1).

    Uses extracted_fields if available, otherwise scans for common patterns.
    """
    metadata = {}

    scan_max_row = max(template.data_start_row - 1, 1)

    # If template has extracted_fields with positions, use them
    if template.extracted_fields:
        for field in template.extracted_fields:
            key = field.get("key", "")
            label = field.get("label", "")
            if not key:
                continue
            # Try to find value in header area by scanning cells
            # extracted_fields typically has label/key but not exact positions,
            # so scan the header rows for the label text and grab the adjacent cell
            found = False
            for row in ws.iter_rows(min_row=1, max_row=scan_max_row, values_only=False):
                for cell in row:
                    if cell.value and label and str(cell.value).strip() == label.strip():
                        # Value is likely in the next cell to the right
                        next_col = cell.column + 1
                        if next_col <= ws.max_column:
                            val_cell = ws.cell(row=cell.row, column=next_col)
                            if val_cell.value is not None:
                                metadata[key] = val_cell.value
                                found = True
                                break
                if found:
                    break

    # Also scan for common metadata patterns in first few rows
    for row in ws.iter_rows(min_row=1, max_row=scan_max_row, values_only=False):
        for cell in row:
            val = str(cell.value).strip() if cell.value else ""
            val_lower = val.lower()
            # Common label → key mappings
            label_map = {
                "po number": "po_number", "po no": "po_number", "purchase order": "po_number",
                "delivery date": "delivery_date", "deliver on": "delivery_date",
                "ship name": "ship_name", "vessel": "ship_name",
                "currency": "currency",
                "destination": "destination_port",
            }
            for pattern, meta_key in label_map.items():
                if pattern in val_lower and meta_key not in metadata:
                    next_col = cell.column + 1
                    if next_col <= ws.max_column:
                        val_cell = ws.cell(row=cell.row, column=next_col)
                        if val_cell.value is not None:
                            metadata[meta_key] = val_cell.value

    return metadata
